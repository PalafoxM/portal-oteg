from typing import Any
from django import http
from django.shortcuts import render
from django.http import HttpRequest, JsonResponse, HttpResponseRedirect
from django.urls import reverse_lazy, reverse
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from back.models import *
from back.forms import *
from web.models import *
from django.shortcuts import render
from django.http import HttpResponse
import csv
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_GET
from django.views.decorators.http import require_POST
from django.shortcuts import get_object_or_404
#serializers
# render to string
from django.template.loader import render_to_string
from django.utils.decorators import method_decorator
# para archivo excel
from django.views import View
from django.contrib import messages
from openpyxl import load_workbook
import csv
import os
from datetime import datetime
from django.urls import reverse
import openpyxl
from django.http import HttpResponse
import json
from config.diccionarios import clean_str_col, homologar_columna_categoria, homologar_columna_destino

def is_ajax(request):
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'

class FuenteInfoEntornoNacional (ListView):
    model = FuenteInfoEntornoN 
    form_class = FuenteInfoEntornoNForm
    template_name = 'back/fuente_info_entorno_nacional/viewer.html'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs) :
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'search':
                data = []
                for i in FuenteInfoEntornoN.objects.all():
                    data.append(i.toJSON())
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Listado de Fuentes de Informacion Nacional'
        context['create_url'] = reverse_lazy('dashboard:fuente_info_entorno_nacional_create')
        context['entity'] = 'Nacional'
        context['is_fuente'] = True
        context['carga_masiva_url'] = reverse_lazy('dashboard:fuente_info_entorno_nacional_carga_masiva')
        return context  
    
class FuenteInfoEntornoNacionalCreate(CreateView):
    model = FuenteInfoEntornoN
    form_class = FuenteInfoEntornoNForm
    template_name = 'back/fuente_info_entorno_nacional/create.html'
    success_url = reverse_lazy('dashboard:fuente_info_entorno_nacional')

    def get_object(self, **kwargs):
        queryset = self.get_queryset()
        try:
            return queryset.get(**kwargs)
        except queryset.model.DoesNotExist:
            return None

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        replace_data = request.POST.get('replace_data')
        if form.is_valid():
            # Check if there is already a record with the same fecha, destino and categoria

            fecha = form.cleaned_data['fecha']
            entidad = form.cleaned_data['entidad']

            try:
                existing_object = self.get_object(fecha=fecha, entidad=entidad)

            except Certificacion.DoesNotExist:
                existing_object = None

            existing_catalogo = CatalogoEntidad.objects.filter(entidad__iexact= entidad).exists()
            # ALTER TABLE mytable MODIFY mycolumn VARCHAR(255) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci;

            # If there is no existing data, save the new data
            if not existing_catalogo:

                if not existing_catalogo:
                    data = {
                        'success': False,
                        'missingData': True,
                        'destino': entidad,
                        'message': 'No existe el la entidad en el catalogo',
                    }
                    return JsonResponse(data)

            # If there is existing data and replace_data is True, delete the existing data
            if existing_object and request.POST.get('replace_data') == 'on':
                # existing_object.delete()
                # Save the new data
                # self.object = form.save()

                for field in form.cleaned_data:
                    if field == 'replace_data':
                        continue
                    setattr(existing_object, field, form.cleaned_data[field])

                existing_object.save()

                data = {
                    'success': True,
                    'message': 'Data created successfully.',
                    'url': self.success_url,

                }
                return JsonResponse(data)

            # If there is existing data and replace_data is False, return an error

            if existing_object:
                data = FuenteInfoEntornoN.objects.filter(fecha=fecha, entidad=entidad)

                data_list = list(data.values('entidad', 'fecha', 'cuartos_disponibles_promedio', 'cuartos_disponibles','cuartos_ocupados','cuartos_ocupados_nacionales','cuartos_ocupados_extranjeros' ,'cuartos_ocupados_sin_clasificar',
    'llegada_de_turistas',
    'llegada_de_turistas_nacionales',
    'llegada_de_turistas_extranjeros',
    'turistas_noche',
    'turistas_noche_nacionales',
    'turistas_noche_extranjeros',
    'porcentaje_de_ocupacion',
    'porcentaje_de_ocupacion_nacionales',
    'porcentaje_de_ocupacion_extranjeros',
    'porcentaje_de_ocupacion_sin_clasificar',
    'densidad',
    'densidad_nacionales',
    'densidad_extranjeros',
    'estadia_promedio',
    'estadia_promedio_nacionales',
    'estadia_promedio_extranjeros' 
    ))
                data_list2 = list(form.cleaned_data.values())
                table_html = render_to_string('back/fuente_info_entorno_nacional/table.html',{'data_list': data_list, 'actual': True, 'data_list2': data_list2})

                datajsn = {
                    'success': False,
                    'message': 'Hubo un error al crear registro.',
                    'errors': 'Ya existe un registro con la misma fecha, destino , origen y museo o zona arqueologica',
                    'existing_object': table_html
                }

                return JsonResponse(datajsn)
            else:
                self.object = form.save()
                data = {
                    'success': True,
                    'message': 'Data created successfully.',
                    'url': self.success_url
                }
                return JsonResponse(data)
        else:
            data = {
                'success': False,
                'message': 'Error creating data.',
                'errors': form.errors,
                'format_errors': True
            }
            return JsonResponse(data)

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Hubo un error al crear registro.',
            'errors': form.errors
        }
        return JsonResponse(data)

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Registro creado exitosamente.',
            'url': self.success_url
        }
        return JsonResponse(data)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Crear una fuente'
        context['entity'] = 'Glosario'
        context['list_url'] = reverse_lazy('dashboard:fuente_info_entorno_nacional')
        context['action'] = 'add'
        return context

class FuenteInfoEntornoNacionalUpdate (UpdateView):
    model =     FuenteInfoEntornoN
    form_class = FuenteInfoEntornoNForm
    template_name = 'back/fuente_info_entorno_nacional/create.html'
    success_url = reverse_lazy('dashboard:fuente_info_entorno_nacional')

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Hubo un error al subir los datos.',
            'errors': form.errors
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Fuente creada exitosamente.',
            'url': self.success_url
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['list_url'] = reverse_lazy('dashboard:fuente_info_entorno_nacional')
            # Set the widget for the 'destino' field to read-only text input
        context['form'].fields['fecha'].widget = forms.TextInput(attrs={'readonly': 'readonly'})
        context['form'].fields['entidad'].widget = forms.TextInput(attrs={'readonly': 'readonly'})
        context['title'] = 'Editar fuente'

        context['edit_msg'] = 'Los Campos Destino y Fecha no pueden ser editados' 

        return context
    
class FuenteInfoEntornoNacionalDelete (DeleteView):
    model = FuenteInfoEntornoN
    success_url = reverse_lazy('dashboard:fuente_info_entorno_nacional')
    
    def post(self, request: HttpRequest, *args: str, **kwargs: Any) -> HttpResponse:
        return super().post(request, *args, **kwargs)

class EntornoNacionalCargaMasivaView(View):
    form_class = CargaMasivaForm
    template_name = 'back/fuente_info_entorno_nacional/carga_masiva.html'
    success_url = reverse_lazy('dashboard:fuente_info_entorno_nacional')

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form, 'title': 'Carga Masiva de fuente_info entorno nacional'})

    def convert_to_serializable(self, obj):
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        raise TypeError(
            f'Object of type {obj.__class__.__name__} is not JSON serializable')

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        archivo = request.FILES.get('archivo', None)
        if archivo:
            extension = os.path.splitext(archivo.name)[1]
            if extension == '.xlsx':
                registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas = self.procesar_archivo_xlsx(
                    archivo)
            elif extension == '.csv':
                registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas = self.procesar_archivo_csv(
                    archivo)
            else:
                messages.error(
                    request, 'El archivo debe ser un archivo .xlsx o .csv')
                registros_incorrectos.append(
                    "El archivo debe ser un archivo .xlsx o .csv")
        else:
            messages.error(request, 'Debe seleccionar un archivo')
            registros_incorrectos.append("Debe seleccionar un archivo")

        if len(registros_incorrectos) > 0 or len(registros_existentes) > 0:
            messages.error(request, 'Hay errores de registros')
            datos_json = json.dumps(
                registros_incorrectos, default=self.convert_to_serializable)

            return render(request, self.template_name, {
                'form': form,
                'title': 'Carga Masiva de fuente_info entorno nacional',
                'registros_correctos': registros_correctos,
                'registros_incorrectos': registros_incorrectos,
                'registros_existentes': registros_existentes,
                'descargar_url': datos_json,
                'num_filas_procesadas': num_filas_procesadas,
            })

        else:
            return HttpResponseRedirect(reverse('dashboard:fuente_info_entorno_nacional'))

    def procesar_archivo_xlsx(self, archivo):
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        try:
            workbook = load_workbook(filename=archivo, read_only=True)
            worksheet = workbook.active
            filas = list(worksheet.rows)
            for i, row in enumerate(filas):
                if i == 0:
                    continue  # Ignorar la primera fila si es el encabezado
                num_filas_procesadas += 1

                # Limpieza de datos
                entidad = clean_str_col(row[0].value)

                fecha = row[1].value
                fecha_str = str(fecha)
                fecha_str = fecha_str.split()[0] if fecha_str else ''
                fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').strftime('%Y-%m-%d')

                cuartos_disponibles_promedio = row[2].value
                cuartos_disponibles = row[3].value
                cuartos_ocupados = row[4].value
                cuartos_ocupados_nacionales = row[5].value
                cuartos_ocupados_extranjeros = row[6].value
                cuartos_ocupados_sin_clasificar = row[7].value
                llegada_de_turistas = row[8].value
                llegada_de_turistas_nacionales = row[9].value
                llegada_de_turistas_extranjeros = row[10].value
                turistas_noche = row[11].value
                turistas_noche_nacionales = row[12].value
                turistas_noche_extranjeros = row[13].value
                porcentaje_de_ocupacion = row[14].value
                porcentaje_de_ocupacion_nacionales = row[15].value
                porcentaje_de_ocupacion_extranjeros = row[16].value
                porcentaje_de_ocupacion_sin_clasificar = row[17].value
                densidad = row[18].value
                densidad_nacionales = row[19].value
                densidad_extranjeros = row[20].value
                estadia_promedio = row[21].value
                estadia_promedio_nacionales = row[22].value
                estadia_promedio_extranjeros = row[23].value

                datos = {
                    'entidad': entidad,
                    'fecha': fecha_str,
                    'cuartos_disponibles_promedio': cuartos_disponibles_promedio,
                    'cuartos_disponibles': cuartos_disponibles,
                    'cuartos_ocupados': cuartos_ocupados,
                    'cuartos_ocupados_nacionales': cuartos_ocupados_nacionales,
                    'cuartos_ocupados_extranjeros': cuartos_ocupados_extranjeros,
                    'cuartos_ocupados_sin_clasificar': cuartos_ocupados_sin_clasificar,
                    'llegada_de_turistas': llegada_de_turistas,
                    'llegada_de_turistas_nacionales': llegada_de_turistas_nacionales,
                    'llegada_de_turistas_extranjeros': llegada_de_turistas_extranjeros,
                    'turistas_noche': turistas_noche,
                    'turistas_noche_nacionales': turistas_noche_nacionales,
                    'turistas_noche_extranjeros': turistas_noche_extranjeros,
                    'porcentaje_de_ocupacion': porcentaje_de_ocupacion,
                    'porcentaje_de_ocupacion_nacionales': porcentaje_de_ocupacion_nacionales,
                    'porcentaje_de_ocupacion_extranjeros': porcentaje_de_ocupacion_extranjeros,
                    'porcentaje_de_ocupacion_sin_clasificar': porcentaje_de_ocupacion_sin_clasificar,
                    'densidad': densidad,
                    'densidad_nacionales': densidad_nacionales,
                    'densidad_extranjeros': densidad_extranjeros,
                    'estadia_promedio': estadia_promedio,
                    'estadia_promedio_nacionales': estadia_promedio_nacionales,
                    'estadia_promedio_extranjeros': estadia_promedio_extranjeros,
                }


                try:
                    if entidad not in CatalagoDestino.objects.values_list('entidad', flat=True):
                        print(f"la entidad {entidad} no está en la tabla CatalagoDestinoAeropuerto")
                        registros_incorrectos.append(datos)
                        continue

                    # Buscar si la fila ya existe en la base de datos
                    existente = FuenteInfoEntornoN.objects.filter(
                        entidad=entidad,
                        fecha=fecha_obj,
                    )
                    if existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {row} ya existe en la base de datos")
                        registros_existentes.append(datos)
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        db = FuenteInfoEntornoN(
                            entidad=entidad,
                            fecha=fecha_obj,
                            cuartos_disponibles_promedio=cuartos_disponibles_promedio,
                            cuartos_disponibles=cuartos_disponibles,
                            cuartos_ocupados=cuartos_ocupados,
                            cuartos_ocupados_nacionales=cuartos_ocupados_nacionales,
                            cuartos_ocupados_extranjeros=cuartos_ocupados_extranjeros,
                            cuartos_ocupados_sin_clasificar=cuartos_ocupados_sin_clasificar,
                            llegada_de_turistas=llegada_de_turistas,
                            llegada_de_turistas_nacionales=llegada_de_turistas_nacionales,
                            llegada_de_turistas_extranjeros=llegada_de_turistas_extranjeros,
                            turistas_noche=turistas_noche,
                            turistas_noche_nacionales=turistas_noche_nacionales,
                            turistas_noche_extranjeros=turistas_noche_extranjeros,
                            porcentaje_de_ocupacion=porcentaje_de_ocupacion,
                            porcentaje_de_ocupacion_nacionales=porcentaje_de_ocupacion_nacionales,
                            porcentaje_de_ocupacion_extranjeros=porcentaje_de_ocupacion_extranjeros,
                            porcentaje_de_ocupacion_sin_clasificar=porcentaje_de_ocupacion_sin_clasificar,
                            densidad=densidad,
                            densidad_nacionales=densidad_nacionales,
                            densidad_extranjeros=densidad_extranjeros,
                            estadia_promedio=estadia_promedio,
                            estadia_promedio_nacionales=estadia_promedio_nacionales,
                            estadia_promedio_extranjeros=estadia_promedio_extranjeros
                        )
                        db.save()
                        registros_correctos.append(datos)
                except (ValueError, TypeError) as e:
                    print(f"Error al procesar la fila {datos}: {e}")
                    registros_incorrectos.append(datos)

        except FileNotFoundError:
            print(f"El archivo {archivo} no se pudo abrir")

        return registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas

    def procesar_archivo_csv(self, archivo):
        archivo = self.request.FILES['archivo']
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        try:
            datos = csv.DictReader(
                archivo.read().decode('latin-1').splitlines())
            # print(datos)
            for row in datos:
                num_filas_procesadas += 1

                # Limpieza de datos
                entidad = clean_str_col(row['entidad'])

                # entidad = row['entidad']
                fecha = row['fecha']
                cuartos_disponibles_promedio = row['cuartos_disponibles_promedio']
                cuartos_disponibles = row['cuartos_disponibles']
                cuartos_ocupados = row['cuartos_ocupados']
                cuartos_ocupados_nacionales = row['cuartos_ocupados_nacionales']
                cuartos_ocupados_extranjeros = row['cuartos_ocupados_extranjeros']
                cuartos_ocupados_sin_clasificar = row['cuartos_ocupados_sin_clasificar']
                llegada_de_turistas = row['llegada_de_turistas']
                llegada_de_turistas_nacionales = row['llegada_de_turistas_nacionales']
                llegada_de_turistas_extranjeros = row['llegada_de_turistas_extranjeros']
                turistas_noche = row['turistas_noche']
                turistas_noche_nacionales = row['turistas_noche_nacionales']
                turistas_noche_extranjeros = row['turistas_noche_extranjeros']
                porcentaje_de_ocupacion = row['porcentaje_de_ocupacion']
                porcentaje_de_ocupacion_nacionales = row['porcentaje_de_ocupacion_nacionales']
                porcentaje_de_ocupacion_extranjeros = row['porcentaje_de_ocupacion_extranjeros']
                porcentaje_de_ocupacion_sin_clasificar = row['porcentaje_de_ocupacion_sin_clasificar']
                densidad = row['densidad']
                densidad_nacionales = row['densidad_nacionales']
                densidad_extranjeros = row['densidad_extranjeros']
                estadia_promedio = row['estadia_promedio']
                estadia_promedio_nacionales = row['estadia_promedio_nacionales']
                estadia_promedio_extranjeros = row['estadia_promedio_extranjeros']

                datos = {
                    'entidad': entidad,
                    'fecha': fecha,
                    'cuartos_disponibles_promedio': cuartos_disponibles_promedio,
                    'cuartos_disponibles': cuartos_disponibles,
                    'cuartos_ocupados': cuartos_ocupados,
                    'cuartos_ocupados_nacionales': cuartos_ocupados_nacionales,
                    'cuartos_ocupados_extranjeros': cuartos_ocupados_extranjeros,
                    'cuartos_ocupados_sin_clasificar': cuartos_ocupados_sin_clasificar,
                    'llegada_de_turistas': llegada_de_turistas,
                    'llegada_de_turistas_nacionales': llegada_de_turistas_nacionales,
                    'llegada_de_turistas_extranjeros': llegada_de_turistas_extranjeros,
                    'turistas_noche': turistas_noche,
                    'turistas_noche_nacionales': turistas_noche_nacionales,
                    'turistas_noche_extranjeros': turistas_noche_extranjeros,
                    'porcentaje_de_ocupacion': porcentaje_de_ocupacion,
                    'porcentaje_de_ocupacion_nacionales': porcentaje_de_ocupacion_nacionales,
                    'porcentaje_de_ocupacion_extranjeros': porcentaje_de_ocupacion_extranjeros,
                    'porcentaje_de_ocupacion_sin_clasificar': porcentaje_de_ocupacion_sin_clasificar,
                    'densidad': densidad,
                    'densidad_nacionales': densidad_nacionales,
                    'densidad_extranjeros': densidad_extranjeros,
                    'estadia_promedio': estadia_promedio,
                    'estadia_promedio_nacionales': estadia_promedio_nacionales,
                    'estadia_promedio_extranjeros': estadia_promedio_extranjeros,
                }


                try:
                    if entidad not in CatalagoDestino.objects.values_list('entidad', flat=True):
                        print(f"la entidad {entidad} no está en la tabla CatalagoDestinoAeropuerto")
                        registros_incorrectos.append(datos)
                        continue

                    # Buscar si la fila ya existe en la base de datos
                    existente = FuenteInfoEntornoN.objects.filter(
                        entidad=entidad,
                        fecha=fecha,
                    )
                    if existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {row} ya existe en la base de datos")
                        registros_existentes.append(datos)
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        db = FuenteInfoEntornoN(
                            entidad=entidad,
                            fecha=fecha,
                            cuartos_disponibles_promedio=cuartos_disponibles_promedio,
                            cuartos_disponibles=cuartos_disponibles,
                            cuartos_ocupados=cuartos_ocupados,
                            cuartos_ocupados_nacionales=cuartos_ocupados_nacionales,
                            cuartos_ocupados_extranjeros=cuartos_ocupados_extranjeros,
                            cuartos_ocupados_sin_clasificar=cuartos_ocupados_sin_clasificar,
                            llegada_de_turistas=llegada_de_turistas,
                            llegada_de_turistas_nacionales=llegada_de_turistas_nacionales,
                            llegada_de_turistas_extranjeros=llegada_de_turistas_extranjeros,
                            turistas_noche=turistas_noche,
                            turistas_noche_nacionales=turistas_noche_nacionales,
                            turistas_noche_extranjeros=turistas_noche_extranjeros,
                            porcentaje_de_ocupacion=porcentaje_de_ocupacion,
                            porcentaje_de_ocupacion_nacionales=porcentaje_de_ocupacion_nacionales,
                            porcentaje_de_ocupacion_extranjeros=porcentaje_de_ocupacion_extranjeros,
                            porcentaje_de_ocupacion_sin_clasificar=porcentaje_de_ocupacion_sin_clasificar,
                            densidad=densidad,
                            densidad_nacionales=densidad_nacionales,
                            densidad_extranjeros=densidad_extranjeros,
                            estadia_promedio=estadia_promedio,
                            estadia_promedio_nacionales=estadia_promedio_nacionales,
                            estadia_promedio_extranjeros=estadia_promedio_extranjeros
                        )
                        db.save()
                        registros_correctos.append(datos)
                except (ValueError, TypeError) as e:
                    print(f"Error al procesar la fila {datos}: {e}")
                    registros_incorrectos.append(datos)
        except FileNotFoundError:
            print(f"No se encontró el archivo {archivo}")
        except Exception as e:
            print(f"Error al procesar el archivo {archivo}: {e}")
        return registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas


class EntornoNacionalDescargarArchivoView(View):

    def crear_archivo_excel(self, registros_incorrectos):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Obtener los nombres y verbose_name de los campos del modelo FuenteInfoEntornoN
        fields = FuenteInfoEntornoN._meta.get_fields()
        column_labels = [field.verbose_name for field in fields if field.name != 'id']
        column_names = [field.name for field in fields if field.name != 'id']

        # Escribir los encabezados de las columnas
        for i, campo in enumerate(column_labels):
            columna = i + 1
            worksheet.cell(row=1, column=columna, value=campo)

        # Obtener los datos del modelo FuenteInfoEntornoN
        datos = registros_incorrectos

        # Escribir los valores en las celdas correspondientes
        fila = 2
        for registro in datos:
            for i, campo in enumerate(column_names):
                if campo != 'id':  # Omitir la clave 'id'
                    columna = i + 1
                    valor = registro[campo]
                    worksheet.cell(row=fila, column=columna, value=valor)
            fila += 1

        # Set the column widths to auto-fit
        for column in worksheet.columns:
            max_length = 0
            column_name = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_name].width = adjusted_width

        # Create the response with the Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=gasto_derrama_registros_incorrectos.xls'

        # workbook.save(response)
        return workbook

    def post(self, request, *args, **kwargs):
        # Obtener los registros incorrectos del cuerpo de la petición
        registros_incorrectos = json.loads(request.body)

        # Crear y enviar el archivo de Excel con las filas incorrectas
        workbook = self.crear_archivo_excel(registros_incorrectos)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=gasto_derrama_registros_incorrectos.xlsx'
        workbook.save(response)
        return response