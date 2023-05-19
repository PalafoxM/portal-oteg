from django.shortcuts import render
from django.http import JsonResponse, HttpResponseRedirect
from django.urls import reverse_lazy, reverse
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from back.models import *
from back.forms import *
from web.models import *
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.template.loader import render_to_string
# para archivo excel
from django.views import View
from django.contrib import messages
from openpyxl import load_workbook
import csv
import os
import datetime
from django.urls import reverse
import openpyxl
from django.http import HttpResponse
import json
from config.diccionarios import clean_str_col, homologar_columna_categoria, homologar_columna_destino


def is_ajax(request):
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'


class FuenteInfoZonasArqueologicas (ListView):
    model = zonas_arqueologicas_museos
    template_name = 'back/fuente_info_zonas_arq/viewer.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Listado de Fuentes de Informacion Zonas Arqueologicas'
        context['create_url'] = reverse_lazy(
            'dashboard:fuente_info_zonas_arqueologicas_create')
        context['entity'] = 'Zonas Arqueologicas'
        context['is_fuente'] = True
        context['carga_masiva_url'] = reverse_lazy('dashboard:fuente_zonas_arqueologicas_carga_masiva')
        return context


class FuenteInfoZonasArqueologicasCreate (CreateView):
    model = zonas_arqueologicas_museos
    form_class = ZonasArqueologicasMuseosForm
    template_name = 'back/fuente_info_zonas_arq/create_update_fuentes_info.html'
    success_url = reverse_lazy('dashboard:fuente_info_zonas_arqueologicas')

    def get_object(self, **kwargs):
        queryset = self.get_queryset()
        try:
            return queryset.get(**kwargs)
        except queryset.model.DoesNotExist:
            return None

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        replace_data = request.POST.get('replace_data')
        if form.is_valid():
            # Check if there is already a record with the same fecha, destino and categoria

            fecha = form.cleaned_data['fecha']
            destino = form.cleaned_data['destino']
            origen = form.cleaned_data['origen_visitante']
            museo_zona_arqueologica_form = form.cleaned_data['museo_zona_arqueologica']

            try:
                existing_object = self.get_object(fecha=fecha, destino=destino, origen_visitante=origen, museo_zona_arqueologica=museo_zona_arqueologica_form)

            except DataTour.DoesNotExist:
                existing_object = None

            existing_catalogo = catalogo_destinos.objects.filter(
                destino=destino).exists()
            museo_zona_arqueologica = catalogo_zonaz_arq_museos.objects.filter(museo_zona_arqueologica=museo_zona_arqueologica_form).exists()

            # If there is no existing data, save the new data
            if not museo_zona_arqueologica or not existing_catalogo:

                if not museo_zona_arqueologica and not existing_catalogo:

                    data = {
                        'success': False,
                        'missingData': True,
                        'museo_Z_A': museo_zona_arqueologica_form,
                        'destino': destino,
                        'message': 'No existe la categoria o el destino en el catalogo',
                    }
                    return JsonResponse(data)

                if not museo_zona_arqueologica:
                    data = {
                        'success': False,
                        'missingData': True,
                        'museo_Z_A': museo_zona_arqueologica_form,
                        'message': 'No existe la Zona Arqueologica o Museo en el catalogo',
                    }
                    return JsonResponse(data)

                if not existing_catalogo:
                    data = {
                        'success': False,
                        'missingData': True,
                        'destino': destino,
                        'message': 'No existe el destino en el catalogo',
                    }
                    return JsonResponse(data)

            # If there is existing data and replace_data is True, delete the existing data
            if existing_object and request.POST.get('replace_data') == 'on':
                # existing_object.delete()
                # Save the new data
                # self.object = form.save()

                for field in form.cleaned_data:
                    if field == 'replace_data':
                        continue
                    setattr(existing_object, field, form.cleaned_data[field])
                existing_object.save()

                data = {
                    'success': True,
                    'message': 'Data created successfully.',
                    'url': self.success_url,

                }
                return JsonResponse(data)

            # If there is existing data and replace_data is False, return an error

            if existing_object:
                data = zonas_arqueologicas_museos.objects.filter(
                    fecha=fecha, destino=destino, origen_visitante=origen, museo_zona_arqueologica=museo_zona_arqueologica_form)

                data_list = list(data.values(
                    'fecha', 'destino', 'museo_zona_arqueologica', 'origen_visitante', 'visitantes', 'tipo'))
                data_list2 = list(form.cleaned_data.values())
                table_html = render_to_string('back/fuente_info_zonas_arq/table.html', {
                                              'data_list': data_list, 'actual': True, 'data_list2': data_list2})

                datajsn = {
                    'success': False,
                    'message': 'Hubo un error al crear registro.',
                    'errors': 'Ya existe un registro con la misma fecha, destino , origen y museo o zona arqueologica',
                    'existing_object': table_html
                }

                return JsonResponse(datajsn)
            else:
                self.object = form.save()
                data = {
                    'success': True,
                    'message': 'Data created successfully.',
                    'url': self.success_url
                }
                return JsonResponse(data)
        else:
            data = {
                'success': False,
                'message': 'Error creating data.',
                'errors': form.errors,
                'format_errors': True
            }
            return JsonResponse(data)

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Hubo un error al crear registro.',
            'errors': form.errors
        }
        return JsonResponse(data)

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Registro creado exitosamente.',
            'url': self.success_url
        }
        return JsonResponse(data)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Crear una fuente'
        context['entity'] = 'Glosario'
        context['list_url'] = reverse_lazy('dashboard:fuente_info_zonas_arqueologicas')
        context['action'] = 'add'
        return context

class FuenteInfoZonasArqueologicasUpdate (UpdateView):
    model = zonas_arqueologicas_museos
    form_class = ZonasArqueologicasMuseosForm
    template_name = 'back/fuente_info_zonas_arq/create_update_fuentes_info.html'
    success_url = reverse_lazy('dashboard:fuente_info_zonas_arqueologicas')

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Hubo un error al subir los datos.',
            'errors': form.errors
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Fuente creada exitosamente.',
            'url': self.success_url
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['list_url'] = reverse_lazy('dashboard:fuente_info_zonas_arqueologicas')
            # Set the widget for the 'destino' field to read-only text input
        context['form'].fields['destino'].widget = forms.TextInput(attrs={'readonly': 'readonly'})
        # Set the widget for the 'fecha' field to read-only text input
        context['form'].fields['museo_zona_arqueologica'].widget = forms.TextInput(attrs={'readonly': 'readonly'})
        # Set the widget for the 'categoria' field to read-only text input
        context['form'].fields['fecha'].widget = forms.TextInput(attrs={'readonly': 'readonly'})
        # Set the widget for the 'categoria' field to read-only text input
        context['form'].fields['origen_visitante'].widget = forms.TextInput(attrs={'readonly': 'readonly'})
        context['title'] = 'Editar fuente'
        context['edit_msg'] = 'Los Campos Destino, Fecha, Museo o Zona Arqueologica y Origen Visitante no pueden ser editados' 

        return context

class FuenteInfoZonasArqueologicasDelete (DeleteView):
    model = zonas_arqueologicas_museos
    success_url = reverse_lazy('dashboard:fuente_info_zonas_arqueologicas')

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        success_url = self.get_success_url()
        self.object.delete()
        return HttpResponseRedirect(success_url)

class ZonasArqueoCargaMasivaView(View):
    form_class = CargaMasivaForm
    template_name = 'back/fuente_info_zonas_arq/carga_masiva.html'
    success_url = reverse_lazy('dashboard:inversion_publica_list')

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form, 'title': 'Carga Masiva de Museos ZA'})


    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        archivo = request.FILES.get('archivo', None)
        if archivo:
            extension = os.path.splitext(archivo.name)[1]
            if extension == '.xlsx':
                registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas = self.procesar_archivo_xlsx(archivo)
            elif extension == '.csv':
                registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas = self.procesar_archivo_csv(archivo)
            else:
                messages.error(request, 'El archivo debe ser un archivo .xlsx o .csv')
                registros_incorrectos.append("El archivo debe ser un archivo .xlsx o .csv")
        else:
            messages.error(request, 'Debe seleccionar un archivo')
            registros_incorrectos.append("Debe seleccionar un archivo")

        if len(registros_incorrectos) > 0 or len(registros_existentes) > 0:
            messages.error(request, 'Hay errores de registros')
            datos_json = json.dumps(registros_incorrectos)
            
            return render(request, self.template_name, {
                'form': form,
                'title': 'Carga Masiva de Museos ZA',
                'registros_correctos': registros_correctos,
                'registros_incorrectos': registros_incorrectos,
                'registros_existentes': registros_existentes,
                'descargar_url': datos_json,
                'num_filas_procesadas': num_filas_procesadas,
            })
            
        else:
            return HttpResponseRedirect(reverse('dashboard:inversion_publica_list'))
        
        

    def procesar_archivo_xlsx(self, archivo):
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        try:
            workbook = load_workbook(filename=archivo, read_only=True)
            worksheet = workbook.active
            filas = list(worksheet.rows)
            for i, row in enumerate(filas):
                if i == 0:
                    continue # Ignorar la primera fila si es el encabezado
                num_filas_procesadas += 1
                # Limpieza de datos
                fecha_str = row[3].value.date().strftime('%Y-%m-%d') if len(row) > 3 and row[3].value else ''
                fecha_obj = datetime.datetime.strptime(fecha_str, '%Y-%m-%d').date() if fecha_str else None
                json_fecha = json.dumps(fecha_obj.strftime('%Y-%m-%d')) if fecha_obj else ''
                json_fecha = json_fecha.strip('"') if json_fecha else ''

                origen_visitante = row[4].value if len(row) > 4 else ''

                visitantes = row[5].value if len(row) > 5 else 0

                # Limpieza de datos
                destino = clean_str_col(row[0].value if len(row) > 0 else '')
                tipo = clean_str_col(row[1].value if len(row) > 1 else '')
                nombre = clean_str_col(row[2].value if len(row) > 1 else '')

                # Homologación de datos
                destino = homologar_columna_destino(destino)
                # tipo = homologar_columna_destino(tipo)
                # nombre = homologar_columna_destino(nombre)
                datos = {
                    "destino": destino,
                    "tipo": tipo,
                    "nombre": nombre,
                    "fecha": json_fecha,
                    "origen_visitante": origen_visitante,
                    "visitantes": visitantes
                }

                try:
                    # Validar los datos
                    # fecha_obj = datetime.datetime.strptime(fecha_str, '%d/%m/%Y').date()
                    visitantes_int = int(visitantes)

                    if destino not in CatalagoDestino.objects.values_list('destino', flat=True):
                        print(f"El destino {destino} no está en la tabla CatalagoDestino")
                        registros_incorrectos.append(datos)
                        continue
                    # Validar si el tipo_visitante es válido
                    if not CatalagoZAMuseos.objects.filter(tipo=tipo).exists():
                        print(f"El tipo_visitante: {tipo} no está en la tabla CatalagoTipoVisistante")
                        registros_incorrectos.append(datos)
                        continue
                    # Validar si el tipo_visitante es válido
                    if not CatalagoZAMuseos.objects.filter(nombre=nombre).exists():
                        print(f"El tipo_visitante: {nombre} no está en la tabla CatalagoTipoVisistante")
                        registros_incorrectos.append(datos)
                        continue

                    # Buscar si la fila ya existe en la base de datos
                    existente = zonas_arqueologicas_museos.objects.filter(
                        destino = destino, 
                        tipo = tipo,
                        fecha = fecha_obj, 
                        nombre = nombre)
                    if existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {row} ya existe en la base de datos")
                        registros_existentes.append(datos)
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        db = zonas_arqueologicas_museos(
                            destino = destino,
                            tipo = tipo,
                            nombre = nombre,
                            fecha = fecha_obj,
                            origen_visitante = origen_visitante,
                            visitantes = visitantes_int,
                        )
                        db.save()
                        registros_correctos.append(datos)
                except (ValueError, TypeError) as e:
                    # Si los datos no son válidos, se guarda el número de fila en la lista de registros incorrectos
                    registros_incorrectos.append(datos)
                    
        except FileNotFoundError:
                print(f"El archivo {archivo} no se pudo abrir")
                
        return registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas
    
    def procesar_archivo_csv(self, archivo):
        archivo = self.request.FILES['archivo']
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        try:
            datos = csv.DictReader(archivo.read().decode('latin-1').splitlines())
            # print(datos)
            for row in datos:
                num_filas_procesadas += 1

                fecha = row['fecha']
                fecha_str = str(fecha)
                fecha_str = fecha_str.split()[0] if fecha_str else ''  # Eliminar la parte de la hora si existe la fecha
                fecha_obj = datetime.datetime.strptime(fecha_str, '%Y-%m-%d').date()
                # Serializar la fecha en formato JSON
                json_fecha = json.dumps(fecha_obj.strftime('%Y-%m-%d'))
                json_fecha = str(json_fecha).strip('"')
                origen_visitante = row['origen_visitante']
                visitantes = row['visitantes']

                # Limpieza de datos
                destino = clean_str_col(row['destino'])
                tipo = clean_str_col(row['tipo'])
                nombre = clean_str_col(row['nombre'])

                # Homologación de datos
                destino = homologar_columna_destino(destino)
                # tipo = homologar_columna_destino(tipo)
                # nombre = homologar_columna_destino(nombre)
                datos = {
                    "destino": destino,
                    "tipo": tipo,
                    "nombre": nombre,
                    "fecha": json_fecha,
                    "origen_visitante": origen_visitante,
                    "visitantes": visitantes
                }

                try:
                    # Validar los datos
                    fecha_obj = datetime.datetime.strptime(fecha_str, '%d/%m/%Y').date()
                    visitantes_int = int(visitantes)

                    if destino not in CatalagoDestino.objects.values_list('destino', flat=True):
                        print(f"El destino {destino} no está en la tabla CatalagoDestino")
                        registros_incorrectos.append(row)
                        continue
                    # Validar si el tipo_visitante es válido
                    if not CatalagoZAMuseos.objects.filter(tipo=tipo).exists():
                        print(f"El tipo_visitante: {tipo} no está en la tabla CatalagoTipoVisistante")
                        registros_incorrectos.append(datos)
                        continue
                    # Validar si el tipo_visitante es válido
                    if not CatalagoZAMuseos.objects.filter(nombre=nombre).exists():
                        print(f"El tipo_visitante: {nombre} no está en la tabla CatalagoTipoVisistante")
                        registros_incorrectos.append(datos)
                        continue

                    # Buscar si la fila ya existe en la base de datos
                    existente = zonas_arqueologicas_museos.objects.filter(
                        destino = destino, 
                        tipo = tipo,
                        fecha = fecha_obj, 
                        nombre = nombre)
                    if existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {row} ya existe en la base de datos")
                        registros_existentes.append(row)
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        db = zonas_arqueologicas_museos(
                            destino = destino,
                            tipo = tipo,
                            nombre = nombre,
                            fecha = fecha_obj,
                            origen_visitante = origen_visitante,
                            visitantes = visitantes_int,
                        )
                        db.save()
                        registros_correctos.append(row)
                except (ValueError, TypeError) as e:
                    print(f"Error al procesar la fila {row}: {e}")
                    registros_incorrectos.append(row)
        except FileNotFoundError:
            print(f"No se encontró el archivo {archivo}")
        except Exception as e:
            print(f"Error al procesar el archivo {archivo}: {e}")
        return registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas


class ZonasArqueoDescargarArchivoView(View):

    def crear_archivo_excel(self, registros_incorrectos):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Escribir encabezados de columna
        worksheet['A1'] = 'destino'
        worksheet['B1'] = 'tipo'
        worksheet['C1'] = 'nombre'
        worksheet['D1'] = 'fecha'
        worksheet['E1'] = 'origen_visitante'
        worksheet['F1'] = 'visitantes'

        # Add the incorrect rows to the worksheet
        for i, row in enumerate(registros_incorrectos):
            fila = i + 2
            worksheet.cell(row=fila, column=1, value=row['destino'])
            worksheet.cell(row=fila, column=2, value=row['tipo'])
            worksheet.cell(row=fila, column=3, value=row['nombre'])
            worksheet.cell(row=fila, column=4, value=row['fecha'])
            worksheet.cell(row=fila, column=5, value=row['origen_visitante'])
            worksheet.cell(row=fila, column=6, value=row['visitantes'])



        # Set the column widths to auto-fit
        for column in worksheet.columns:
            max_length = 0
            column_name = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_name].width = adjusted_width

        # Create the response with the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=gasto_derrama_registros_incorrectos.xls'

        

        # workbook.save(response)
        return workbook
    
    def post(self, request, *args, **kwargs):
        # Obtener los registros incorrectos del cuerpo de la petición
        registros_incorrectos = json.loads(request.body)

        # Crear y enviar el archivo de Excel con las filas incorrectas
        workbook = self.crear_archivo_excel(registros_incorrectos)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=gasto_derrama_registros_incorrectos.xlsx'
        workbook.save(response)
        return response   