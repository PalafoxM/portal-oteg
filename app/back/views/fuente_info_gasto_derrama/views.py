from django.shortcuts import render
from django.http import JsonResponse, HttpResponseRedirect
from django.urls import reverse_lazy , reverse
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from back.models import  *
from back.forms import *
from web.models import *
from django.shortcuts import render

# para archivo excel
from django.views import View
from django.contrib import messages
from openpyxl import load_workbook
import csv
import os
import datetime
from django.urls import reverse
import openpyxl
from django.http import HttpResponse
import json
from config.diccionarios import clean_str_col, homologar_columna_categoria, homologar_columna_destino
from django.contrib.auth.decorators import login_required, permission_required
from django.utils.decorators import method_decorator
from django.http import Http404
from django.core.exceptions import PermissionDenied
from back.mixins import *
from django.contrib.auth.decorators import user_passes_test

def es_admin_o_superadmin(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)


def is_ajax(request):
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'




class FuenteInfoGastoDerrama (SuperAdminOrAdminMixin, LoginRequiredMixin, ListView):
    model = GastoDerrama
    template_name = 'back/fuente_info_gasto_derrama/viewer.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Listado de Fuentes de Informacion Gasto Derrama'
        context['create_url'] = reverse_lazy('dashboard:fuente_info_gasto_derrama_create')
        context['entity'] = 'Categorias'
        context['is_fuente']    = True
        context['carga_masiva_url'] = reverse_lazy('dashboard:fuente_gasto_derrama_carga_masiva')
        return context

    
class FuenteInfoGastoDerramaCreate (SuperAdminOrAdminMixin, LoginRequiredMixin, CreateView):
    model = GastoDerrama
    form_class = GastoDerramaForm
    template_name ='back/components/create_update.html'
    success_url = reverse_lazy('dashboard:fuente_info_gasto_derrama')

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        if form.is_valid():
            self.object = form.save()
            data = {
                'success': True,
                'message': 'fuente Creada exitosamente.',
                'url': self.success_url
            }
            return JsonResponse(data)
        else:
            data = {
                'success': False,
                'message': 'Hubo un error al crear registro.',
                'errors': form.errors
            }
            return JsonResponse(data)

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Hubo un error al crear registro.',
            'errors': form.errors
        }
        return JsonResponse(data)

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Registro creado exitosamente.',
            'url': self.success_url
        }
        return JsonResponse(data)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Crear una fuente'
        context['entity'] = 'Glosario'
        context['list_url'] = reverse_lazy('dashboard:fuente_info_gasto_derrama')
        context['action'] = 'add'
        return context
    


class FuenteInfoGastoDerramaUpdate (SuperAdminOrAdminMixin, LoginRequiredMixin, UpdateView):
    model = GastoDerrama
    form_class = GastoDerramaForm
    template_name = 'back/components/create_update.html'
    success_url = reverse_lazy('dashboard:fuente_info_gasto_derrama')

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Hubo un error al crear el evento.',
            'errors': form.errors
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Evento creado exitosamente.',
            'url': self.success_url
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(instance=self.object)
        context['list_url'] = reverse_lazy('dashboard:fuente_info_gasto_derrama')
        return context

    
class FuenteInfoGastoDerramaDelete(SuperAdminOrAdminMixin, LoginRequiredMixin, DeleteView):
    model = GastoDerrama
    success_url = reverse_lazy('dashboard:fuente_info_gasto_derrama')

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()

        try:
            self.object.delete()
            return JsonResponse({'message': 'Eliminación exitosa.'})
        except Exception as e:
            return JsonResponse({'error': 'Error al eliminar el registro.'}, status=500)

    
class GastoDerramaCargaMasivaView(SuperAdminOrAdminMixin, LoginRequiredMixin, View):
    form_class = CargaMasivaForm
    template_name = 'back/fuente_info_gasto_derrama/carga_masiva.html'
    success_url = reverse_lazy('dashboard:fuente_info_gasto_derrama')

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form, 'title': 'Carga Masiva Gasto Derrama'})


    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        archivo = request.FILES.get('archivo', None)
        if archivo:
            extension = os.path.splitext(archivo.name)[1]
            if extension == '.xlsx':
                registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas = self.procesar_archivo_xlsx(archivo)
            elif extension == '.csv':
                registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas = self.procesar_archivo_csv(archivo)
            else:
                messages.error(request, 'El archivo debe ser un archivo .xlsx o .csv')
                registros_incorrectos.append("El archivo debe ser un archivo .xlsx o .csv")
        else:
            messages.error(request, 'Debe seleccionar un archivo')
            registros_incorrectos.append("Debe seleccionar un archivo")

        if len(registros_incorrectos) > 0 or len(registros_existentes) > 0:
            messages.error(request, 'Hay errores de registros')
            datos_json = json.dumps(registros_incorrectos)
            
            return render(request, self.template_name, {
                'form': form,
                'title': 'Carga Masiva Gasto Derrama',
                'registros_correctos': registros_correctos,
                'registros_incorrectos': registros_incorrectos,
                'registros_existentes': registros_existentes,
                'descargar_url': datos_json,
                'num_filas_procesadas': num_filas_procesadas,
            })
            
        else:
            return HttpResponseRedirect(reverse('dashboard:fuente_info_gasto_derrama'))
        
        

    def procesar_archivo_xlsx(self, archivo):
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        try:
            workbook = load_workbook(filename=archivo, read_only=True)
            worksheet = workbook.active
            filas = list(worksheet.rows)
            for i, row in enumerate(filas):
                if i == 0:
                    continue # Ignorar la primera fila si es el encabezado
                num_filas_procesadas += 1
                # Limpieza de datos
                if len(row) >= 4:
                    destino = clean_str_col(row[3].value)
                else:
                    # Manejar el caso cuando la tupla no tiene suficientes elementos
                    destino = None  # o cualquier otro valor predeterminado o acción que desees realizar
                
                if len(row) >= 3:
                    tipo_visitante = clean_str_col(row[2].value)
                else:
                    # Manejar el caso cuando la tupla no tiene suficientes elementos
                    tipo_visitante = None  # o cualquier otro valor predeterminado o acción que desees realizar

                # Homologación de datos
                destino = homologar_columna_destino(destino)
                tipo_visitante = CatalagoTipoVisistante.homologar_tipo_visitante(tipo_visitante)
                
                gasto_diario_prom = row[0].value if len(row) > 0 else 0
                ano = row[1].value if len(row) > 1 else 0
                participacion_en_hospedaje = row[4].value if len(row) > 4 else 0
                estadia_promedio = row[5].value if len(row) > 5 else 0


                datos = {
                        'gasto_diario_prom': gasto_diario_prom,
                        'ano': ano,
                        'tipo_visitante': tipo_visitante,
                        'destino': destino,
                        'participacion_en_hospedaje': participacion_en_hospedaje,
                        'estadia_promedio': estadia_promedio
                    }
                try:
                    # Validar los datos
                    gasto_diario_prom_float = float(gasto_diario_prom)
                    ano_int = int(ano)
                    participacion_en_hospedaje_float = float(participacion_en_hospedaje)
                    estadia_promedio_float = float(estadia_promedio)
                    

                    # Validar si el destino es válido
                    if destino not in CatalagoDestino.objects.values_list('destino', flat=True):
                        print(f"El destino {destino} no está en la tabla CatalagoDestino")
                        registros_incorrectos.append(datos)
                        continue

                    # Validar si el tipo_visitante es válido
                    if not CatalagoTipoVisistante.objects.filter(tipo_visitante=tipo_visitante).exists():
                        print(f"El tipo_visitante: {tipo_visitante} no está en la tabla CatalagoTipoVisistante")
                        registros_incorrectos.append(datos)
                        continue

                    # Buscar si la fila ya existe en la base de datos
                    inventario_existente = GastoDerrama.objects.filter(ano=ano, destino=destino, tipo_visitante=tipo_visitante)
                    if inventario_existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {row} ya existe en la base de datos")
                        registros_existentes.append(datos)
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        gastoDiario = GastoDerrama(
                            gasto_diario_prom=gasto_diario_prom_float,
                            ano=ano_int,
                            tipo_visitante=tipo_visitante,
                            destino=destino,
                            participacion_en_hospedaje=participacion_en_hospedaje_float,
                            estadia_promedio=estadia_promedio_float
                        )
                        gastoDiario.save()
                        registros_correctos.append(datos)
                except (ValueError, TypeError) as e:
                    # Si los datos no son válidos, se guarda el número de fila en la lista de registros incorrectos
                    registros_incorrectos.append(datos)
                    
        except FileNotFoundError:
                print(f"El archivo {archivo} no se pudo abrir")
                
        return registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas
    
    def procesar_archivo_csv(self, archivo):
        archivo = self.request.FILES['archivo']
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        try:
            datos = csv.DictReader(archivo.read().decode('latin-1').splitlines())
            # print(datos)
            for row in datos:
                num_filas_procesadas += 1

                # Limpieza de datos
                destino = clean_str_col(row['destino'])
                tipo_visitante = clean_str_col(row['tipo_visitante'])

                # Homologación de datos
                destino = homologar_columna_destino(destino)
                tipo_visitante = CatalagoTipoVisistante.homologar_tipo_visitante(tipo_visitante)

                # Validar si el destino es válido
                if destino not in CatalagoDestino.objects.values_list('destino', flat=True):
                    print(f"El destino {destino} no está en la tabla CatalagoDestino")
                    registros_incorrectos.append(datos)
                    continue
                # Validar si el tipo_visitante es válido
                if not CatalagoTipoVisistante.objects.filter(tipo_visitante=tipo_visitante).exists():
                    print(f"El tipo_visitante: {tipo_visitante} no está en la tabla CatalagoTipoVisistante")
                    registros_incorrectos.append(datos)
                    continue

                # Accede a los valores de cada fila utilizando los nombres de las columnas del archivo CSV
                gasto_diario_prom = row['gasto_diario_prom']
                ano = row['ano']
                participacion_en_hospedaje = row['participacion_en_hospedaje']
                estadia_promedio = row['estadia_promedio']

                try:
                    # Validar los datos
                    gasto_diario_prom_float = float(gasto_diario_prom)
                    ano_int = int(ano)
                    participacion_en_hospedaje_float = float(participacion_en_hospedaje)
                    estadia_promedio_float = float(estadia_promedio)

                    # Buscar si la fila ya existe en la base de datos
                    inventario_existente = GastoDerrama.objects.filter(ano=ano, destino=destino, tipo_visitante=tipo_visitante)
                    if inventario_existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {row} ya existe en la base de datos")
                        registros_existentes.append(datos)
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        gastoDiario = GastoDerrama(
                            gasto_diario_prom=gasto_diario_prom_float,
                            ano=ano_int,
                            tipo_visitante=tipo_visitante,
                            destino=destino,
                            participacion_en_hospedaje=participacion_en_hospedaje_float,
                            estadia_promedio=estadia_promedio_float
                        )
                        gastoDiario.save()
                        registros_correctos.append(datos)
                except (ValueError, TypeError) as e:
                    print(f"Error al procesar la fila {row}: {e}")
                    registros_incorrectos.append(datos)
        except FileNotFoundError:
            print(f"No se encontró el archivo {archivo}")
        except Exception as e:
            print(f"Error al procesar el archivo {archivo}: {e}")
        return registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas

    
class GastoDerramaDescargarArchivoView(SuperAdminOrAdminMixin, LoginRequiredMixin, View):

    def crear_archivo_excel(self, registros_incorrectos):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Escribir encabezados de columna
        worksheet['A1'] = 'gasto_diario_prom'
        worksheet['B1'] = 'Año'
        worksheet['C1'] = 'tipo_visitante'
        worksheet['D1'] = 'destino'
        worksheet['E1'] = 'participacion_en_hospedaje'
        worksheet['F1'] = 'estadia_promedio'

        # Add the incorrect rows to the worksheet
        for i, row in enumerate(registros_incorrectos):
            fila = i + 2
            worksheet.cell(row=fila, column=1, value=row['gasto_diario_prom'])
            worksheet.cell(row=fila, column=2, value=row['ano'])
            worksheet.cell(row=fila, column=3, value=row['tipo_visitante'])
            worksheet.cell(row=fila, column=4, value=row['destino'])
            worksheet.cell(row=fila, column=5, value=row['participacion_en_hospedaje'])
            worksheet.cell(row=fila, column=6, value=row['estadia_promedio'])



        # Set the column widths to auto-fit
        for column in worksheet.columns:
            max_length = 0
            column_name = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_name].width = adjusted_width

        # Create the response with the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=gasto_derrama_registros_incorrectos.xls'

        

        # workbook.save(response)
        return workbook
    
    def post(self, request, *args, **kwargs):
        # Obtener los registros incorrectos del cuerpo de la petición
        registros_incorrectos = json.loads(request.body)

        # Crear y enviar el archivo de Excel con las filas incorrectas
        workbook = self.crear_archivo_excel(registros_incorrectos)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=gasto_derrama_registros_incorrectos.xlsx'
        workbook.save(response)
        return response   