from typing import Any, Dict
from django.shortcuts import render
from django.http import HttpRequest, JsonResponse, HttpResponseRedirect
from django.urls import reverse_lazy, reverse
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from back.models import *
from back.forms import *
from web.models import *
from django.shortcuts import render
from django.http import HttpResponse
import csv
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_GET
from django.views.decorators.http import require_POST
from django.template.loader import render_to_string
# para archivo excel
from datetime import datetime, date
from django.views import View
from django.contrib import messages
from openpyxl import load_workbook
import csv
import os
from django.urls import reverse
import openpyxl
from django.http import HttpResponse
import json
from config.diccionarios import clean_str_col, homologar_columna_categoria, homologar_columna_destino
from django.contrib.auth.decorators import login_required, permission_required
from django.utils.decorators import method_decorator
from django.http import Http404
from django.core.exceptions import PermissionDenied

from django.contrib.auth.decorators import user_passes_test

def es_admin_o_superadmin(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)


def is_ajax(request):
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'

@method_decorator(login_required(login_url='/auth/login_user'), name='dispatch')
@method_decorator(permission_required('auth.view_banner', raise_exception=True), name='dispatch')
@method_decorator(user_passes_test(es_admin_o_superadmin, login_url='404'), name='dispatch')
class FuenteInfoAirbnb (ListView):
    model = Airbnb
    template_name = 'back/fuente_info_airbnb/viewer.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Listado de Fuentes de Informacion Airbnb'
        context['create_url'] = reverse_lazy('dashboard:fuente_info_airbnb_create')
        context['entity'] = 'Airbnb'
        context['is_fuente'] = True
        context['carga_masiva_url'] = reverse_lazy('dashboard:fuente_info_airbnb_carga_masiva')
        return context

@method_decorator(login_required(login_url='/auth/login_user'), name='dispatch')
@method_decorator(permission_required('auth.view_banner', raise_exception=True), name='dispatch')
@method_decorator(user_passes_test(es_admin_o_superadmin, login_url='404'), name='dispatch')
class FuenteInfoAirbnbCreate (CreateView):
    model = Airbnb
    form_class = AirbnbForm
    template_name = 'back/fuente_info_airbnb/create.html'
    success_url = reverse_lazy('dashboard:fuente_info_airbnb')

    def get_object(self, **kwargs):
        queryset = self.get_queryset()
        try:
            return queryset.get(**kwargs)
        except queryset.model.DoesNotExist:
            return None


    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        if form.is_valid():
            # Check if there is already a record with the same fecha, destino and categoria
            fechaI = form.cleaned_data['fecha_inicio']
            destino = form.cleaned_data['destino']
            fechaF = form.cleaned_data['fecha_fin']

            try:
                existing_object = self.get_object(fecha_inicio=fechaI, destino=destino, fecha_fin=fechaF)

            except Certificacion.DoesNotExist:
                existing_object = None

            existing_catalogo = CatalagoDestino.objects.filter(destino__iexact=destino).exists()
            # ALTER TABLE mytable MODIFY mycolumn VARCHAR(255) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci;

            # If there is no existing data, save the new data
            if not existing_catalogo:

                if not existing_catalogo:
                    data = {
                        'success': False,
                        'missingData': True,
                        'destino': destino,
                        'message': 'No existe el destino en el catalogo',
                    }
                    return JsonResponse(data)

            # If there is existing data and replace_data is True, delete the existing data
            if existing_object and request.POST.get('replace_data') == 'on':
                # existing_object.delete()
                # Save the new data
                # self.object = form.save()

                for field in form.cleaned_data:
                    if field == 'replace_data':
                        continue
                    setattr(existing_object, field, form.cleaned_data[field])

                existing_object.save()

                data = {
                    'success': True,
                    'message': 'Data created successfully.',
                    'url': self.success_url,

                }
                return JsonResponse(data)

            # If there is existing data and replace_data is False, return an error

            if existing_object:
                data =  Airbnb.objects.filter(fecha_inicio=fechaI, destino=destino, fecha_fin=fechaF)

                data_list = list(data.values( 'fecha_inicio', 'fecha_fin' , 'destino', 'propiedad_renta','porcentaje_ocupacion','tarifa_promedio'))
                data_list2 = list(form.cleaned_data.values())
                table_html = render_to_string('back/fuente_info_airbnb/table.html',{'data_list': data_list, 'actual': True, 'data_list2': data_list2})

                datajsn = {
                    'success': False,
                    'message': 'Hubo un error al crear registro.',
                    'errors': 'Ya existe un registro con la misma fecha, destino , origen y museo o zona arqueologica',
                    'existing_object': table_html
                }

                return JsonResponse(datajsn)
            
            else:
                self.object = form.save()
                data = {
                    'success': True,
                    'message': 'Data created successfully.',
                    'url': self.success_url
                }
                return JsonResponse(data)
        else:
            data = {
                'success': False,
                'message': 'Error creating data.',
                'errors': form.errors,
                'format_errors': True
            }
            return JsonResponse(data)

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Hubo un error al crear registro.',
            'errors': form.errors
        }
        return JsonResponse(data)

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Registro creado exitosamente.',
            'url': self.success_url
        }
        return JsonResponse(data)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Crear una fuente'
        context['entity'] = 'Glosario'
        context['list_url'] = reverse_lazy('dashboard:fuente_info_airbnb')
        context['action'] = 'add'
        return context

@method_decorator(login_required(login_url='/auth/login_user'), name='dispatch')
@method_decorator(permission_required('auth.view_banner', raise_exception=True), name='dispatch')
@method_decorator(user_passes_test(es_admin_o_superadmin, login_url='404'), name='dispatch')
class FuenteInfoAirbnbUpdate (UpdateView): 
    model = Airbnb
    form_class = AirbnbForm
    template_name = 'back/fuente_info_airbnb/create.html'
    success_url = reverse_lazy('dashboard:fuente_info_airbnb')

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Hubo un error al subir los datos.',
            'errors': form.errors
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Fuente creada exitosamente.',
            'url': self.success_url
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['list_url'] = reverse_lazy('dashboard:fuente_info_airbnb')
            # Set the widget for the 'destino' field to read-only text input
        context['form'].fields['destino'].widget = forms.TextInput(attrs={'readonly': 'readonly'})
        # Set the widget for the 'fecha' field to read-only text input
        context['form'].fields['fecha_inicio'].widget = forms.TextInput(attrs={'readonly': 'readonly'})
        # Set the widget for the 'categoria' field to read-only text input
        context['form'].fields['fecha_fin'].widget = forms.TextInput(attrs={'readonly': 'readonly'})
        context['title'] = 'Editar fuente'
        context['edit_msg'] = 'Los Campos Fecha Inicio , Destino y Fecha Fin no se pueden editar'

        return context
    
@method_decorator(login_required(login_url='/auth/login_user'), name='dispatch')
@method_decorator(permission_required('auth.view_banner', raise_exception=True), name='dispatch')
@method_decorator(user_passes_test(es_admin_o_superadmin, login_url='404'), name='dispatch')
class FuenteInfoAirbnbDelete (DeleteView):
    model = Airbnb
    success_url = reverse_lazy('dashboard:fuente_info_airbnb')
    
    def post(self, request: HttpRequest, *args: str, **kwargs: Any) -> HttpResponse:
        return super().post(request, *args, **kwargs)
@method_decorator(login_required(login_url='/auth/login_user'), name='dispatch')
@method_decorator(permission_required('auth.view_banner', raise_exception=True), name='dispatch')
@method_decorator(user_passes_test(es_admin_o_superadmin, login_url='404'), name='dispatch')
class AirbnbCargaMasivaView(View):
    form_class = CargaMasivaForm
    template_name = 'back/fuente_info_airbnb/carga_masiva.html'
    success_url = reverse_lazy('dashboard:fuente_info_airbnb')

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form, 'title': 'Carga Masiva de Airbnb'})

    def convert_to_serializable(self, obj):
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        raise TypeError(
            f'Object of type {obj.__class__.__name__} is not JSON serializable')

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        archivo = request.FILES.get('archivo', None)
        if archivo:
            extension = os.path.splitext(archivo.name)[1]
            if extension == '.xlsx':
                registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas = self.procesar_archivo_xlsx(
                    archivo)
            elif extension == '.csv':
                registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas = self.procesar_archivo_csv(
                    archivo)
            else:
                messages.error(
                    request, 'El archivo debe ser un archivo .xlsx o .csv')
                registros_incorrectos.append(
                    "El archivo debe ser un archivo .xlsx o .csv")
        else:
            messages.error(request, 'Debe seleccionar un archivo')
            registros_incorrectos.append("Debe seleccionar un archivo")

        if len(registros_incorrectos) > 0 or len(registros_existentes) > 0:
            messages.error(request, 'Hay errores de registros')
            datos_json = json.dumps(
                registros_incorrectos, default=self.convert_to_serializable)

            return render(request, self.template_name, {
                'form': form,
                'title': 'Carga Masiva de Airbnb',
                'registros_correctos': registros_correctos,
                'registros_incorrectos': registros_incorrectos,
                'registros_existentes': registros_existentes,
                'descargar_url': datos_json,
                'num_filas_procesadas': num_filas_procesadas,
            })

        else:
            return HttpResponseRedirect(reverse('dashboard:fuente_info_airbnb'))

    def procesar_archivo_xlsx(self, archivo):
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        try:
            workbook = load_workbook(filename=archivo, read_only=True)
            worksheet = workbook.active
            filas = list(worksheet.rows)
            for i, row in enumerate(filas):
                if i == 0:
                    continue  # Ignorar la primera fila si es el encabezado
                num_filas_procesadas += 1

                # Limpieza de datos
                destino = clean_str_col(row[2].value)

                # Homologación de datos
                destino = homologar_columna_destino(destino)

                fecha_inicio = row[0].value
                fecha_inicio_str = str(fecha_inicio)
                print(fecha_inicio_str)
                fecha_inicio_str = fecha_inicio_str.split()[0] if fecha_inicio_str else ''
                fecha_inicio_obj = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').strftime('%Y-%m-%d')
                
                fecha_fin = row[1].value
                fecha_fin_str = str(fecha_fin)
                fecha_fin_str = fecha_fin_str.split()[0] if fecha_fin_str else ''
                fecha_fin_obj = datetime.strptime(fecha_fin_str, '%Y-%m-%d').strftime('%Y-%m-%d')
                
                propiedad_renta = row[3].value
                porcentaje_ocupacion = row[4].value
                tarifa_promedio = row[5].value

                datos = {
                    'fecha_inicio':fecha_inicio_str , 
                    'fecha_fin':fecha_fin_str , 
                    'destino':destino , 
                    'propiedad_renta':propiedad_renta , 
                    'porcentaje_ocupacion':porcentaje_ocupacion,
                    'tarifa_promedio':tarifa_promedio , 
                }

                try:
                    if destino not in CatalagoDestino.objects.values_list('destino', flat=True):
                        print(f"El destino {destino} no está en la tabla CatalagoDestinoAeropuerto")
                        registros_incorrectos.append(datos)
                        continue

                    # Buscar si la fila ya existe en la base de datos
                    existente = Airbnb.objects.filter(
                        fecha_inicio = fecha_inicio_obj,
                        fecha_fin = fecha_fin_obj,
                        destino = destino
                    )
                    if existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {row} ya existe en la base de datos")
                        registros_existentes.append(datos)
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        db = Airbnb(
                            fecha_inicio = fecha_inicio_obj,
                            fecha_fin = fecha_fin_obj,
                            destino = destino,
                            propiedad_renta = propiedad_renta,
                            porcentaje_ocupacion= porcentaje_ocupacion,
                            tarifa_promedio = tarifa_promedio,
                        )
                        db.save()
                        registros_correctos.append(datos)
                except (ValueError, TypeError) as e:
                    print(f"Error al procesar la fila {datos}: {e}")
                    registros_incorrectos.append(datos)

        except FileNotFoundError:
            print(f"El archivo {archivo} no se pudo abrir")

        return registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas

    def procesar_archivo_csv(self, archivo):
        archivo = self.request.FILES['archivo']
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        try:
            datos = csv.DictReader(
                archivo.read().decode('latin-1').splitlines())
            # print(datos)
            for row in datos:
                num_filas_procesadas += 1

                # Limpieza de datos
                destino = clean_str_col(row['destino'])

                # Homologación de datos
                destino = homologar_columna_destino(destino)

                fecha_inicio = row['fecha_inicio ']
                fecha_fin = row['fecha_fin ']
                
                propiedad_renta = row['propiedad_renta ']
                porcentaje_ocupacion = row['porcentaje_ocupacion']
                tarifa_promedio = row['tarifa_promedio ']

                datos = {
                    'fecha_inicio':fecha_inicio , 
                    'fecha_fin':fecha_fin , 
                    'destino':destino , 
                    'propiedad_renta':propiedad_renta , 
                    'porcentaje_ocupacion':porcentaje_ocupacion,
                    'tarifa_promedio':tarifa_promedio , 
                }

                try:
                    if destino not in CatalagoDestino.objects.values_list('destino', flat=True):
                        print(f"El destino {destino} no está en la tabla CatalagoDestinoAeropuerto")
                        registros_incorrectos.append(datos)
                        continue

                    # Buscar si la fila ya existe en la base de datos
                    existente = Airbnb.objects.filter(
                        fecha_inicio = fecha_inicio,
                        fecha_fin = fecha_fin,
                        destino = destino
                    )
                    if existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {row} ya existe en la base de datos")
                        registros_existentes.append(datos)
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        db = Airbnb(
                            fecha_inicio = fecha_inicio,
                            fecha_fin = fecha_fin,
                            destino = destino,
                            propiedad_renta = propiedad_renta,
                            porcentaje_ocupacion=porcentaje_ocupacion,
                            tarifa_promedio = tarifa_promedio,
                        )
                        db.save()
                        registros_correctos.append(datos)
                except (ValueError, TypeError) as e:
                    print(f"Error al procesar la fila {datos}: {e}")
                    registros_incorrectos.append(datos)
        except FileNotFoundError:
            print(f"No se encontró el archivo {archivo}")
        except Exception as e:
            print(f"Error al procesar el archivo {archivo}: {e}")
        return registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas

@method_decorator(login_required(login_url='/auth/login_user'), name='dispatch')
@method_decorator(permission_required('auth.view_banner', raise_exception=True), name='dispatch')
@method_decorator(user_passes_test(es_admin_o_superadmin, login_url='404'), name='dispatch')
class AirbnbDescargarArchivoView(View):

    def crear_archivo_excel(self, registros_incorrectos):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Obtener los nombres y verbose_name de los campos del modelo Airbnb
        fields = Airbnb._meta.get_fields()
        column_labels = [field.verbose_name for field in fields if field.name != 'id']
        column_names = [field.name for field in fields if field.name != 'id']

        # Escribir los encabezados de las columnas
        for i, campo in enumerate(column_labels):
            columna = i + 1
            worksheet.cell(row=1, column=columna, value=campo)

        # Obtener los datos del modelo Airbnb
        datos = registros_incorrectos

        # Escribir los valores en las celdas correspondientes
        fila = 2
        for registro in datos:
            for i, campo in enumerate(column_names):
                if campo != 'id':  # Omitir la clave 'id'
                    columna = i + 1
                    valor = registro[campo]
                    worksheet.cell(row=fila, column=columna, value=valor)
            fila += 1

        # Set the column widths to auto-fit
        for column in worksheet.columns:
            max_length = 0
            column_name = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_name].width = adjusted_width

        # Create the response with the Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=gasto_derrama_registros_incorrectos.xls'

        # workbook.save(response)
        return workbook

    def post(self, request, *args, **kwargs):
        # Obtener los registros incorrectos del cuerpo de la petición
        registros_incorrectos = json.loads(request.body)

        # Crear y enviar el archivo de Excel con las filas incorrectas
        workbook = self.crear_archivo_excel(registros_incorrectos)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=gasto_derrama_registros_incorrectos.xlsx'
        workbook.save(response)
        return response
