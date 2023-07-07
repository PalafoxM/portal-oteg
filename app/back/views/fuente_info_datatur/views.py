from typing import Any
from django import http
from django.shortcuts import render
from django.http import JsonResponse, HttpResponseRedirect
from django.urls import reverse_lazy, reverse
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from back.models import *
from back.forms import *
from web.models import *
from django.shortcuts import render
from django.http import HttpResponse
import csv
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_GET
from django.views.decorators.http import require_POST
from django.shortcuts import get_object_or_404
#serializers
# render to string
from django.template.loader import render_to_string
from django.utils.decorators import method_decorator
# para archivo excel
from django.views import View
from django.contrib import messages
from openpyxl import load_workbook
import csv
import os
import datetime
from django.urls import reverse
import openpyxl
from django.http import HttpResponse
import json
from config.diccionarios import clean_str_col, homologar_columna_categoria, homologar_columna_destino
from django.contrib.auth.decorators import login_required, permission_required
from django.utils.decorators import method_decorator
from django.http import Http404
from django.core.exceptions import PermissionDenied
from back.mixins import *
from django.contrib.auth.decorators import user_passes_test

def es_admin_o_superadmin(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)




def is_ajax(request):
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'



class FuenteInfoDatatur (SuperAdminOrAdminMixin, LoginRequiredMixin, ListView):
    model = DataTour
    template_name = 'back/fuente_info_datatur/viewer.html'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs) :
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'search':
                data = []
                for i in DataTour.objects.all():
                    data.append(i.toJSON())
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)


    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Listado de Fuentes de Informacion de DataTour'
        context['create_url'] = reverse_lazy(
            'dashboard:fuente_info_datatour_create')
        context['entity'] = 'Categorias'
        context['list'] = 'dashboard:fuente_info_datatour'
        context['is_fuente'] = True
        context['carga_masiva_url'] = reverse_lazy('dashboard:fuente_info_datatour_carga_masiva')
        return context



class FuenteInfoDataturCreate (SuperAdminOrAdminMixin, LoginRequiredMixin, CreateView):
    model = DataTour
    form_class = DataTurForm
    template_name = 'back/fuente_info_datatur/create.html'
    success_url = reverse_lazy('dashboard:fuente_info_datatour')

    def get_object(self, **kwargs):
        queryset = self.get_queryset()
        try:
            return queryset.get(**kwargs)
        except queryset.model.DoesNotExist:
            return None

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        replace_data = request.POST.get('replace_data')
        if form.is_valid():
            # Check if there is already a record with the same fecha, destino and categoria
       
            fecha = form.cleaned_data['fecha']
            destino = form.cleaned_data['destino']
            categoria = form.cleaned_data['categoria']
            try:
                existing_object = self.get_object(fecha=fecha, destino=destino, categoria=categoria)
             

            except DataTour.DoesNotExist:
                existing_object = None 

            existing_category = CatalagoCategoria.objects.filter(categoria=categoria).exists()
            existing_catalogo = CatalagoDestino.objects.filter(destino=destino).exists()



            # If there is no existing data, save the new data
            if not existing_category or not existing_catalogo:

                if not existing_category and not existing_catalogo:
    
                    data = {
                        'success': False,
                        'missingData': True,
                        'categoria': categoria,
                        'destino': destino, 
                        'message': 'No existe la categoria o el destino en el catalogo',
                    }
                    return JsonResponse(data)
                
                if not existing_category:
                    data = {
                        'success': False,
                        'missingData': True,
                        'categoria': categoria,
                        'message': 'No existe la categoria en el catalogo',
                    }
                    return JsonResponse(data)
                
                if not existing_catalogo:
                    data = {
                        'success': False,
                        'missingData': True,
                        'destino': destino,
                        'message': 'No existe el destino en el catalogo',
                    }
                    return JsonResponse(data)
            
                       

            # If there is existing data and replace_data is True, delete the existing data
            if existing_object and request.POST.get('replace_data') == 'on':
                #existing_object.delete()
                # Save the new data
                #self.object = form.save()

                for field in form.cleaned_data:
                    if field == 'replace_data':
                        continue
                    setattr(existing_object, field, form.cleaned_data[field])
                existing_object.save()

                data = {
                    'success': True,
                    'message': 'Data created successfully.',
                    'url': self.success_url,
                    
                }
                return JsonResponse(data)
            
            # If there is existing data and replace_data is False, return an error

            if existing_object  :
                data = DataTour.objects.filter(fecha=fecha, destino=destino, categoria=categoria)
                
                data_list = list(data.values('destino',
'categoria',
    'fecha',
    'cuartos_registrados_fin_periodo',
    'cuartos_disponibles_promedio',
    'cuartos_disponibles',
    'cuartos_ocupados',
    'cuartos_ocupados_residentes',
    'cuartos_ocupados_no_residentes',
    'llegadas_de_turistas',
    'llegadas_de_turistas_residentes',
    'llegadas_de_turistas_no_residentes',
    'turistas_noche',
    'turistas_noche_residentes',
    'turistas_noche_no_residentes',
    'porcentaje_de_ocupacion',
    'porcentaje_de_ocupacion_residentes',
    'porcentaje_de_ocupacion_no_residentes',
    'estadia_promedio',
    'estadia_promedio_residentes',
    'estadia_promedio_no_residentes',
    'densidad_de_ocupacion',
    'densidad_de_ocupacion_residentes',
    'densidad_de_ocupacion_no_residentes',))                                                 
                data_list2 =  list(form.cleaned_data.values())
                table_html = render_to_string('back/fuente_info_datatur/data_tour_table.html', {'data_list': data_list , 'actual':True , 'data_list2':data_list2})                            
                
                datajsn = {
                        'success': False,
                        'message': 'Hubo un error al crear registro.',
                        'errors': 'Ya existe un registro con la misma fecha, destino y categoria.',
                        'existing_object': table_html
                }

                return JsonResponse(datajsn)
            else:
                self.object = form.save()
                data = {
                    'success': True,
                    'message': 'Data created successfully.',
                    'url': self.success_url
                }
                return JsonResponse(data)
        else:
            data = {
                'success': False,
                'message': 'Error creating data.',
                'errors': form.errors,
                'format_errors':True
            }
            return JsonResponse(data)

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Hubo un error al crear registro.',
            'errors': form.errors
        }
        return JsonResponse(data)

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Registro creado exitosamente.',
            'url': self.success_url
        }
        return JsonResponse(data)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Crear una fuente'
        context['entity'] = 'Glosario'
        context['list_url'] = reverse_lazy('dashboard:fuente_info_datatour')
        context['action'] = 'add'
        return context


class FuenteInfoDataturUpdate(SuperAdminOrAdminMixin, LoginRequiredMixin, UpdateView):
    
    model = DataTour
    form_class = DataTurForm
    template_name = 'back/components/create_update_fuentes_info.html'
    success_url = reverse_lazy('dashboard:fuente_info_datatour')

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Hubo un error al crear el evento.',
            'errors': form.errors
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Evento creado exitosamente.',
            'url': self.success_url
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['list_url'] = reverse_lazy('dashboard:fuente_info_datatour')
            # Set the widget for the 'destino' field to read-only text input
        context['form'].fields['destino'].widget = forms.TextInput(attrs={'readonly': 'readonly'})
        # Set the widget for the 'fecha' field to read-only text input
        context['form'].fields['fecha'].widget = forms.TextInput(attrs={'readonly': 'readonly'})
        # Set the widget for the 'categoria' field to read-only text input
        context['form'].fields['categoria'].widget = forms.TextInput(attrs={'readonly': 'readonly'})
        context['title'] = 'Editar fuente'
        context['edit_msg'] = 'Los Campos Destino, Fecha y Categoria no pueden ser editados'    

        return context



class FuenteInfoDataturDelete(SuperAdminOrAdminMixin, LoginRequiredMixin, DeleteView):
    model = DataTour
    success_url = reverse_lazy('dashboard:fuente_info_datatour')

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        success_url = self.get_success_url()
        self.object.delete()
        return HttpResponseRedirect(success_url)


class DataturCargaMasivaView(SuperAdminOrAdminMixin, LoginRequiredMixin, View):
    form_class = CargaMasivaForm
    template_name = 'back/fuente_info_datatur/carga_masiva.html'
    success_url = reverse_lazy('dashboard:fuente_info_datatour')

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form, 'title': 'Carga Masiva Datatur'})


    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        archivo = request.FILES.get('archivo', None)
        if archivo:
            extension = os.path.splitext(archivo.name)[1]
            if extension == '.xlsx':
                registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas = self.procesar_archivo_xlsx(archivo)
            elif extension == '.csv':
                registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas = self.procesar_archivo_csv(archivo)
            else:
                messages.error(request, 'El archivo debe ser un archivo .xlsx o .csv')
                registros_incorrectos.append("El archivo debe ser un archivo .xlsx o .csv")
        else:
            messages.error(request, 'Debe seleccionar un archivo')
            registros_incorrectos.append("Debe seleccionar un archivo")

        if len(registros_incorrectos) > 0 or len(registros_existentes) > 0:
            messages.error(request, 'Hay errores de registros')

            datos_json = json.dumps(registros_incorrectos)
            
            return render(request, self.template_name, {
                'form': form,
                'title': 'Carga Masiva',
                'registros_correctos': registros_correctos,
                'registros_incorrectos': registros_incorrectos,
                'registros_existentes': registros_existentes,
                'descargar_url': datos_json,
                'num_filas_procesadas': num_filas_procesadas,
            })
            
        else:
            return HttpResponseRedirect(reverse('dashboard:inventario_hotelero_list'))
        
        

    def procesar_archivo_xlsx(self, archivo):
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        try:
            workbook = load_workbook(filename=archivo, read_only=True)
            worksheet = workbook.active
            filas = list(worksheet.rows)
            for i, row in enumerate(filas):
                if i == 0:
                    continue # Ignorar la primera fila si es el encabezado
                num_filas_procesadas += 1

                # Limpieza de datos
                destino = clean_str_col(row[0].value)
                categoria = clean_str_col(row[1].value)

                # Homologación de datos
                destino = homologar_columna_destino(destino)
                categoria = homologar_columna_categoria(categoria)

                fecha = row[2].value
                # Validar los datos
                fecha_str = str(fecha)
                fecha_str = fecha_str.split()[0] if fecha_str else ''  # Eliminar la parte de la hora si existe la fecha
                fecha_obj = datetime.datetime.strptime(fecha_str, '%Y-%m-%d').date()
                # Serializar la fecha en formato JSON
                json_fecha = json.dumps(fecha_obj.strftime('%Y-%m-%d'))
                json_fecha = str(json_fecha).strip('"')

                
                cuartos_registrados_fin_periodo = row[3].value
                cuartos_disponibles_promedio = row[4].value
                cuartos_disponibles = row[5].value
                cuartos_ocupados = row[6].value
                cuartos_ocupados_residentes = row[7].value
                cuartos_ocupados_no_residentes = row[8].value
                llegadas_de_turistas = row[9].value
                llegadas_de_turistas_residentes = row[10].value
                llegadas_de_turistas_no_residentes = row[11].value
                turistas_noche = row[12].value
                turistas_noche_residentes = row[13].value
                turistas_noche_no_residentes = row[14].value
                porcentaje_de_ocupacion = row[15].value
                porcentaje_de_ocupacion_residentes = row[16].value
                porcentaje_de_ocupacion_no_residentes = row[17].value
                estadia_promedio = row[18].value
                estadia_promedio_residentes = row[19].value
                estadia_promedio_no_residentes = row[20].value
                densidad_de_ocupacion = row[21].value
                densidad_de_ocupacion_residentes = row[22].value
                densidad_de_ocupacion_no_residentes = row[23].value

                data = {
                    'destino': destino,
                    'fecha': json_fecha,
                    'categoria': categoria,
                    'cuartos_registrados_fin_periodo': cuartos_registrados_fin_periodo,
                    'cuartos_disponibles_promedio': cuartos_disponibles_promedio,
                    'cuartos_disponibles': cuartos_disponibles,
                    'cuartos_ocupados': cuartos_ocupados,
                    'cuartos_ocupados_residentes': cuartos_ocupados_residentes,
                    'cuartos_ocupados_no_residentes': cuartos_ocupados_no_residentes,
                    'llegadas_de_turistas': llegadas_de_turistas,
                    'llegadas_de_turistas_residentes': llegadas_de_turistas_residentes,
                    'llegadas_de_turistas_no_residentes': llegadas_de_turistas_no_residentes,
                    'turistas_noche': turistas_noche,
                    'turistas_noche_residentes': turistas_noche_residentes,
                    'turistas_noche_no_residentes': turistas_noche_no_residentes,
                    'porcentaje_de_ocupacion': porcentaje_de_ocupacion,
                    'porcentaje_de_ocupacion_residentes': porcentaje_de_ocupacion_residentes,
                    'porcentaje_de_ocupacion_no_residentes': porcentaje_de_ocupacion_no_residentes,
                    'estadia_promedio': estadia_promedio,
                    'estadia_promedio_residentes': estadia_promedio_residentes,
                    'estadia_promedio_no_residentes': estadia_promedio_no_residentes,
                    'densidad_de_ocupacion': densidad_de_ocupacion,
                    'densidad_de_ocupacion_residentes': densidad_de_ocupacion_residentes,
                    'densidad_de_ocupacion_no_residentes': densidad_de_ocupacion_no_residentes
                }

                # Validar si el destino y categoria son válidos
                if destino not in CatalagoDestino.objects.values_list('destino', flat=True):
                    print(f"El destino {destino} no está en la tabla CatalagoDestino")
                    registros_incorrectos.append(data)
                    continue
                if categoria not in CatalagoCategoria.objects.values_list('categoria', flat=True):
                    print(f"La categoría {categoria} no está en la tabla CatalagoCategoria")
                    registros_incorrectos.append(data)
                    continue



                try:
                    

                    # Buscar si la fila ya existe en la base de datos
                    inventario_existente = InventarioHotelero.objects.filter(destino=destino, fecha=fecha_obj, categoria=categoria)
                    if inventario_existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {row} ya existe en la base de datos")
                        registros_existentes.append(data)

                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        inventario = inventario = InventarioHotelero(
                            destino=destino, 
                            fecha=fecha_obj, 
                            categoria=categoria, 
                            cuartos_registrados_fin_periodo = cuartos_registrados_fin_periodo,
                            cuartos_disponibles_promedio = cuartos_disponibles_promedio,
                            cuartos_disponibles = cuartos_disponibles,
                            cuartos_ocupados = cuartos_ocupados,
                            cuartos_ocupados_residentes = cuartos_ocupados_residentes,
                            cuartos_ocupados_no_residentes = cuartos_ocupados_no_residentes,
                            llegadas_de_turistas = llegadas_de_turistas,
                            llegadas_de_turistas_residentes = llegadas_de_turistas_residentes,
                            llegadas_de_turistas_no_residentes = llegadas_de_turistas_no_residentes,
                            turistas_noche = turistas_noche,
                            turistas_noche_residentes = turistas_noche_residentes,
                            turistas_noche_no_residentes = turistas_noche_no_residentes,
                            porcentaje_de_ocupacion = porcentaje_de_ocupacion,
                            porcentaje_de_ocupacion_residentes = porcentaje_de_ocupacion_residentes,
                            porcentaje_de_ocupacion_no_residentes = porcentaje_de_ocupacion_no_residentes,
                            estadia_promedio = estadia_promedio,
                            estadia_promedio_residentes = estadia_promedio_residentes,
                            estadia_promedio_no_residentes = estadia_promedio_no_residentes,
                            densidad_de_ocupacion = densidad_de_ocupacion,
                            densidad_de_ocupacion_residentes = densidad_de_ocupacion_residentes,
                            densidad_de_ocupacion_no_residentes = densidad_de_ocupacion_no_residentes
                        )
                        inventario.save()
                        registros_correctos.append(data)
                except (ValueError, TypeError) as e:
                    # Si los datos no son válidos, se guarda el número de fila en la lista de registros incorrectos
                    registros_incorrectos.append(data)
                    
        except FileNotFoundError:
                print(f"El archivo {archivo} no se pudo abrir")
                
        return registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas
    
    def procesar_archivo_csv(self, archivo):
        archivo = self.request.FILES['archivo']
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        try:
            datos = csv.DictReader(archivo.read().decode('latin-1').splitlines())
            # print(datos)
            for row in datos:
                num_filas_procesadas += 1

                # Limpieza de datos
                destino = clean_str_col(row['destino'])
                categoria = clean_str_col(row['categoria'])

                # Homologación de datos
                destino = homologar_columna_destino(destino)
                categoria = homologar_columna_categoria(categoria)

                # Validar si el destino y categoria son válidos
                if destino not in CatalagoDestino.objects.values_list('destino', flat=True):
                    print(f"El destino {destino} no está en la tabla CatalagoDestino")
                    registros_incorrectos.append(datos)
                    continue
                if categoria not in CatalagoCategoria.objects.values_list('categoria', flat=True):
                    print(f"La categoría {categoria} no está en la tabla CatalagoCategoria")
                    registros_incorrectos.append(datos)
                    continue


                destino = row['destino']
                categoria = row['categoria']
                fecha = row['fecha']
                cuartos_registrados_fin_periodo = float(row['cuartos_registrados_fin_periodo'])
                cuartos_disponibles_promedio = float(row['cuartos_disponibles_promedio'])
                cuartos_disponibles = float(row['cuartos_disponibles'])
                cuartos_ocupados = float(row['cuartos_ocupados'])
                cuartos_ocupados_residentes = float(row['cuartos_ocupados_residentes'])
                cuartos_ocupados_no_residentes = float(row['cuartos_ocupados_no_residentes'])
                llegadas_de_turistas = float(row['llegadas_de_turistas'])
                llegadas_de_turistas_residentes = float(row['llegadas_de_turistas_residentes'])
                llegadas_de_turistas_no_residentes = float(row['llegadas_de_turistas_no_residentes'])
                turistas_noche = float(row['turistas_noche'])
                turistas_noche_residentes = float(row['turistas_noche_residentes'])
                turistas_noche_no_residentes = float(row['turistas_noche_no_residentes'])
                porcentaje_de_ocupacion = float(row['porcentaje_de_ocupacion'])
                porcentaje_de_ocupacion_residentes = float(row['porcentaje_de_ocupacion_residentes'])
                porcentaje_de_ocupacion_no_residentes = float(row['porcentaje_de_ocupacion_no_residentes'])
                estadia_promedio = float(row['estadia_promedio'])
                estadia_promedio_residentes = float(row['estadia_promedio_residentes'])
                estadia_promedio_no_residentes = float(row['estadia_promedio_no_residentes'])
                densidad_de_ocupacion = float(row['densidad_de_ocupacion'])
                densidad_de_ocupacion_residentes = float(row['densidad_de_ocupacion_residentes'])
                densidad_de_ocupacion_no_residentes = float(row['densidad_de_ocupacion_no_residentes'])


                try:
                    # Validar los datos
                    fecha_str = str(fecha)
                    fecha_obj = datetime.datetime.strptime(fecha_str, '%d/%m/%Y').date()
                    

                    # Buscar si la fila ya existe en la base de datos
                    inventario_existente = InventarioHotelero.objects.filter(destino=destino, fecha=fecha_obj, categoria=categoria)
                    if inventario_existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {row} ya existe en la base de datos")
                        registros_existentes.append(datos)
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        inventario = InventarioHotelero(
                            destino=destino, 
                            fecha=fecha_obj, 
                            categoria=categoria, 
                            cuartos_registrados_fin_periodo = cuartos_registrados_fin_periodo,
                            cuartos_disponibles_promedio = cuartos_disponibles_promedio,
                            cuartos_disponibles = cuartos_disponibles,
                            cuartos_ocupados = cuartos_ocupados,
                            cuartos_ocupados_residentes = cuartos_ocupados_residentes,
                            cuartos_ocupados_no_residentes = cuartos_ocupados_no_residentes,
                            llegadas_de_turistas = llegadas_de_turistas,
                            llegadas_de_turistas_residentes = llegadas_de_turistas_residentes,
                            llegadas_de_turistas_no_residentes = llegadas_de_turistas_no_residentes,
                            turistas_noche = turistas_noche,
                            turistas_noche_residentes = turistas_noche_residentes,
                            turistas_noche_no_residentes = turistas_noche_no_residentes,
                            porcentaje_de_ocupacion = porcentaje_de_ocupacion,
                            porcentaje_de_ocupacion_residentes = porcentaje_de_ocupacion_residentes,
                            porcentaje_de_ocupacion_no_residentes = porcentaje_de_ocupacion_no_residentes,
                            estadia_promedio = estadia_promedio,
                            estadia_promedio_residentes = estadia_promedio_residentes,
                            estadia_promedio_no_residentes = estadia_promedio_no_residentes,
                            densidad_de_ocupacion = densidad_de_ocupacion,
                            densidad_de_ocupacion_residentes = densidad_de_ocupacion_residentes,
                            densidad_de_ocupacion_no_residentes = densidad_de_ocupacion_no_residentes
                        )
                        inventario.save()
                        registros_correctos.append(datos)
                except (ValueError, TypeError) as e:
                    print(f"Error al procesar la fila {row}: {e}")
                    registros_incorrectos.append(datos)
        except FileNotFoundError:
            print(f"No se encontró el archivo {archivo}")
        except Exception as e:
            print(f"Error al procesar el archivo {archivo}: {e}")
        return registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas



class DescargarArchivoDataturView(SuperAdminOrAdminMixin, LoginRequiredMixin, View):

    def crear_archivo_excel(self, registros_incorrectos):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Add headers to the worksheet
        worksheet['A1'] = 'destino'
        worksheet['B1'] = 'categoria'
        worksheet['C1'] = 'fecha'
        worksheet['D1'] = 'cuartos_registrados_fin_periodo'
        worksheet['E1'] = 'cuartos_disponibles_promedio'
        worksheet['F1'] = 'cuartos_disponibles'
        worksheet['G1'] = 'cuartos_ocupados'
        worksheet['H1'] = 'cuartos_ocupados_residentes'
        worksheet['I1'] = 'cuartos_ocupados_no_residentes'
        worksheet['J1'] = 'llegadas_de_turistas'
        worksheet['K1'] = 'llegadas_de_turistas_residentes'
        worksheet['L1'] = 'llegadas_de_turistas_no_residentes'
        worksheet['M1'] = 'turistas_noche'
        worksheet['N1'] = 'turistas_noche_residentes'
        worksheet['O1'] = 'turistas_noche_no_residentes'
        worksheet['P1'] = 'porcentaje_de_ocupacion'
        worksheet['Q1'] = 'porcentaje_de_ocupacion_residentes'
        worksheet['R1'] = 'porcentaje_de_ocupacion_no_residentes'
        worksheet['S1'] = 'estadia_promedio'
        worksheet['T1'] = 'estadia_promedio_residentes'
        worksheet['U1'] = 'estadia_promedio_no_residentes'
        worksheet['V1'] = 'densidad_de_ocupacion'
        worksheet['W1'] = 'densidad_de_ocupacion_residentes'
        worksheet['X1'] = 'densidad_de_ocupacion_no_residentes'

        # Add the incorrect rows to the worksheet
        # Add the incorrect rows to the worksheet
        for i, row in enumerate(registros_incorrectos):
            fila = i + 2
            worksheet.cell(row=fila, column=1, value=row['destino'])
            worksheet.cell(row=fila, column=2, value=row['categoria'])
            worksheet.cell(row=fila, column=3, value=row['fecha'])
            worksheet.cell(row=fila, column=4, value=row['cuartos_registrados_fin_periodo'])
            worksheet.cell(row=fila, column=5, value=row['cuartos_disponibles_promedio'])
            worksheet.cell(row=fila, column=6, value=row['cuartos_disponibles'])
            worksheet.cell(row=fila, column=7, value=row['cuartos_ocupados'])
            worksheet.cell(row=fila, column=8, value=row['cuartos_ocupados_residentes'])
            worksheet.cell(row=fila, column=9, value=row['cuartos_ocupados_no_residentes'])
            worksheet.cell(row=fila, column=10, value=row['llegadas_de_turistas'])
            worksheet.cell(row=fila, column=11, value=row['llegadas_de_turistas_residentes'])
            worksheet.cell(row=fila, column=12, value=row['llegadas_de_turistas_no_residentes'])
            worksheet.cell(row=fila, column=13, value=row['turistas_noche'])
            worksheet.cell(row=fila, column=14, value=row['turistas_noche_residentes'])
            worksheet.cell(row=fila, column=15, value=row['turistas_noche_no_residentes'])
            worksheet.cell(row=fila, column=16, value=row['porcentaje_de_ocupacion'])
            worksheet.cell(row=fila, column=18, value=row['porcentaje_de_ocupacion_no_residentes'])
            worksheet.cell(row=fila, column=19, value=row['estadia_promedio'])
            worksheet.cell(row=fila, column=20, value=row['estadia_promedio_residentes'])
            worksheet.cell(row=fila, column=21, value=row['estadia_promedio_no_residentes'])
            worksheet.cell(row=fila, column=22, value=row['densidad_de_ocupacion'])
            worksheet.cell(row=fila, column=23, value=row['densidad_de_ocupacion_residentes'])
            worksheet.cell(row=fila, column=24, value=row['densidad_de_ocupacion_no_residentes'])



        # Set the column widths to auto-fit
        for column in worksheet.columns:
            max_length = 0
            column_name = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_name].width = adjusted_width

        # Create the response with the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=datatur_registros_incorrectos.xls'

        

        # workbook.save(response)
        return workbook
    
    def post(self, request, *args, **kwargs):
        # Obtener los registros incorrectos del cuerpo de la petición
        registros_incorrectos = json.loads(request.body)

        # Crear y enviar el archivo de Excel con las filas incorrectas
        workbook = self.crear_archivo_excel(registros_incorrectos)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=datatur_registros_incorrectos.xlsx'
        workbook.save(response)
        return response   