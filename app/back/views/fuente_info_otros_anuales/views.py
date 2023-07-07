
from django.shortcuts import render
from django.http import JsonResponse, HttpResponseRedirect
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from back.models import *
from back.forms import *
from web.models import *

from django.template.loader import render_to_string
# para archivo excel
from django.views import View
from django.contrib import messages
from openpyxl import load_workbook
import csv
import os
import datetime
from django.urls import reverse
import openpyxl
from django.http import HttpResponse
import json
from config.diccionarios import clean_str_col, homologar_columna_categoria, homologar_columna_destino
from django.contrib.auth.decorators import login_required, permission_required
from django.utils.decorators import method_decorator
from django.http import Http404
from django.core.exceptions import PermissionDenied
from back.mixins import *
from django.contrib.auth.decorators import user_passes_test

def es_admin_o_superadmin(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)



def is_ajax(request):
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'



class FuenteInfoOtrosAnuales (SuperAdminOrAdminMixin, LoginRequiredMixin, ListView):
    model = otros_anuales
    template_name = 'back/fuente_info_otros_anuales/viewer.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Listado de Fuentes de Informacion Otros'
        context['create_url'] = reverse_lazy('dashboard:fuente_info_otros_anuales_create')
        context['entity'] = 'Otros Anuales'
        context['is_fuente'] = True
        context['carga_masiva_url'] = reverse_lazy('dashboard:fuente_otros_anuales_carga_masiva')
        return context



class FuenteInfoOtrosAnualesCreate (SuperAdminOrAdminMixin, LoginRequiredMixin, CreateView):
    model = otros_anuales
    form_class = OtrosAnualesForm
    success_url = reverse_lazy('dashboard:fuente_info_otros_anuales')
    template_name = 'back/fuente_info_otros_anuales/create.html'

    def get_object(self, **kwargs):
        queryset = self.get_queryset()
        try:
            return queryset.get(**kwargs)
        except queryset.model.DoesNotExist:
            return None

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        replace_data = request.POST.get('replace_data')
        if form.is_valid():
            # Check if there is already a record with the same fecha, destino and categoria
       
            anio = form.cleaned_data['ano']

            try:
                existing_object = self.get_object(ano=anio)
             
            except otros_anuales.DoesNotExist:
                existing_object = None
    

            # If there is existing data and replace_data is True, delete the existing data
            if existing_object and request.POST.get('replace_data') == 'on':
                #existing_object.delete()
                # Save the new data
                #self.object = form.save()

                for field in form.cleaned_data:
                    if field == 'replace_data':
                        continue
                    setattr(existing_object, field, form.cleaned_data[field])
                existing_object.save()

                data = {
                    'success': True,
                    'message': 'Data created successfully.',
                    'url': self.success_url,
                    
                }
                return JsonResponse(data)
            
            # If there is existing data and replace_data is False, return an error

            if existing_object  :
                data = otros_anuales.objects.filter(ano=anio)
                data_list = list(data.values('ano','PIB_sector_72','PIB_actividades_terciarias','basura_generada_persona_diaria_Kg'))                                      
                data_list2 =  list(form.cleaned_data.values())
                table_html = render_to_string('back/fuente_info_otros_anuales/table.html', {'data_list': data_list , 'actual':True , 'data_list2':data_list2})                            
                
                datajsn = {
                        'success': False,
                        'message': 'Hubo un error al crear registro.',
                        'errors': 'Ya existe un registro con la misma fecha, destino y categoria.',
                        'existing_object': table_html
                }

                return JsonResponse(datajsn)
            else:
                self.object = form.save()
                data = {
                    'success': True,
                    'message': 'Data created successfully.',
                    'url': self.success_url
                }
                return JsonResponse(data)
        else:
            data = {
                'success': False,
                'message': 'Error creating data.',
                'errors': form.errors,
                'format_errors':True
            }
            return JsonResponse(data)

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Hubo un error al crear registro.',
            'errors': form.errors
        }
        return JsonResponse(data)

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Registro creado exitosamente.',
            'url': self.success_url
        }
        return JsonResponse(data)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Crear una fuente'
        context['entity'] = 'Glosario'
        context['list_url'] = reverse_lazy('dashboard:fuente_info_otros_anuales')
        context['action'] = 'add'
        return context
    


class FuenteInfoOtrosAnualesUpdate (SuperAdminOrAdminMixin, LoginRequiredMixin, UpdateView):
    model = otros_anuales
    form_class = OtrosAnualesForm
    success_url = reverse_lazy('dashboard:fuente_info_otros_anuales')
    template_name = 'back/fuente_info_otros_anuales/view_editor.html'

    
    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Hubo un error al crear el evento.',
            'errors': form.errors
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Evento creado exitosamente.',
            'url': self.success_url
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['list_url'] = reverse_lazy('dashboard:fuente_info_otros_anuales')
        # Set the widget for the 'destino' field to read-only text input
        context['form'].fields['ano'].widget = forms.TextInput(attrs={'readonly': 'readonly'})
      
        context['title'] = 'Editar fuente'
        context['edit_msg'] = 'El campo Año no pueden ser editado'    

        return context



class FuenteInfoOtrosAnualesDelete (SuperAdminOrAdminMixin, LoginRequiredMixin, DeleteView):
    model = otros_anuales
    success_url = reverse_lazy('dashboard:fuente_info_otros_anuales')

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        success_url = self.get_success_url()
        self.object.delete()
        return HttpResponseRedirect(success_url)


class OtrosAnualesCargaMasivaView(SuperAdminOrAdminMixin, LoginRequiredMixin, View):
    form_class = CargaMasivaForm
    template_name = 'back/fuente_info_otros_anuales/carga_masiva.html'
    success_url = reverse_lazy('dashboard:fuente_info_otros_anuales')

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form, 'title': 'Carga Masiva Otros Anuales'})


    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        archivo = request.FILES.get('archivo', None)
        if archivo:
            extension = os.path.splitext(archivo.name)[1]
            if extension == '.xlsx':
                registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas = self.procesar_archivo_xlsx(archivo)
            elif extension == '.csv':
                registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas = self.procesar_archivo_csv(archivo)
            else:
                messages.error(request, 'El archivo debe ser un archivo .xlsx o .csv')
                registros_incorrectos.append("El archivo debe ser un archivo .xlsx o .csv")
        else:
            messages.error(request, 'Debe seleccionar un archivo')
            registros_incorrectos.append("Debe seleccionar un archivo")

        if len(registros_incorrectos) > 0 or len(registros_existentes) > 0:
            messages.error(request, 'Hay errores de registros')
            datos_json = json.dumps(registros_incorrectos)
            
            return render(request, self.template_name, {
                'form': form,
                'title': 'Carga Masiva Otros Anuales',
                'registros_correctos': registros_correctos,
                'registros_incorrectos': registros_incorrectos,
                'registros_existentes': registros_existentes,
                'descargar_url': datos_json,
                'num_filas_procesadas': num_filas_procesadas,
            })
            
        else:
            return HttpResponseRedirect(reverse('dashboard:fuente_info_otros_anuales'))
        
        

    def procesar_archivo_xlsx(self, archivo):
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        try:
            workbook = load_workbook(filename=archivo, read_only=True)
            worksheet = workbook.active
            filas = list(worksheet.rows)
            for i, row in enumerate(filas):
                if i == 0:
                    continue # Ignorar la primera fila si es el encabezado

                num_filas_procesadas += 1
                # Limpieza de datos
                ano = clean_str_col(row[0].value)
                PIB_sector_72 = clean_str_col(row[1].value)
                PIB_actividades_terciarias = clean_str_col(row[2].value)
                basura_generada_persona_diaria_Kg = clean_str_col(row[3].value)
                datos = {
                    'ano':ano,
                    'PIB_sector_72':PIB_sector_72,
                    'PIB_actividades_terciarias':PIB_actividades_terciarias,
                    'basura_generada_persona_diaria_Kg':basura_generada_persona_diaria_Kg,
                }

                try:
                    # Validar los datos
                    ano_int = int(ano)
                    PIB_sector_72_float = float(PIB_sector_72)
                    PIB_actividades_terciarias_float = float(PIB_actividades_terciarias)
                    basura_generada_persona_diaria_Kg_float = float(basura_generada_persona_diaria_Kg)

                    # Buscar si la fila ya existe en la base de datos
                    inventario_existente = otros_anuales.objects.filter(
                        ano=ano_int, 
                        PIB_sector_72=PIB_sector_72_float, 
                        PIB_actividades_terciarias=PIB_actividades_terciarias_float, 
                        basura_generada_persona_diaria_Kg=basura_generada_persona_diaria_Kg_float
                    )
                    if inventario_existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {row} ya existe en la base de datos")
                        registros_existentes.append(datos)
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        otrosAnuales = otros_anuales(
                            ano=ano_int, 
                            PIB_sector_72=PIB_sector_72_float, 
                            PIB_actividades_terciarias=PIB_actividades_terciarias_float, 
                            basura_generada_persona_diaria_Kg=basura_generada_persona_diaria_Kg_float
                        )
                        otrosAnuales.save()
                        registros_correctos.append(datos)
                except (ValueError, TypeError) as e:
                    # Si los datos no son válidos, se guarda el número de fila en la lista de registros incorrectos
                    registros_incorrectos.append(datos)
                    
        except FileNotFoundError:
                print(f"El archivo {archivo} no se pudo abrir")
                
        return registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas
    
    def procesar_archivo_csv(self, archivo):
        archivo = self.request.FILES['archivo']
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        try:
            datos = csv.DictReader(archivo.read().decode('latin-1').splitlines())
            # print(datos)
            for row in datos:
                num_filas_procesadas += 1

                # Accede a los valores de cada fila utilizando los nombres de las columnas del archivo CSV
                # Limpieza de datos
                ano = clean_str_col(row['ano'])
                PIB_sector_72 = clean_str_col(row['PIB_sector_72'])
                PIB_actividades_terciarias = clean_str_col(row['PIB_actividades_terciarias'])
                basura_generada_persona_diaria_Kg = clean_str_col(row['basura_generada_persona_diaria_Kg'])

                try:
                    # Validar los datos
                    ano_int = int(ano)
                    PIB_sector_72_float = float(PIB_sector_72)
                    PIB_actividades_terciarias_float = float(PIB_actividades_terciarias)
                    basura_generada_persona_diaria_Kg_float = float(basura_generada_persona_diaria_Kg)

                    # Buscar si la fila ya existe en la base de datos
                    inventario_existente = otros_anuales.objects.filter(
                        ano=ano_int, 
                        PIB_sector_72=PIB_sector_72_float, 
                        PIB_actividades_terciarias=PIB_actividades_terciarias_float, 
                        basura_generada_persona_diaria_Kg=basura_generada_persona_diaria_Kg_float
                    )
                    if inventario_existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {row} ya existe en la base de datos")
                        registros_existentes.append(row)
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        otrosAnuales = otros_anuales(
                            ano=ano_int, 
                            PIB_sector_72=PIB_sector_72_float, 
                            PIB_actividades_terciarias=PIB_actividades_terciarias_float, 
                            basura_generada_persona_diaria_Kg=basura_generada_persona_diaria_Kg_float
                        )
                        otrosAnuales.save()
                        registros_correctos.append(row)
                except (ValueError, TypeError) as e:
                    print(f"Error al procesar la fila {row}: {e}")
                    registros_incorrectos.append(row)
        except FileNotFoundError:
            print(f"No se encontró el archivo {archivo}")
        except Exception as e:
            print(f"Error al procesar el archivo {archivo}: {e}")
        return registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas


class OtrosAnualeDescargarArchivoView(SuperAdminOrAdminMixin, LoginRequiredMixin, View):

    def crear_archivo_excel(self, registros_incorrectos):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Escribir encabezados de columna
        worksheet['A1'] = 'año'
        worksheet['B1'] = 'PIB_sector_72'
        worksheet['C1'] = 'PIB_actividades_terciarias'
        worksheet['D1'] = 'basura_generada_persona_diaria_Kg'

        # Add the incorrect rows to the worksheet
        for i, row in enumerate(registros_incorrectos):
            fila = i + 2
            worksheet.cell(row=fila, column=1, value=row['ano'])
            worksheet.cell(row=fila, column=2, value=row['PIB_sector_72'])
            worksheet.cell(row=fila, column=3, value=row['PIB_actividades_terciarias'])
            worksheet.cell(row=fila, column=4, value=row['basura_generada_persona_diaria_Kg'])



        # Set the column widths to auto-fit
        for column in worksheet.columns:
            max_length = 0
            column_name = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_name].width = adjusted_width

        # Create the response with the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=gasto_derrama_registros_incorrectos.xls'

        

        # workbook.save(response)
        return workbook
    
    def post(self, request, *args, **kwargs):
        # Obtener los registros incorrectos del cuerpo de la petición
        registros_incorrectos = json.loads(request.body)

        # Crear y enviar el archivo de Excel con las filas incorrectas
        workbook = self.crear_archivo_excel(registros_incorrectos)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=gasto_derrama_registros_incorrectos.xlsx'
        workbook.save(response)
        return response   