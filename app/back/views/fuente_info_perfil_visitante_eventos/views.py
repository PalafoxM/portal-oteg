from typing import Any
from django import http
from django.shortcuts import render
from django.http import HttpRequest, JsonResponse, HttpResponseRedirect
from django.urls import reverse_lazy, reverse
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from back.models import *
from back.forms import *
from web.models import *
from django.shortcuts import render
from django.http import HttpResponse
import csv
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_GET
from django.views.decorators.http import require_POST
from django.shortcuts import get_object_or_404
# serializers
# render to string
from django.template.loader import render_to_string
from django.utils.decorators import method_decorator
# para archivo excel
from django.views import View
from django.contrib import messages
from openpyxl import load_workbook
import csv
import os
from datetime import datetime
from django.urls import reverse
import openpyxl
from django.http import HttpResponse
import json
from config.diccionarios import clean_str_col, homologar_columna_categoria, homologar_columna_destino
from django.contrib.auth.decorators import login_required, permission_required
from django.utils.decorators import method_decorator
from django.http import Http404
from django.core.exceptions import PermissionDenied

from django.contrib.auth.decorators import user_passes_test

def es_admin_o_superadmin(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)


def is_ajax(request):
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'


@method_decorator(login_required(login_url='/auth/login_user'), name='dispatch')
@method_decorator(permission_required('auth.view_banner', raise_exception=True), name='dispatch')
@method_decorator(user_passes_test(es_admin_o_superadmin, login_url='404'), name='dispatch')
class FuenteInfoPerfilVisitanteEventos(ListView):
    model = FuenteInfoPerfilVisitanteEvento
    template_name = 'back/fuente_info_perfil_visitante_eventos/viewer.html'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'search':
                data = []
                for i in FuenteInfoPerfilVisitanteEvento.objects.all():
                    data.append(i.toJSON())
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Listado de Fuentes de Información de PV Eventos'
        context['create_url'] = reverse_lazy(
            'dashboard:fuente_info_perfil_visitante_eventos_create')
        context['entity'] = 'Fuentes de Información de PV Eventos'
        context['is_fuente'] = True
        context['carga_masiva_url'] = reverse_lazy('dashboard:fuente_info_perfil_visitante_eventos_carga_masiva')
        return context


@method_decorator(login_required(login_url='/auth/login_user'), name='dispatch')
@method_decorator(permission_required('auth.view_banner', raise_exception=True), name='dispatch')
@method_decorator(user_passes_test(es_admin_o_superadmin, login_url='404'), name='dispatch')
class FuenteInfoPerfilVisitanteEventosCreate (CreateView):
    model = FuenteInfoPerfilVisitanteEvento
    form_class = FuenteInfoPerfilVisitanteEventoForm
    template_name = 'back/fuente_info_perfil_visitante_eventos/create.html'
    success_url = reverse_lazy(
        'dashboard:fuente_info_perfil_visitante_eventos')

    def get_object(self, **kwargs):
        queryset = self.get_queryset()
        try:
            return queryset.get(**kwargs)
        except queryset.model.DoesNotExist:
            return None

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        replace_data = request.POST.get('replace_data')
        if form.is_valid():
            # Check if there is already a record with the same fecha, destino and categoria

            fecha = form.cleaned_data['fecha']
            destino = form.cleaned_data['destino']
            evento = form.cleaned_data['nombre_evento']

            try:
                existing_object = self.get_object(
                    fecha=fecha, destino=destino, nombre_evento=evento)

            except FuenteInfoPerfilVisitanteEvento.DoesNotExist:
                existing_object = None

            existing_catalogo = CatalagoDestino.objects.filter(
                destino=destino).exists()
            # ALTER TABLE mytable MODIFY mycolumn VARCHAR(255) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci;

            # If there is no existing data, save the new data
            if not existing_catalogo:

                if not existing_catalogo:
                    data = {
                        'success': False,
                        'missingData': True,
                        'destino': destino,
                        'message': 'No existe el la entidad en el catalogo',
                    }
                    return JsonResponse(data)

            # If there is existing data and replace_data is True, delete the existing data
            if existing_object and request.POST.get('replace_data') == 'on':
                # existing_object.delete()
                # Save the new data
                # self.object = form.save()

                for field in form.cleaned_data:
                    if field == 'replace_data':
                        continue
                    setattr(existing_object, field, form.cleaned_data[field])

                existing_object.save()

                data = {
                    'success': True,
                    'message': 'Data created successfully.',
                    'url': self.success_url,

                }
                return JsonResponse(data)

            # If there is existing data and replace_data is False, return an error

            if existing_object:
                data = FuenteInfoPerfilVisitanteEvento.objects.filter(
                    fecha=fecha, destino=destino, nombre_evento=evento)
                data_list = list(data.values('ano',
                                             'folio',
                                             'fecha',
                                             'destino',
                                             'nombre_evento',
                                             'segmento',
                                             'tipo_participante',
                                             'residencia',
                                             'tipo_asistente',
                                             'municipio',
                                             'estado',
                                             'pais',
                                             'origen',
                                             'tipo_hospedaje',
                                             'tipo_visitante',
                                             'grupo_viaje',
                                             'acompanantes_maxmin',
                                             'nps_evento',
                                             'nps_evento_categoria',
                                             'edad',
                                             'nse',
                                             'sexo',
                                             'codigo_encuesta_ano'
                                             ))
                data_list2 = list(form.cleaned_data.values())
                table_html = render_to_string('back/fuente_info_perfil_visitante_eventos/table.html', {
                                              'data_list': data_list, 'actual': True, 'data_list2': data_list2})

                datajsn = {
                    'success': False,
                    'message': 'Hubo un error al crear registro.',
                    'errors': 'Ya existe un registro con la misma fecha, destino , origen y museo o zona arqueologica',
                    'existing_object': table_html
                }

                return JsonResponse(datajsn)
            else:
                self.object = form.save()
                data = {
                    'success': True,
                    'message': 'Data created successfully.',
                    'url': self.success_url
                }
                return JsonResponse(data)
        else:
            data = {
                'success': False,
                'message': 'Error creating data.',
                'errors': form.errors,
                'format_errors': True
            }
            return JsonResponse(data)

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Hubo un error al crear registro.',
            'errors': form.errors
        }
        return JsonResponse(data)

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Registro creado exitosamente.',
            'url': self.success_url
        }
        return JsonResponse(data)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Crear una fuente'
        context['entity'] = 'Glosario'
        context['list_url'] = reverse_lazy(
            'dashboard:fuente_info_perfil_visitante_eventos')
        context['action'] = 'add'
        return context


@method_decorator(login_required(login_url='/auth/login_user'), name='dispatch')
@method_decorator(permission_required('auth.view_banner', raise_exception=True), name='dispatch')
@method_decorator(user_passes_test(es_admin_o_superadmin, login_url='404'), name='dispatch')
class FuenteInfoPerfilVisitanteEventosUpdate(UpdateView):
    model = FuenteInfoPerfilVisitanteEvento
    form_class = FuenteInfoPerfilVisitanteEventoForm
    template_name = 'back/fuente_info_perfil_visitante_eventos/create.html'
    success_url = reverse_lazy(
        'dashboard:fuente_info_perfil_visitante_eventos')

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Hubo un error al subir los datos.',
            'errors': form.errors
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Fuente creada exitosamente.',
            'url': self.success_url
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['list_url'] = reverse_lazy(
            'dashboard:fuente_info_perfil_visitante_eventos')
        # Set the widget for the 'destino' field to read-only text input
        context['form'].fields['ano'].widget = forms.TextInput(
            attrs={'readonly': 'readonly'})
        context['form'].fields['folio'].widget = forms.TextInput(
            attrs={'readonly': 'readonly'})
        context['form'].fields['fecha'].widget = forms.TextInput(
            attrs={'readonly': 'readonly'})
        context['form'].fields['destino'].widget = forms.TextInput(
            attrs={'readonly': 'readonly'})
        context['form'].fields['nombre_evento'].widget = forms.TextInput(
            attrs={'readonly': 'readonly'})

        context['title'] = 'Editar fuente'

        context['edit_msg'] = 'Los Campos año , folio, fecha, destino y nombre_evento no se pueden editar'

        return context


@method_decorator(login_required(login_url='/auth/login_user'), name='dispatch')
@method_decorator(permission_required('auth.view_banner', raise_exception=True), name='dispatch')
@method_decorator(user_passes_test(es_admin_o_superadmin, login_url='404'), name='dispatch')
class FuenteInfoPerfilVisitanteEventosDelete(DeleteView):
    model = FuenteInfoPerfilVisitanteEvento
    success_url = reverse_lazy(
        'dashboard:fuente_info_perfil_visitante_eventos')

    def post(self, request: HttpRequest, *args: str, **kwargs: Any) -> HttpResponse:
        return super().post(request, *args, **kwargs)

@method_decorator(login_required(login_url='/auth/login_user'), name='dispatch')
@method_decorator(permission_required('auth.view_banner', raise_exception=True), name='dispatch')
@method_decorator(user_passes_test(es_admin_o_superadmin, login_url='404'), name='dispatch')
class PerfilVisitanteEventosCargaMasivaView(View):
    form_class = CargaMasivaForm
    template_name = 'back/fuente_info_perfil_visitante_eventos/carga_masiva.html'
    success_url = reverse_lazy('dashboard:fuente_info_perfil_visitante_eventos')

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form, 'title': 'Carga Masiva de Perfil Visitante Eventos'})


    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        archivo = request.FILES.get('archivo', None)
        if archivo:
            extension = os.path.splitext(archivo.name)[1]
            if extension == '.xlsx':
                registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas = self.procesar_archivo_xlsx(archivo)
            elif extension == '.csv':
                registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas = self.procesar_archivo_csv(archivo)
            else:
                messages.error(request, 'El archivo debe ser un archivo .xlsx o .csv')
                registros_incorrectos.append("El archivo debe ser un archivo .xlsx o .csv")
        else:
            messages.error(request, 'Debe seleccionar un archivo')
            registros_incorrectos.append("Debe seleccionar un archivo")

        if len(registros_incorrectos) > 0 or len(registros_existentes) > 0:
            messages.error(request, 'Hay errores de registros')
            datos_json = json.dumps(registros_incorrectos)
            
            return render(request, self.template_name, {
                'form': form,
                'title': 'Carga Masiva de Perfil Visitante Eventos',
                'registros_correctos': registros_correctos,
                'registros_incorrectos': registros_incorrectos,
                'registros_existentes': registros_existentes,
                'descargar_url': datos_json,
                'num_filas_procesadas': num_filas_procesadas,
            })
            
        else:
            return HttpResponseRedirect(reverse('dashboard:fuente_info_aerolinea'))
        
    def procesar_archivo_xlsx(self, archivo):
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        try:
            workbook = load_workbook(filename=archivo, read_only=True)
            worksheet = workbook.active
            filas = list(worksheet.rows)
            for i, row in enumerate(filas):
                if i == 0:
                    continue # Ignorar la primera fila si es el encabezado
                num_filas_procesadas += 1
                
                fecha_str = row[2].value.date().strftime('%Y-%m-%d') if len(row) > 2 and row[2].value else ''
                fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date() if fecha_str else ''

                # Limpieza de datos
                destino = clean_str_col(row[3].value)

                # Homologación de datos
                destino = homologar_columna_destino(destino)
                
                ano = row[0].value
                folio = row[1].value
                nombre_evento = row[4].value
                segmento = row[5].value
                tipo_participante = row[6].value
                residencia = row[7].value
                tipo_asistente = row[8].value
                municipio = row[9].value
                estado = row[10].value
                pais = row[11].value
                origen = row[12].value
                tipo_hospedaje = row[13].value
                tipo_visitante = row[14].value
                grupo_viaje = row[15].value
                acompanantes_maxmin = row[16].value
                nps_evento = row[17].value
                nps_evento_categoria = row[18].value
                edad = row[19].value
                nse = row[20].value
                sexo = row[21].value
                codigo_encuesta_ano = row[22].value


                datos = {
                    'ano': ano,
                    'folio': folio,
                    'fecha': fecha_str,
                    'destino': destino,
                    'nombre_evento': nombre_evento,
                    'segmento': segmento,
                    'tipo_participante': tipo_participante,
                    'residencia': residencia,
                    'tipo_asistente': tipo_asistente,
                    'municipio': municipio,
                    'estado': estado,
                    'pais': pais,
                    'origen': origen,
                    'tipo_hospedaje': tipo_hospedaje,
                    'tipo_visitante': tipo_visitante,
                    'grupo_viaje': grupo_viaje,
                    'acompanantes_maxmin': acompanantes_maxmin,
                    'nps_evento': nps_evento,
                    'nps_evento_categoria': nps_evento_categoria,
                    'edad': edad,
                    'nse': nse,
                    'sexo': sexo,
                    'codigo_encuesta_ano': codigo_encuesta_ano
                }

                try:
                    if destino not in CatalagoDestino.objects.values_list('destino', flat=True):
                        print(f"El destino {destino} no está en la tabla CatalagoDestino")
                        registros_incorrectos.append(datos)
                        continue

                    # Buscar si la fila ya existe en la base de datos
                    existente = FuenteInfoPerfilVisitanteEvento.objects.filter(
                        fecha=fecha_obj, 
                        destino=destino, 
                        nombre_evento=nombre_evento
                    )
                    if existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {row} ya existe en la base de datos")
                        registros_existentes.append(datos)
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        db = FuenteInfoPerfilVisitanteEvento(
                            ano=ano,
                            folio=folio,
                            fecha=fecha_obj,
                            destino=destino,
                            nombre_evento=nombre_evento,
                            segmento=segmento,
                            tipo_participante=tipo_participante,
                            residencia=residencia,
                            tipo_asistente=tipo_asistente,
                            municipio=municipio,
                            estado=estado,
                            pais=pais,
                            origen=origen,
                            tipo_hospedaje=tipo_hospedaje,
                            tipo_visitante=tipo_visitante,
                            grupo_viaje=grupo_viaje,
                            acompanantes_maxmin=acompanantes_maxmin,
                            nps_evento=nps_evento,
                            nps_evento_categoria=nps_evento_categoria,
                            edad=edad,
                            nse=nse,
                            sexo=sexo,
                            codigo_encuesta_ano=codigo_encuesta_ano
                        )
                        db.save()
                        registros_correctos.append(datos)
                except (ValueError, TypeError) as e:
                    print(f"Error al procesar la fila {datos}: {e}")
                    registros_incorrectos.append(datos)
                    
        except FileNotFoundError:
                print(f"El archivo {archivo} no se pudo abrir")
                
        return registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas
    
    def procesar_archivo_csv(self, archivo):
        archivo = self.request.FILES['archivo']
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        try:
            datos = csv.DictReader(archivo.read().decode('latin-1').splitlines())
            # print(datos)
            for row in datos:
                num_filas_procesadas += 1

                fecha = row['fecha']
                fecha_str = str(fecha)
                fecha_str = fecha_str.split()[0] if fecha_str else ''  # Eliminar la parte de la hora si existe la fecha
                fecha_obj = datetime.datetime.strptime(fecha_str, '%Y-%m-%d').date()

                # Limpieza de datos
                destino = clean_str_col(row['destino'])

                # Homologación de datos
                destino = homologar_columna_destino(destino)

                ano = row['ano']
                folio = row['folio']
                nombre_evento = row['nombre_evento']
                segmento = row['segmento']
                tipo_participante = row['tipo_participante']
                residencia = row['residencia']
                tipo_asistente = row['tipo_asistente']
                municipio = row['municipio']
                estado = row['estado']
                pais = row['pais']
                origen = row['origen']
                tipo_hospedaje = row['tipo_hospedaje']
                tipo_visitante = row['tipo_visitante']
                grupo_viaje = row['grupo_viaje']
                acompanantes_maxmin = row['acompanantes_maxmin']
                nps_evento = row['nps_evento']
                nps_evento_categoria = row['nps_evento_categoria']
                edad = row['edad']
                nse = row['nse']
                sexo = row['sexo']
                codigo_encuesta_ano = row['codigo_encuesta_ano']


                datos = {
                    'ano': ano,
                    'folio': folio,
                    'fecha': fecha_str,
                    'destino': destino,
                    'nombre_evento': nombre_evento,
                    'segmento': segmento,
                    'tipo_participante': tipo_participante,
                    'residencia': residencia,
                    'tipo_asistente': tipo_asistente,
                    'municipio': municipio,
                    'estado': estado,
                    'pais': pais,
                    'origen': origen,
                    'tipo_hospedaje': tipo_hospedaje,
                    'tipo_visitante': tipo_visitante,
                    'grupo_viaje': grupo_viaje,
                    'acompanantes_maxmin': acompanantes_maxmin,
                    'nps_evento': nps_evento,
                    'nps_evento_categoria': nps_evento_categoria,
                    'edad': edad,
                    'nse': nse,
                    'sexo': sexo,
                    'codigo_encuesta_ano': codigo_encuesta_ano
                }

                try:
                    if destino not in CatalagoDestino.objects.values_list('destino', flat=True):
                        print(f"El destino {destino} no está en la tabla CatalagoDestino")
                        registros_incorrectos.append(datos)
                        continue

                    # Buscar si la fila ya existe en la base de datos
                    existente = FuenteInfoPerfilVisitanteEvento.objects.filter(
                        fecha=fecha_obj, 
                        destino=destino, 
                        nombre_evento=nombre_evento
                    )
                    if existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {row} ya existe en la base de datos")
                        registros_existentes.append(datos)
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        db = FuenteInfoPerfilVisitanteEvento(
                            ano=ano,
                            folio=folio,
                            fecha=fecha,
                            destino=destino,
                            nombre_evento=nombre_evento,
                            segmento=segmento,
                            tipo_participante=tipo_participante,
                            residencia=residencia,
                            tipo_asistente=tipo_asistente,
                            municipio=municipio,
                            estado=estado,
                            pais=pais,
                            origen=origen,
                            tipo_hospedaje=tipo_hospedaje,
                            tipo_visitante=tipo_visitante,
                            grupo_viaje=grupo_viaje,
                            acompanantes_maxmin=acompanantes_maxmin,
                            nps_evento=nps_evento,
                            nps_evento_categoria=nps_evento_categoria,
                            edad=edad,
                            nse=nse,
                            sexo=sexo,
                            codigo_encuesta_ano=codigo_encuesta_ano
                        )
                        db.save()
                        registros_correctos.append(datos)
                except (ValueError, TypeError) as e:
                    print(f"Error al procesar la fila {datos}: {e}")
                    registros_incorrectos.append(datos)
        except FileNotFoundError:
            print(f"No se encontró el archivo {archivo}")
        except Exception as e:
            print(f"Error al procesar el archivo {archivo}: {e}")
        return registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas


@method_decorator(login_required(login_url='/auth/login_user'), name='dispatch')
@method_decorator(permission_required('auth.view_banner', raise_exception=True), name='dispatch')
@method_decorator(user_passes_test(es_admin_o_superadmin, login_url='404'), name='dispatch')
class PerfilVisitanteEventosDescargarArchivoView(View):

    def crear_archivo_excel(self, registros_incorrectos):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Escribir encabezados de columna
        worksheet['A1'] = 'Año'
        worksheet['B1'] = 'Folio'
        worksheet['C1'] = 'Fecha'
        worksheet['D1'] = 'Destino'
        worksheet['E1'] = 'Nombre del Evento'
        worksheet['F1'] = 'Segmento'
        worksheet['G1'] = 'Tipo de Participante'
        worksheet['H1'] = 'Residencia'
        worksheet['I1'] = 'Tipo de Asistente'
        worksheet['J1'] = 'Municipio'
        worksheet['K1'] = 'Estado'
        worksheet['L1'] = 'País'
        worksheet['M1'] = 'Origen'
        worksheet['N1'] = 'Tipo de Hospedaje'
        worksheet['O1'] = 'Tipo de Visitante'
        worksheet['P1'] = 'Grupo de Viaje'
        worksheet['Q1'] = 'Acompañantes MaxMin'
        worksheet['R1'] = 'NPS del Evento'
        worksheet['S1'] = 'Categoría del NPS del Evento'
        worksheet['T1'] = 'Edad'
        worksheet['U1'] = 'NSE'
        worksheet['V1'] = 'Sexo'
        worksheet['W1'] = 'Código de Encuesta Año'

        # Add the incorrect rows to the worksheet
        for i, row in enumerate(registros_incorrectos):
            fila = i + 2
            worksheet.cell(row=fila, column=1, value=row['ano'])
            worksheet.cell(row=fila, column=2, value=row['folio'])
            worksheet.cell(row=fila, column=3, value=row['fecha'])
            worksheet.cell(row=fila, column=4, value=row['destino'])
            worksheet.cell(row=fila, column=5, value=row['nombre_evento'])
            worksheet.cell(row=fila, column=6, value=row['segmento'])
            worksheet.cell(row=fila, column=7, value=row['tipo_participante'])
            worksheet.cell(row=fila, column=8, value=row['residencia'])
            worksheet.cell(row=fila, column=9, value=row['tipo_asistente'])
            worksheet.cell(row=fila, column=10, value=row['municipio'])
            worksheet.cell(row=fila, column=11, value=row['estado'])
            worksheet.cell(row=fila, column=12, value=row['pais'])
            worksheet.cell(row=fila, column=13, value=row['origen'])
            worksheet.cell(row=fila, column=14, value=row['tipo_hospedaje'])
            worksheet.cell(row=fila, column=15, value=row['tipo_visitante'])
            worksheet.cell(row=fila, column=16, value=row['grupo_viaje'])
            worksheet.cell(row=fila, column=17, value=row['acompanantes_maxmin'])
            worksheet.cell(row=fila, column=18, value=row['nps_evento'])
            worksheet.cell(row=fila, column=19, value=row['nps_evento_categoria'])
            worksheet.cell(row=fila, column=20, value=row['edad'])
            worksheet.cell(row=fila, column=21, value=row['nse'])
            worksheet.cell(row=fila, column=22, value=row['sexo'])
            worksheet.cell(row=fila, column=23, value=row['codigo_encuesta_ano'])



        # Set the column widths to auto-fit
        for column in worksheet.columns:
            max_length = 0
            column_name = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_name].width = adjusted_width

        # Create the response with the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=gasto_derrama_registros_incorrectos.xls'

        

        # workbook.save(response)
        return workbook
    
    def post(self, request, *args, **kwargs):
        # Obtener los registros incorrectos del cuerpo de la petición
        registros_incorrectos = json.loads(request.body)

        # Crear y enviar el archivo de Excel con las filas incorrectas
        workbook = self.crear_archivo_excel(registros_incorrectos)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=gasto_derrama_registros_incorrectos.xlsx'
        workbook.save(response)
        return response       