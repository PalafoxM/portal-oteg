from django.shortcuts import render
from typing import Any, Dict
from django.http import HttpRequest, JsonResponse, HttpResponseRedirect
from django.urls import reverse_lazy, reverse
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from back.models import *
from back.forms import *
from web.models import *
from django.shortcuts import render
from django.http import HttpResponse
import csv
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_GET
from django.views.decorators.http import require_POST
from django.shortcuts import get_object_or_404
#serializers
# render to string
from django.template.loader import render_to_string
# para archivo excel
from django.views import View
from django.contrib import messages
from openpyxl import load_workbook
import csv
import os
from datetime import datetime
from django.urls import reverse
import openpyxl
from django.http import HttpResponse
import json
from config.diccionarios import clean_str_col, homologar_columna_categoria, homologar_columna_destino
from django.contrib.auth.decorators import login_required, permission_required
from django.utils.decorators import method_decorator
from django.http import Http404
from django.core.exceptions import PermissionDenied
from back.mixins import *
from django.contrib.auth.decorators import user_passes_test

def es_admin_o_superadmin(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)


def is_ajax(request):
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'


class FuenteInfoDiscapacidad (SuperAdminOrAdminMixin, LoginRequiredMixin, ListView):
    model = Discapacidad
    template_name = 'back/fuente_info_discapacidad/list.html'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'search':
                data = []
                for i in Discapacidad.objects.all():
                    data.append(i.toJSON())
            else:
                data.append({'error': 'Ha ocurrido un error'})
        except Exception as e:
            data.append({'error': str(e)})
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Listado de Fuentes de Informacion de Empleo Discapacidad'
        context['create_url'] = reverse_lazy('dashboard:fuente_info_discapacidad_create')
        context['entity'] = 'Empleo'
        context['is_fuente'] = True
        context['carga_masiva_url'] = reverse_lazy('dashboard:fuente_info_discapacidad_carga_masiva')

        return context


class FuenteInfoDiscapacidadCreate (SuperAdminOrAdminMixin, LoginRequiredMixin, CreateView):
    model = Discapacidad
    form_class = DiscapacidadForm
    template_name = 'back/fuente_info_discapacidad/create.html'
    success_url = reverse_lazy('dashboard:fuente_info_discapacidad')

    def get_object(self, **kwargs):
        queryset = self.get_queryset()
        try:
            return queryset.get(**kwargs)
        except queryset.model.DoesNotExist:
            return None
    
    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)

        if form.is_valid():
            # Check if there is already a record with the same fecha, destino and categoria
   
            fecha_inicio = form.cleaned_data['fecha_inicio']
            fecha_fin = form.cleaned_data['fecha_fin']
            try:
                existing_object = self.get_object( fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)

            except inversion_privada.DoesNotExist:
                existing_object = None

            # If there is existing data and replace_data is True, delete the existing data
            if existing_object and request.POST.get('replace_data') == 'on':
                # existing_object.delete()
                # Save the new data
                # self.object = form.save()

                for field in form.cleaned_data:
                    if field == 'replace_data':
                        continue
                    setattr(existing_object, field, form.cleaned_data[field])
                existing_object.save()

                data = {
                    'success': True,
                    'message': 'Data created successfully.',
                    'url': self.success_url,
                }
                return JsonResponse(data)

            # If there is existing data and replace_data is False, return an error

            if existing_object:
                data = empleo.objects.filter(fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)

                data_list = list(data.values('fecha_inicio', 'fecha_fin', 'hombres_empleados_gto', 'mujeres_empleadas_gto', 'hombres_empleados_sec_72_gto','mujeres_empleadas_sec_72_gto','hombres_empleados_sec_72_nac','mujeres_empleadas_sec_72_nac'))
                data_list2 = list(form.cleaned_data.values())

                table_html = render_to_string('back/fuente_info_discapacidad/table.html', {'data_list': data_list, 'actual': True, 'data_list2': data_list2})

                datajsn = {
                    'success': False,
                    'message': 'Hubo un error al crear registro.',
                    'errors': 'Ya existe un registro con la misma fecha, destino , origen y museo o zona arqueologica',
                    'existing_object': table_html
                }

                return JsonResponse(datajsn)
            else:

                self.object = form.save()
                data = {
                    'success': True,
                    'message': 'Data created successfully.',
                    'url': self.success_url
                }
                return JsonResponse(data)
        else:
            data = {
                'success': False,
                'message': 'Error creating data.',
                'errors': form.errors,
                'format_errors': True
            }
            return JsonResponse(data)

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Hubo un error al crear registro.',
            'errors': form.errors
        }
        return JsonResponse(data)

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Registro creado exitosamente.',
            'url': self.success_url
        }
        return JsonResponse(data)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Crear una fuente'
        context['entity'] = 'Empleo'
        context['list_url'] = reverse_lazy('dashboard:fuente_info_discapacidad')
        context['action'] = 'add'
        return context


class FuenteInfoDiscapacidadUpdate (SuperAdminOrAdminMixin, LoginRequiredMixin, UpdateView):
    model = Discapacidad
    form_class = DiscapacidadForm
    template_name = 'back/fuente_info_discapacidad/view_editor.html'
    success_url = reverse_lazy('dashboard:fuente_info_discapacidad')

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Hubo un error al subir los datos.',
            'errors': form.errors
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Fuente creada exitosamente.',
            'url': self.success_url
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['list_url'] = reverse_lazy('dashboard:fuente_info_discapacidad')
            # Set the widget for the 'destino' field to read-only text input
        context['title'] = 'Editar fuente'
        context['edit_msg'] = 'Los campos fecha inicio y fecha fin no son editables'

        return context


class FuenteInfoDiscapacidadDelete (SuperAdminOrAdminMixin, LoginRequiredMixin, DeleteView):
    model = Discapacidad
    success_url = reverse_lazy('dashboard:fuente_info_discapacidad')

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()

        try:
            self.object.delete()
            return JsonResponse({'message': 'Eliminación exitosa.'})
        except Exception as e:
            return JsonResponse({'error': 'Error al eliminar el registro.'}, status=500)


class DiscapacidadCargaMasivaView(SuperAdminOrAdminMixin, LoginRequiredMixin, View):
    form_class = CargaMasivaForm
    template_name = 'back/fuente_info_discapacidad/carga_masiva.html'
    success_url = reverse_lazy('dashboard:fuente_info_discapacidad')

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form, 'title': 'Carga Masiva de Empleo Discapacidad'})


    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        archivo = request.FILES.get('archivo', None)
        if archivo:
            extension = os.path.splitext(archivo.name)[1]
            if extension == '.xlsx':
                registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas = self.procesar_archivo_xlsx(archivo)
            elif extension == '.csv':
                registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas = self.procesar_archivo_csv(archivo)
            else:
                messages.error(request, 'El archivo debe ser un archivo .xlsx o .csv')
                registros_incorrectos.append("El archivo debe ser un archivo .xlsx o .csv")
        else:
            messages.error(request, 'Debe seleccionar un archivo')
            registros_incorrectos.append("Debe seleccionar un archivo")

        if len(registros_incorrectos) > 0 or len(registros_existentes) > 0:
            messages.error(request, 'Hay errores de registros')
            datos_json = json.dumps(registros_incorrectos)
            
            return render(request, self.template_name, {
                'form': form,
                'title': 'Carga Masiva de Empleo Discapacidad',
                'registros_correctos': registros_correctos,
                'registros_incorrectos': registros_incorrectos,
                'registros_existentes': registros_existentes,
                'descargar_url': datos_json,
                'num_filas_procesadas': num_filas_procesadas,
            })
            
        else:
            return HttpResponseRedirect(reverse('dashboard:fuente_info_discapacidad'))
        
        

    def procesar_archivo_xlsx(self, archivo):
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        try:
            workbook = load_workbook(filename=archivo, read_only=True)
            worksheet = workbook.active
            filas = list(worksheet.rows)
            for i, row in enumerate(filas):
                if i == 0:
                    continue # Ignorar la primera fila si es el encabezado

                if not row or all(cell.value is None for cell in row):
                    continue  # Salta filas vacías

                num_filas_procesadas += 1
                # Limpieza de datos
                fecha_str = row[1].value.date().strftime('%d-%m-%Y') if len(row) > 0 and row[0].value else ''
                fecha_obj = datetime.strptime(fecha_str, '%d-%m-%Y').date() if fecha_str else ''

                giro_comercial           = row[2].value if len(row) > 2 and row[2].value else 0
                empleos_fijos_h          = row[3].value if len(row) > 3 and row[3].value else 0
                empleos_fijos_m          = row[4].value if len(row) > 4 and row[4].value else 0
                empleos_temporales_h     = row[5].value if len(row) > 5 and row[5].value else 0
                empleos_temporales_m     = row[6].value if len(row) > 6 and row[6].value else 0
                empleados_discapacidad_h = row[7].value if len(row) > 7 and row[7].value else 0
                empleados_discapacidad_m = row[8].value if len(row) > 8 and row[8].value else 0

                # Limpieza de datos
                destino = clean_str_col(row[0].value)

                # Homologación de datos
                destino = homologar_columna_destino(destino)          
                
                datos = {
                    "destino": destino,
                    "fecha": fecha_str,
                    "giro_comercial": giro_comercial,
                    "empleos_fijos_h": empleos_fijos_h,
                    "empleos_fijos_m": empleos_fijos_m,
                    "empleos_temporales_h": empleos_temporales_h,
                    "empleos_temporales_m": empleos_temporales_m,
                    "empleados_discapacidad_h": empleados_discapacidad_h,
                    "empleados_discapacidad_m": empleados_discapacidad_m,
                }

                try:
                    # Validar si el destino y categoria son válidos
                    if destino not in CatalagoDestino.objects.values_list('destino', flat=True):
                        print(f"El destino {destino} no está en la tabla CatalagoDestino")
                        registros_incorrectos.append(datos)
                        continue

                    # Buscar si la fila ya existe en la base de datos
                    existente = Discapacidad.objects.filter(
                        fecha = fecha_obj, 
                        destino = destino,
                        giro_comercial = giro_comercial
                        )
                    if existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {datos} ya existe en la base de datos")
                        registros_existentes.append(datos)
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        db = Discapacidad(
                            fecha = fecha_obj,
                            destino = destino,
                            giro_comercial = giro_comercial,
                            empleos_fijos_h = empleos_fijos_h,
                            empleos_fijos_m = empleos_fijos_m,
                            empleos_temporales_h = empleos_temporales_h,
                            empleos_temporales_m = empleos_temporales_m,
                            empleados_discapacidad_h = empleados_discapacidad_h,
                            empleados_discapacidad_m = empleados_discapacidad_m,
                        )
                        db.save()
                        registros_correctos.append(datos)
                except (ValueError, TypeError) as e:
                    print(f"Error al procesar la fila {datos}: {e}")
                    registros_incorrectos.append(datos)
                    
        except FileNotFoundError:
                print(f"El archivo {archivo} no se pudo abrir")
                
        return registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas
    
    def procesar_archivo_csv(self, archivo):
        archivo = self.request.FILES['archivo']
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        try:
            datos = csv.DictReader(archivo.read().decode('latin-1').splitlines())
            # print(datos)
            for row in datos:
                num_filas_procesadas += 1

                fecha = row['fecha']
                fecha_str = str(fecha)
                fecha_str = fecha_str.split()[0] if fecha_str else ''  # Eliminar la parte de la hora si existe la fecha
                fecha_obj = datetime.datetime.strptime(fecha_str, '%Y-%m-%d').date()

                giro_comercial = row['giro_comercial']
                empleos_fijos_h = row['empleos_fijos_h']
                empleos_fijos_m = row['empleos_fijos_m']
                empleos_temporales_h = row['empleos_temporales_h']
                empleos_temporales_m = row['empleos_temporales_m']
                empleados_discapacidad_h = row['empleados_discapacidad_h']
                empleados_discapacidad_m = row['empleados_discapacidad_m']

                # Limpieza de datos
                destino = clean_str_col(row['destino'])

                # Homologación de datos
                destino = homologar_columna_destino(destino)          
                
                datos = {
                    "destino": destino,
                    "fecha": fecha_str,
                    "giro_comercial": giro_comercial,
                    "empleos_fijos_h": empleos_fijos_h,
                    "empleos_fijos_m": empleos_fijos_m,
                    "empleos_temporales_h": empleos_temporales_h,
                    "empleos_temporales_m": empleos_temporales_m,
                    "empleados_discapacidad_h": empleados_discapacidad_h,
                    "empleados_discapacidad_m": empleados_discapacidad_m,
                }

                try:
                    # Validar si el destino y categoria son válidos
                    if destino not in CatalagoDestino.objects.values_list('destino', flat=True):
                        print(f"El destino {destino} no está en la tabla CatalagoDestino")
                        registros_incorrectos.append(datos)
                        continue

                    # Buscar si la fila ya existe en la base de datos
                    existente = Discapacidad.objects.filter(
                        fecha = fecha_obj, 
                        destino = destino,
                        giro_comercial = giro_comercial
                        )
                    if existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {datos} ya existe en la base de datos")
                        registros_existentes.append(datos)
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        db = Discapacidad(
                            fecha = fecha_obj,
                            destino = destino,
                            giro_comercial = giro_comercial,
                            empleos_fijos_h = empleos_fijos_h,
                            empleos_fijos_m = empleos_fijos_m,
                            empleos_temporales_h = empleos_temporales_h,
                            empleos_temporales_m = empleos_temporales_m,
                            empleados_discapacidad_h = empleados_discapacidad_h,
                            empleados_discapacidad_m = empleados_discapacidad_m,
                        )
                        db.save()
                        registros_correctos.append(datos)
                except (ValueError, TypeError) as e:
                    print(f"Error al procesar la fila {datos}: {e}")
                    registros_incorrectos.append(datos)
        except FileNotFoundError:
            print(f"No se encontró el archivo {archivo}")
        except Exception as e:
            print(f"Error al procesar el archivo {archivo}: {e}")
        return registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas


class DiscapacidadDescargarArchivoView(SuperAdminOrAdminMixin, LoginRequiredMixin, View):

    def crear_archivo_excel(self, registros_incorrectos):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Escribir encabezados de columna
        worksheet['A1'] = "destino"
        worksheet['B1'] = "fecha"
        worksheet['C1'] = "giro_comercial"
        worksheet['D1'] = "empleos_fijos_h"
        worksheet['E1'] = "empleos_fijos_m"
        worksheet['F1'] = "empleos_temporales_h"
        worksheet['G1'] = "empleos_temporales_m"
        worksheet['H1'] = "empleados_discapacidad_h"
        worksheet['I1'] = "empleados_discapacidad_m"

        # Add the incorrect rows to the worksheet
        for i, row in enumerate(registros_incorrectos):
            fila = i + 2
            worksheet.cell(row=fila, column=1, value=row['destino'])
            worksheet.cell(row=fila, column=2, value=row['fecha'])
            worksheet.cell(row=fila, column=3, value=row['giro_comercial'])
            worksheet.cell(row=fila, column=4, value=row['empleos_fijos_h'])
            worksheet.cell(row=fila, column=5, value=row['empleos_fijos_m'])
            worksheet.cell(row=fila, column=6, value=row['empleos_temporales_h'])
            worksheet.cell(row=fila, column=7, value=row['empleos_temporales_m'])
            worksheet.cell(row=fila, column=8, value=row['empleados_discapacidad_h'])
            worksheet.cell(row=fila, column=9, value=row['empleados_discapacidad_m'])



        # Set the column widths to auto-fit
        for column in worksheet.columns:
            max_length = 0
            column_name = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_name].width = adjusted_width

        # Create the response with the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=sensibilizacion_registros_incorrectos.xls'

        

        # workbook.save(response)
        return workbook
    
    def post(self, request, *args, **kwargs):
        # Obtener los registros incorrectos del cuerpo de la petición
        registros_incorrectos = json.loads(request.body)

        # Crear y enviar el archivo de Excel con las filas incorrectas
        workbook = self.crear_archivo_excel(registros_incorrectos)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=sensibilizacion_registros_incorrectos.xlsx'
        workbook.save(response)
        return response   