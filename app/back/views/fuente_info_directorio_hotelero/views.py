from typing import Any, Dict
from django.shortcuts import render
from django.http import HttpRequest, JsonResponse, HttpResponseRedirect
from django.urls import reverse_lazy, reverse
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from back.models import *
from back.forms import *
from web.models import *
from django.http import HttpResponse
from django.contrib.auth.mixins import LoginRequiredMixin
from django.template.loader import render_to_string
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime, date

# para archivo excel
from django.views import View
from django.contrib import messages
from openpyxl import load_workbook
import csv
import os
from django.urls import reverse
import openpyxl
from django.http import HttpResponse
import json
from config.diccionarios import clean_str_col, homologar_columna_categoria, homologar_columna_destino


def is_ajax(request):
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'


class FuenteInfoDirectorioHotelero (ListView):
    model = DirectorioHotelero
    template_name = 'back/fuente_info_directorio_hotelero/list.html'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'search':
                data = []
                for i in DirectorioHotelero.objects.all():
                    data.append(i.toJSON())
            else:
                data.append({'error': 'Ha ocurrido un error'})
        except Exception as e:
            data.append({'error': str(e)})
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Listado  de Directorio Hotelero'
        context['create_url'] = reverse_lazy('dashboard:fuente_info_directorio_hotelero_create')
        context['entity'] = 'Fuentes de Informacion de DirectorioHotelero'
        context['is_fuente'] = True
        #urls
        context['delete_url'] = reverse_lazy('dashboard:fuente_info_directorio_hotelero')
        context['update_url'] =  reverse('dashboard:fuente_info_certificacion_delete')  # assuming 'e' is available in your context
        context['carga_masiva_url'] = reverse_lazy('dashboard:fuente_info_directorio_hotelero_carga_masiva')

        return context


class FuenteInfoDirectorioHoteleroCreate (CreateView):
    model = DirectorioHotelero
    form_class = DirectorioHoteleroForm
    template_name = 'back/fuente_info_directorio_hotelero/create.html'
    success_url = reverse_lazy('dashboard:fuente_info_directorio_hotelero')

    def get_object(self, **kwargs):
        queryset = self.get_queryset()
        try:
            return queryset.get(**kwargs)
        except queryset.model.DoesNotExist:
            return None

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        replace_data = request.POST.get('replace_data')
        if form.is_valid():
            # Check if there is already a record with the same fecha, destino and categoria

            nombre_de_la_unidad_economica = form.cleaned_data['nombre_de_la_unidad_economica']
            id_establecimiento = form.cleaned_data['id_establecimiento']

            try:
                existing_object = self.get_object(
                    nombre_de_la_unidad_economica=nombre_de_la_unidad_economica, id_establecimiento=id_establecimiento)

            except DirectorioHotelero.DoesNotExist:
                existing_object = None

            # If there is existing data and replace_data is True, delete the existing data
            if existing_object and request.POST.get('replace_data') == 'on':
                # existing_object.delete()
                # Save the new data
                # self.object = form.save()

                for field in form.cleaned_data:
                    if field == 'replace_data':
                        continue
                    setattr(existing_object, field, form.cleaned_data[field])

                existing_object.save()

                data = {
                    'success': True,
                    'message': 'Data created successfully.',
                    'url': self.success_url,

                }
                return JsonResponse(data)

            # If there is existing data and replace_data is False, return an error

            if existing_object:
                data = DirectorioHotelero.objects.filter(
                    nombre_de_la_unidad_economica=nombre_de_la_unidad_economica, id_establecimiento=id_establecimiento)

                data_list = list(data.values(
                    'id_establecimiento',
                    'nombre_de_la_unidad_economica',
                    'razon_social',
                    'codigo_de_la_clase_de_actividad_scian',
                    'nombre_de_clase_de_la_actividad',
                    'descripcion_estrato_personal_ocupado',
                    'tipo_de_vialidad',
                    'nombre_de_la_vialidad',
                    'tipo_de_entre_vialidad_1',
                    'nombre_de_entre_vialidad_1',
                    'tipo_de_entre_vialidad_2',
                    'nombre_de_entre_vialidad_2',
                    'tipo_de_entre_vialidad_3',
                    'nombre_de_entre_vialidad_3',
                    'numero_exterior_o_kilometro',
                    'letra_exterior',
                    'edificio',
                    'edificio_piso',
                    'numero_interior',
                    'letra_interior',
                    'tipo_de_asentamiento_humano',
                    'nombre_de_asentamiento_humano',
                    'tipo_centro_comercial',
                    'c_industrial_comercial_o_mercado',
                    'numero_de_local',
                    'codigo_postal',
                    'clave_entidad',
                    'entidad_federativa',
                    'clave_municipio',
                    'municipio',
                    'clave_localidad',
                    'localidad',
                    'area_geoestadistica_basica',
                    'manzana',
                    'numero_de_telefono',
                    'correo_electronico',
                    'sitio_en_internet',
                    'tipo_de_establecimiento',
                    'latitud',
                    'longitud',
                    'fecha_de_incorporacion_al_denue',
                    'categoria_turistica',
                    'no_cuartos',
                    'unidades',
                    'espacios_cajones',
                    'no_camas',
                    'cadena',
                    'operadora',
                    'responsable',
                    'cargo',
                    'imss',
                    'inicio_de_operaciones_este_ano',
                    'fecha_de_inicio_de_operaciones',
                    'centro_turistico',
                    'zona',
                    'datatur',
                    'hotel_boutique',
                    'nombre_de_la_cadena',
                    'hoteles_tesoros',

                ))

                fields_list = [
                    'id_establecimiento',
                    'nombre_de_la_unidad_economica',
                    'razon_social',
                    'codigo_de_la_clase_de_actividad_scian',
                    'nombre_de_clase_de_la_actividad',
                    'descripcion_estrato_personal_ocupado',
                    'tipo_de_vialidad',
                    'nombre_de_la_vialidad',
                    'tipo_de_entre_vialidad_1',
                    'nombre_de_entre_vialidad_1',
                    'tipo_de_entre_vialidad_2',
                    'nombre_de_entre_vialidad_2',
                    'tipo_de_entre_vialidad_3',
                    'nombre_de_entre_vialidad_3',
                    'numero_exterior_o_kilometro',
                    'letra_exterior',
                    'edificio',
                    'edificio_piso',
                    'numero_interior',
                    'letra_interior',
                    'tipo_de_asentamiento_humano',
                    'nombre_de_asentamiento_humano',
                    'tipo_centro_comercial',
                    'c_industrial_comercial_o_mercado',
                    'numero_de_local',
                    'codigo_postal',
                    'clave_entidad',
                    'entidad_federativa',
                    'clave_municipio',
                    'municipio',
                    'clave_localidad',
                    'localidad',
                    'area_geoestadistica_basica',
                    'manzana',
                    'numero_de_telefono',
                    'correo_electronico',
                    'sitio_en_internet',
                    'tipo_de_establecimiento',
                    'latitud',
                    'longitud',
                    'fecha_de_incorporacion_al_denue',
                    'categoria_turistica',
                    'no_cuartos',
                    'unidades',
                    'espacios_cajones',
                    'no_camas',
                    'cadena',
                    'operadora',
                    'responsable',
                    'cargo',
                    'imss',
                    'inicio_de_operaciones_este_ano',
                    'fecha_de_inicio_de_operaciones',
                    'centro_turistico',
                    'zona',
                    'datatur',
                    'hotel_boutique',
                    'nombre_de_la_cadena',
                    'hoteles_tesoros',
                ]

                table_headers = ''.join(
                    f'<th style="width: 300px;">{field}</th>' for field in fields_list)

                data_list2 = list(form.cleaned_data.values())

                table_html = render_to_string('back/fuente_info_directorio_hotelero/table.html', {
                                              'data_list': data_list, 'actual': True, 'data_list2': data_list2, "table_headers": table_headers})

                datajsn = {
                    'success': False,
                    'message': 'Hubo un error al crear registro.',
                    'errors': 'Ya existe un registro con la misma fecha, destino , origen y museo o zona arqueologica',
                    'existing_object': table_html
                }

                return JsonResponse(datajsn)
            else:
                self.object = form.save()
                data = {
                    'success': True,
                    'message': 'Data created successfully.',
                    'url': self.success_url
                }
                return JsonResponse(data)
        else:
            data = {
                'success': False,
                'message': 'Error creating data.',
                'errors': form.errors,
                'format_errors': True
            }
            return JsonResponse(data)

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Hubo un error al crear registro.',
            'errors': form.errors
        }
        return JsonResponse(data)

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Registro creado exitosamente.',
            'url': self.success_url
        }
        return JsonResponse(data)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Crear una fuente'
        context['entity'] = 'Glosario'
        context['list_url'] = reverse_lazy(
            'dashboard:fuente_info_directorio_hotelero')
        context['action'] = 'add'
        return context


class FuenteInfoDirectorioHoteleroUpdate (UpdateView):
    model = DirectorioHotelero
    form_class = DirectorioHoteleroForm
    template_name = 'back/fuente_info_directorio_hotelero/view_editor.html'
    success_url = reverse_lazy('dashboard:fuente_info_directorio_hotelero')

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Hubo un error al subir los datos.',
            'errors': form.errors
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Fuente creada exitosamente.',
            'url': self.success_url
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['list_url'] = reverse_lazy(
            'dashboard:fuente_info_directorio_hotelero')
        # Set the widget for the 'destino' field to read-only text input
        # context['form'].fields['ano'].widget = forms.TextInput(attrs={'class': 'form-control', 'readonly': 'readonly'})
        # context['form'].fields['giro'].widget = forms.TextInput(attrs={'class': 'form-control', 'readonly': 'readonly'})
        # context['form'].fields['destino'].widget = forms.TextInput(attrs={'class': 'form-control', 'readonly': 'readonly'})

        context['title'] = 'Editar fuente'
        context['edit_msg'] = 'Los Campos Destino y Año no pueden ser editados'

        return context


class FuenteInfoDirectorioHoteleroDelete (DeleteView):
    model = DirectorioHotelero
    success_url = reverse_lazy('dashboard:fuente_info_directorio_hotelero')

    def post(self, request: HttpRequest, *args: str, **kwargs: Any) -> HttpResponse:
        return super().post(request, *args, **kwargs)


class DirectorioHoteleroCargaMasivaView(View):
    form_class = CargaMasivaForm
    template_name = 'back/fuente_info_directorio_hotelero/carga_masiva.html'
    success_url = reverse_lazy('dashboard:fuente_info_directorio_hotelero')

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form, 'title': 'Carga Masiva de DirectorioHotelero'})

    def convert_to_serializable(self, obj):
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        raise TypeError(
            f'Object of type {obj.__class__.__name__} is not JSON serializable')

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        archivo = request.FILES.get('archivo', None)
        if archivo:
            extension = os.path.splitext(archivo.name)[1]
            if extension == '.xlsx':
                registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas = self.procesar_archivo_xlsx(
                    archivo)
            elif extension == '.csv':
                registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas = self.procesar_archivo_csv(
                    archivo)
            else:
                messages.error(
                    request, 'El archivo debe ser un archivo .xlsx o .csv')
                registros_incorrectos.append(
                    "El archivo debe ser un archivo .xlsx o .csv")
        else:
            messages.error(request, 'Debe seleccionar un archivo')
            registros_incorrectos.append("Debe seleccionar un archivo")

        if len(registros_incorrectos) > 0 or len(registros_existentes) > 0:
            messages.error(request, 'Hay errores de registros')
            datos_json = json.dumps(
                registros_incorrectos, default=self.convert_to_serializable)

            return render(request, self.template_name, {
                'form': form,
                'title': 'Carga Masiva de DirectorioHotelero',
                'registros_correctos': registros_correctos,
                'registros_incorrectos': registros_incorrectos,
                'registros_existentes': registros_existentes,
                'descargar_url': datos_json,
                'num_filas_procesadas': num_filas_procesadas,
            })

        else:
            return HttpResponseRedirect(reverse('dashboard:fuente_info_directorio_hotelero'))

    def procesar_archivo_xlsx(self, archivo):
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        try:
            workbook = load_workbook(filename=archivo, read_only=True)
            worksheet = workbook.active
            filas = list(worksheet.rows)
            for i, row in enumerate(filas):
                if i == 0:
                    continue  # Ignorar la primera fila si es el encabezado
                num_filas_procesadas += 1

                id_establecimiento = row[0].value
                nombre_de_la_unidad_economica = row[1].value
                razon_social = row[2].value
                codigo_de_la_clase_de_actividad_scian = row[3].value
                nombre_de_clase_de_la_actividad = row[4].value
                descripcion_estrato_personal_ocupado = row[5].value
                tipo_de_vialidad = row[6].value
                nombre_de_la_vialidad = row[7].value
                tipo_de_entre_vialidad_1 = row[8].value
                nombre_de_entre_vialidad_1 = row[9].value
                tipo_de_entre_vialidad_2 = row[10].value
                nombre_de_entre_vialidad_2 = row[11].value
                tipo_de_entre_vialidad_3 = row[12].value
                nombre_de_entre_vialidad_3 = row[13].value
                numero_exterior_o_kilometro = row[14].value
                letra_exterior = row[15].value
                edificio = row[16].value
                edificio_piso = row[17].value
                numero_interior = row[18].value
                letra_interior = row[19].value
                tipo_de_asentamiento_humano = row[20].value
                nombre_de_asentamiento_humano = row[21].value
                tipo_centro_comercial = row[22].value
                c_industrial_comercial_o_mercado = row[23].value
                numero_de_local = row[24].value
                codigo_postal = row[25].value
                clave_entidad = row[26].value
                entidad_federativa = row[27].value
                clave_municipio = row[28].value
                municipio = row[29].value
                clave_localidad = row[30].value
                localidad = row[31].value
                area_geoestadistica_basica = row[32].value
                manzana = row[33].value
                numero_de_telefono = row[34].value
                correo_electronico = row[35].value
                sitio_en_internet = row[36].value
                tipo_de_establecimiento = row[37].value
                latitud = row[38].value
                longitud = row[39].value
                fecha_de_incorporacion_al_denue = row[40].value
                categoria_turistica = row[41].value
                no_cuartos = row[42].value
                unidades = row[43].value
                espacios_cajones = row[44].value
                no_camas = row[45].value
                cadena = row[46].value
                operadora = row[47].value
                responsable = row[48].value
                cargo = row[49].value
                imss = row[50].value
                inicio_de_operaciones_este_ano = row[51].value
                fecha_de_inicio_de_operaciones = row[52].value
                centro_turistico = row[53].value
                zona = row[54].value
                datatur = row[55].value
                hotel_boutique = row[56].value
                nombre_de_la_cadena = row[57].value
                hoteles_tesoros = row[58].value

                # Verificar si fecha_de_incorporacion_al_denue es una instancia de datetime
                if isinstance(fecha_de_incorporacion_al_denue, datetime):
                    fecha_de_incorporacion_al_denue_str = fecha_de_incorporacion_al_denue.strftime(
                        '%Y-%m-%d')
                    fecha_de_incorporacion_al_denue_obj = datetime.strptime(
                        fecha_de_incorporacion_al_denue_str, '%Y-%m-%d').date()
                else:
                    fecha_de_incorporacion_al_denue_str = ''
                    fecha_de_incorporacion_al_denue_obj = None

                # Verificar si fecha_de_inicio_de_operaciones es una instancia de datetime
                if isinstance(fecha_de_inicio_de_operaciones, datetime):
                    fecha_de_inicio_de_operaciones_str = fecha_de_inicio_de_operaciones.strftime(
                        '%Y-%m-%d')
                    fecha_de_inicio_de_operaciones_obj = datetime.strptime(
                        fecha_de_inicio_de_operaciones_str, '%Y-%m-%d').date()
                else:
                    fecha_de_inicio_de_operaciones_str = ''
                    fecha_de_inicio_de_operaciones_obj = None

                datos = {
                    "id_establecimiento": id_establecimiento,
                    "nombre_de_la_unidad_economica": nombre_de_la_unidad_economica,
                    "razon_social": razon_social,
                    "codigo_de_la_clase_de_actividad_scian": codigo_de_la_clase_de_actividad_scian,
                    "nombre_de_clase_de_la_actividad": nombre_de_clase_de_la_actividad,
                    "descripcion_estrato_personal_ocupado": descripcion_estrato_personal_ocupado,
                    "tipo_de_vialidad": tipo_de_vialidad,
                    "nombre_de_la_vialidad": nombre_de_la_vialidad,
                    "tipo_de_entre_vialidad_1": tipo_de_entre_vialidad_1,
                    "nombre_de_entre_vialidad_1": nombre_de_entre_vialidad_1,
                    "tipo_de_entre_vialidad_2": tipo_de_entre_vialidad_2,
                    "nombre_de_entre_vialidad_2": nombre_de_entre_vialidad_2,
                    "tipo_de_entre_vialidad_3": tipo_de_entre_vialidad_3,
                    "nombre_de_entre_vialidad_3": nombre_de_entre_vialidad_3,
                    "numero_exterior_o_kilometro": numero_exterior_o_kilometro,
                    "letra_exterior": letra_exterior,
                    "edificio": edificio,
                    "edificio_piso": edificio_piso,
                    "numero_interior": numero_interior,
                    "letra_interior": letra_interior,
                    "tipo_de_asentamiento_humano": tipo_de_asentamiento_humano,
                    "nombre_de_asentamiento_humano": nombre_de_asentamiento_humano,
                    "tipo_centro_comercial": tipo_centro_comercial,
                    "c_industrial_comercial_o_mercado": c_industrial_comercial_o_mercado,
                    "numero_de_local": numero_de_local,
                    "codigo_postal": codigo_postal,
                    "clave_entidad": clave_entidad,
                    "entidad_federativa": entidad_federativa,
                    "clave_municipio": clave_municipio,
                    "municipio": municipio,
                    "clave_localidad": clave_localidad,
                    "localidad": localidad,
                    "area_geoestadistica_basica": area_geoestadistica_basica,
                    "manzana": manzana,
                    "numero_de_telefono": numero_de_telefono,
                    "correo_electronico": correo_electronico,
                    "sitio_en_internet": sitio_en_internet,
                    "tipo_de_establecimiento": tipo_de_establecimiento,
                    "latitud": latitud,
                    "longitud": longitud,
                    "fecha_de_incorporacion_al_denue": fecha_de_incorporacion_al_denue_str,
                    "categoria_turistica": categoria_turistica,
                    "no_cuartos": no_cuartos,
                    "unidades": unidades,
                    "espacios_cajones": espacios_cajones,
                    "no_camas": no_camas,
                    "cadena": cadena,
                    "operadora": operadora,
                    "responsable": responsable,
                    "cargo": cargo,
                    "imss": imss,
                    "inicio_de_operaciones_este_ano": inicio_de_operaciones_este_ano,
                    "fecha_de_inicio_de_operaciones": fecha_de_inicio_de_operaciones_str,
                    "centro_turistico": centro_turistico,
                    "zona": zona,
                    "datatur": datatur,
                    "hotel_boutique": hotel_boutique,
                    "nombre_de_la_cadena": nombre_de_la_cadena,
                    "hoteles_tesoros": hoteles_tesoros,
                }

                try:
                    # if destino not in CatalagoDestino.objects.values_list('destino', flat=True):
                    #     print(f"El destino {destino} no está en la tabla CatalagoDestinoAeropuerto")
                    #     registros_incorrectos.append(datos)
                    #     continue

                    # Buscar si la fila ya existe en la base de datos
                    existente = DirectorioHotelero.objects.filter(
                        nombre_de_la_unidad_economica=nombre_de_la_unidad_economica
                    )
                    if existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {datos} ya existe en la base de datos")
                        registros_existentes.append(datos)
                        registros_incorrectos.append(datos)
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        db = DirectorioHotelero(
                            id_establecimiento=id_establecimiento,
                            nombre_de_la_unidad_economica=nombre_de_la_unidad_economica,
                            razon_social=razon_social,
                            codigo_de_la_clase_de_actividad_scian=codigo_de_la_clase_de_actividad_scian,
                            nombre_de_clase_de_la_actividad=nombre_de_clase_de_la_actividad,
                            descripcion_estrato_personal_ocupado=descripcion_estrato_personal_ocupado,
                            tipo_de_vialidad=tipo_de_vialidad,
                            nombre_de_la_vialidad=nombre_de_la_vialidad,
                            tipo_de_entre_vialidad_1=tipo_de_entre_vialidad_1,
                            nombre_de_entre_vialidad_1=nombre_de_entre_vialidad_1,
                            tipo_de_entre_vialidad_2=tipo_de_entre_vialidad_2,
                            nombre_de_entre_vialidad_2=nombre_de_entre_vialidad_2,
                            tipo_de_entre_vialidad_3=tipo_de_entre_vialidad_3,
                            nombre_de_entre_vialidad_3=nombre_de_entre_vialidad_3,
                            numero_exterior_o_kilometro=numero_exterior_o_kilometro,
                            letra_exterior=letra_exterior,
                            edificio=edificio,
                            edificio_piso=edificio_piso,
                            numero_interior=numero_interior,
                            letra_interior=letra_interior,
                            tipo_de_asentamiento_humano=tipo_de_asentamiento_humano,
                            nombre_de_asentamiento_humano=nombre_de_asentamiento_humano,
                            tipo_centro_comercial=tipo_centro_comercial,
                            c_industrial_comercial_o_mercado=c_industrial_comercial_o_mercado,
                            numero_de_local=numero_de_local,
                            codigo_postal=codigo_postal,
                            clave_entidad=clave_entidad,
                            entidad_federativa=entidad_federativa,
                            clave_municipio=clave_municipio,
                            municipio=municipio,
                            clave_localidad=clave_localidad,
                            localidad=localidad,
                            area_geoestadistica_basica=area_geoestadistica_basica,
                            manzana=manzana,
                            numero_de_telefono=numero_de_telefono,
                            correo_electronico=correo_electronico,
                            sitio_en_internet=sitio_en_internet,
                            tipo_de_establecimiento=tipo_de_establecimiento,
                            latitud=latitud,
                            longitud=longitud,
                            fecha_de_incorporacion_al_denue=fecha_de_incorporacion_al_denue,
                            categoria_turistica=categoria_turistica,
                            no_cuartos=no_cuartos,
                            unidades=unidades,
                            espacios_cajones=espacios_cajones,
                            no_camas=no_camas,
                            cadena=cadena,
                            operadora=operadora,
                            responsable=responsable,
                            cargo=cargo,
                            imss=imss,
                            inicio_de_operaciones_este_ano=inicio_de_operaciones_este_ano,
                            fecha_de_inicio_de_operaciones=fecha_de_inicio_de_operaciones,
                            centro_turistico=centro_turistico,
                            zona=zona,
                            datatur=datatur,
                            hotel_boutique=hotel_boutique,
                            nombre_de_la_cadena=nombre_de_la_cadena,
                            hoteles_tesoros=hoteles_tesoros
                        )
                        db.save()
                        registros_correctos.append(datos)
                except (ValueError, TypeError) as e:
                    print(f"Error al procesar la fila {datos}: {e}")
                    registros_incorrectos.append(datos)

        except FileNotFoundError:
            print(f"El archivo {archivo} no se pudo abrir")

        return registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas

    def procesar_archivo_csv(self, archivo):
        archivo = self.request.FILES['archivo']
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        try:
            datos = csv.DictReader(
                archivo.read().decode('latin-1').splitlines())
            # print(datos)
            for row in datos:
                num_filas_procesadas += 1

                id_establecimiento = row['id_establecimiento']
                nombre_de_la_unidad_economica = row['nombre_de_la_unidad_economica']
                razon_social = row['razon_social']
                codigo_de_la_clase_de_actividad_scian = row['codigo_de_la_clase_de_actividad_scian']
                nombre_de_clase_de_la_actividad = row['nombre_de_clase_de_la_actividad']
                descripcion_estrato_personal_ocupado = row['descripcion_estrato_personal_ocupado']
                tipo_de_vialidad = row['tipo_de_vialidad']
                nombre_de_la_vialidad = row['nombre_de_la_vialidad']
                tipo_de_entre_vialidad_1 = row['tipo_de_entre_vialidad_1']
                nombre_de_entre_vialidad_1 = row['nombre_de_entre_vialidad_1']
                tipo_de_entre_vialidad_2 = row['tipo_de_entre_vialidad_2']
                nombre_de_entre_vialidad_2 = row['nombre_de_entre_vialidad_2']
                tipo_de_entre_vialidad_3 = row['tipo_de_entre_vialidad_3']
                nombre_de_entre_vialidad_3 = row['nombre_de_entre_vialidad_3']
                numero_exterior_o_kilometro = row['numero_exterior_o_kilometro']
                letra_exterior = row['letra_exterior']
                edificio = row['edificio']
                edificio_piso = row['edificio_piso']
                numero_interior = row['numero_interior']
                letra_interior = row['letra_interior']
                tipo_de_asentamiento_humano = row['tipo_de_asentamiento_humano']
                nombre_de_asentamiento_humano = row['nombre_de_asentamiento_humano']
                tipo_centro_comercial = row['tipo_centro_comercial']
                c_industrial_comercial_o_mercado = row['c_industrial_comercial_o_mercado']
                numero_de_local = row['numero_de_local']
                codigo_postal = row['codigo_postal']
                clave_entidad = row['clave_entidad']
                entidad_federativa = row['entidad_federativa']
                clave_municipio = row['clave_municipio']
                municipio = row['municipio']
                clave_localidad = row['clave_localidad']
                localidad = row['localidad']
                area_geoestadistica_basica = row['area_geoestadistica_basica']
                manzana = row['manzana']
                numero_de_telefono = row['numero_de_telefono']
                correo_electronico = row['correo_electronico']
                sitio_en_internet = row['sitio_en_internet']
                tipo_de_establecimiento = row['tipo_de_establecimiento']
                latitud = row['latitud']
                longitud = row['longitud']
                fecha_de_incorporacion_al_denue = row['fecha_de_incorporacion_al_denue']
                categoria_turistica = row['categoria_turistica']
                no_cuartos = row['no_cuartos']
                unidades = row['unidades']
                espacios_cajones = row['espacios_cajones']
                no_camas = row['no_camas']
                cadena = row['cadena']
                operadora = row['operadora']
                responsable = row['responsable']
                cargo = row['cargo']
                imss = row['imss']
                inicio_de_operaciones_este_ano = row['inicio_de_operaciones_este_ano']
                fecha_de_inicio_de_operaciones = row['fecha_de_inicio_de_operaciones']
                centro_turistico = row['centro_turistico']
                zona = row['zona']
                datatur = row['datatur']
                hotel_boutique = row['hotel_boutique']
                nombre_de_la_cadena = row['nombre_de_la_cadena']
                hoteles_tesoros = row['hoteles_tesoros']

                datos = {
                    "id_establecimiento": id_establecimiento,
                    "nombre_de_la_unidad_economica": nombre_de_la_unidad_economica,
                    "razon_social": razon_social,
                    "codigo_de_la_clase_de_actividad_scian": codigo_de_la_clase_de_actividad_scian,
                    "nombre_de_clase_de_la_actividad": nombre_de_clase_de_la_actividad,
                    "descripcion_estrato_personal_ocupado": descripcion_estrato_personal_ocupado,
                    "tipo_de_vialidad": tipo_de_vialidad,
                    "nombre_de_la_vialidad": nombre_de_la_vialidad,
                    "tipo_de_entre_vialidad_1": tipo_de_entre_vialidad_1,
                    "nombre_de_entre_vialidad_1": nombre_de_entre_vialidad_1,
                    "tipo_de_entre_vialidad_2": tipo_de_entre_vialidad_2,
                    "nombre_de_entre_vialidad_2": nombre_de_entre_vialidad_2,
                    "tipo_de_entre_vialidad_3": tipo_de_entre_vialidad_3,
                    "nombre_de_entre_vialidad_3": nombre_de_entre_vialidad_3,
                    "numero_exterior_o_kilometro": numero_exterior_o_kilometro,
                    "letra_exterior": letra_exterior,
                    "edificio": edificio,
                    "edificio_piso": edificio_piso,
                    "numero_interior": numero_interior,
                    "letra_interior": letra_interior,
                    "tipo_de_asentamiento_humano": tipo_de_asentamiento_humano,
                    "nombre_de_asentamiento_humano": nombre_de_asentamiento_humano,
                    "tipo_centro_comercial": tipo_centro_comercial,
                    "c_industrial_comercial_o_mercado": c_industrial_comercial_o_mercado,
                    "numero_de_local": numero_de_local,
                    "codigo_postal": codigo_postal,
                    "clave_entidad": clave_entidad,
                    "entidad_federativa": entidad_federativa,
                    "clave_municipio": clave_municipio,
                    "municipio": municipio,
                    "clave_localidad": clave_localidad,
                    "localidad": localidad,
                    "area_geoestadistica_basica": area_geoestadistica_basica,
                    "manzana": manzana,
                    "numero_de_telefono": numero_de_telefono,
                    "correo_electronico": correo_electronico,
                    "sitio_en_internet": sitio_en_internet,
                    "tipo_de_establecimiento": tipo_de_establecimiento,
                    "latitud": latitud,
                    "longitud": longitud,
                    "fecha_de_incorporacion_al_denue": fecha_de_incorporacion_al_denue,
                    "categoria_turistica": categoria_turistica,
                    "no_cuartos": no_cuartos,
                    "unidades": unidades,
                    "espacios_cajones": espacios_cajones,
                    "no_camas": no_camas,
                    "cadena": cadena,
                    "operadora": operadora,
                    "responsable": responsable,
                    "cargo": cargo,
                    "imss": imss,
                    "inicio_de_operaciones_este_ano": inicio_de_operaciones_este_ano,
                    "fecha_de_inicio_de_operaciones": fecha_de_inicio_de_operaciones,
                    "centro_turistico": centro_turistico,
                    "zona": zona,
                    "datatur": datatur,
                    "hotel_boutique": hotel_boutique,
                    "nombre_de_la_cadena": nombre_de_la_cadena,
                    "hoteles_tesoros": hoteles_tesoros,
                }

                try:
                    # if destino not in CatalagoDestino.objects.values_list('destino', flat=True):
                    #     print(f"El destino {destino} no está en la tabla CatalagoDestinoAeropuerto")
                    #     registros_incorrectos.append(datos)
                    #     continue

                    # Buscar si la fila ya existe en la base de datos
                    existente = DirectorioHotelero.objects.filter(
                        nombre_de_la_unidad_economica=nombre_de_la_unidad_economica
                    )
                    if existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {row} ya existe en la base de datos")
                        registros_existentes.append(datos)
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        db = DirectorioHotelero(
                            id_establecimiento=id_establecimiento,
                            nombre_de_la_unidad_economica=nombre_de_la_unidad_economica,
                            razon_social=razon_social,
                            codigo_de_la_clase_de_actividad_scian=codigo_de_la_clase_de_actividad_scian,
                            nombre_de_clase_de_la_actividad=nombre_de_clase_de_la_actividad,
                            descripcion_estrato_personal_ocupado=descripcion_estrato_personal_ocupado,
                            tipo_de_vialidad=tipo_de_vialidad,
                            nombre_de_la_vialidad=nombre_de_la_vialidad,
                            tipo_de_entre_vialidad_1=tipo_de_entre_vialidad_1,
                            nombre_de_entre_vialidad_1=nombre_de_entre_vialidad_1,
                            tipo_de_entre_vialidad_2=tipo_de_entre_vialidad_2,
                            nombre_de_entre_vialidad_2=nombre_de_entre_vialidad_2,
                            tipo_de_entre_vialidad_3=tipo_de_entre_vialidad_3,
                            nombre_de_entre_vialidad_3=nombre_de_entre_vialidad_3,
                            numero_exterior_o_kilometro=numero_exterior_o_kilometro,
                            letra_exterior=letra_exterior,
                            edificio=edificio,
                            edificio_piso=edificio_piso,
                            numero_interior=numero_interior,
                            letra_interior=letra_interior,
                            tipo_de_asentamiento_humano=tipo_de_asentamiento_humano,
                            nombre_de_asentamiento_humano=nombre_de_asentamiento_humano,
                            tipo_centro_comercial=tipo_centro_comercial,
                            c_industrial_comercial_o_mercado=c_industrial_comercial_o_mercado,
                            numero_de_local=numero_de_local,
                            codigo_postal=codigo_postal,
                            clave_entidad=clave_entidad,
                            entidad_federativa=entidad_federativa,
                            clave_municipio=clave_municipio,
                            municipio=municipio,
                            clave_localidad=clave_localidad,
                            localidad=localidad,
                            area_geoestadistica_basica=area_geoestadistica_basica,
                            manzana=manzana,
                            numero_de_telefono=numero_de_telefono,
                            correo_electronico=correo_electronico,
                            sitio_en_internet=sitio_en_internet,
                            tipo_de_establecimiento=tipo_de_establecimiento,
                            latitud=latitud,
                            longitud=longitud,
                            fecha_de_incorporacion_al_denue=fecha_de_incorporacion_al_denue,
                            categoria_turistica=categoria_turistica,
                            no_cuartos=no_cuartos,
                            unidades=unidades,
                            espacios_cajones=espacios_cajones,
                            no_camas=no_camas,
                            cadena=cadena,
                            operadora=operadora,
                            responsable=responsable,
                            cargo=cargo,
                            imss=imss,
                            inicio_de_operaciones_este_ano=inicio_de_operaciones_este_ano,
                            fecha_de_inicio_de_operaciones=fecha_de_inicio_de_operaciones,
                            centro_turistico=centro_turistico,
                            zona=zona,
                            datatur=datatur,
                            hotel_boutique=hotel_boutique,
                            nombre_de_la_cadena=nombre_de_la_cadena,
                            hoteles_tesoros=hoteles_tesoros
                        )
                        db.save()
                        registros_correctos.append(datos)
                except (ValueError, TypeError) as e:
                    print(f"Error al procesar la fila {datos}: {e}")
                    registros_incorrectos.append(datos)
        except FileNotFoundError:
            print(f"No se encontró el archivo {archivo}")
        except Exception as e:
            print(f"Error al procesar el archivo {archivo}: {e}")
        return registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas


class DirectorioHoteleroDescargarArchivoView(View):

    def crear_archivo_excel(self, registros_incorrectos):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Obtener los nombres de los campos del modelo DirectorioHotelero
        campos = [field.name for field in DirectorioHotelero._meta.get_fields()
                  if field.name != 'id']

        # Escribir los encabezados de las columnas
        for i, campo in enumerate(campos):
            columna = i + 1
            worksheet.cell(row=1, column=columna, value=campo)

        # Obtener los datos del modelo DirectorioHotelero
        datos = registros_incorrectos

        # Escribir los valores en las celdas correspondientes
        fila = 2
        for registro in datos:
            for i, campo in enumerate(campos):
                if campo != 'id':  # Omitir la clave 'id'
                    columna = i + 1
                    valor = registro[campo]
                    worksheet.cell(row=fila, column=columna, value=valor)
            fila += 1

        # Set the column widths to auto-fit
        for column in worksheet.columns:
            max_length = 0
            column_name = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_name].width = adjusted_width

        # Create the response with the Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=gasto_derrama_registros_incorrectos.xls'

        # workbook.save(response)
        return workbook

    def post(self, request, *args, **kwargs):
        # Obtener los registros incorrectos del cuerpo de la petición
        registros_incorrectos = json.loads(request.body)

        # Crear y enviar el archivo de Excel con las filas incorrectas
        workbook = self.crear_archivo_excel(registros_incorrectos)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=gasto_derrama_registros_incorrectos.xlsx'
        workbook.save(response)
        return response
