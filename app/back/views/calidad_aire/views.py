from django.shortcuts import render, redirect
from django.views.generic import CreateView, ListView, DetailView, UpdateView, DeleteView
from django.urls import reverse_lazy
from back.models import  *
from back.forms import *
from django.http import JsonResponse, HttpResponseRedirect
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.views import View
from django.contrib import messages
from openpyxl import load_workbook
import csv
import os
import datetime
from django.urls import reverse

# para archivo excel
import openpyxl
from django.http import HttpResponse
import json
from config.diccionarios import clean_str_col, homologar_columna_categoria, homologar_columna_destino

from django.contrib.auth.decorators import login_required, permission_required
from django.utils.decorators import method_decorator
from django.http import Http404
from django.core.exceptions import PermissionDenied
from back.mixins import *
from django.contrib.auth.decorators import user_passes_test
from django.template.loader import render_to_string

def es_admin_o_superadmin(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)



# Create your views here.
def is_ajax(request):
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'


class CalidadAireListView(SuperAdminOrAdminMixin, LoginRequiredMixin, ListView):
    model = CalidadAire
    template_name = 'back/calidad_aire/list.html'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs) :
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'search':
                data = []
                for i in CalidadAire.objects.all():
                    data.append(i.toJSON())
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Listado de Calidad del Aire'
        context['create_url'] = reverse_lazy('dashboard:calidad_aire_create')
        context['carga_masiva_url'] = reverse_lazy('dashboard:calidad_aire_carga_masiva')
        context['entity'] = 'Calidad del Aire'
        context['d_route'] = 'Fuentes de información > Dashboard'
        context['is_fuente'] = True
        return context


class CalidadAireCreateView(SuperAdminOrAdminMixin, LoginRequiredMixin, CreateView):
    model = CalidadAire
    form_class = CalidadAireForm
    template_name = 'back/calidad_aire/create.html'
    success_url = reverse_lazy('dashboard:calidad_aire_list')

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        replace_data = request.POST.get('replace_data')
        if form.is_valid():
            # Verificar si ya existe un registro con la misma fecha y destino
            fecha = form.cleaned_data['fecha']
            destino = form.cleaned_data['destino']

            try:
                existing_object = self.model.objects.get(fecha=fecha, destino=destino)
            except self.model.DoesNotExist:
                existing_object = None

            # Si no existe un objeto existente, guardar los nuevos datos
            if not existing_object:
                self.object = form.save()
                data = {
                    'success': True,
                    'message': 'Datos creados exitosamente.',
                    'url': self.success_url,
                }
                return JsonResponse(data)

            # Si existe un objeto existente y replace_data es True, actualizar los datos existentes
            if existing_object and replace_data == 'on':
                for field in form.cleaned_data:
                    if field == 'replace_data':
                        continue
                    setattr(existing_object, field, form.cleaned_data[field])
                existing_object.save()

                data = {
                    'success': True,
                    'message': 'Datos actualizados exitosamente.',
                    'url': self.success_url,
                }
                return JsonResponse(data)

            # Si existe un objeto existente y replace_data es False, devolver un error
            if existing_object:
                data = {
                    'success': False,
                    'message': 'Hubo un error al crear el registro.',
                    'errors': 'Ya existe un registro con la misma fecha y destino.',
                }
                return JsonResponse(data)
        else:
            data = {
                'success': False,
                'message': 'Error al crear los datos.',
                'errors': form.errors,
                'format_errors': True,
            }
            return JsonResponse(data)
        

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Ha ocurrido un error al crear un registro.',
            'errors': form.errors
        }
        return JsonResponse(data)

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Registro creado exitosamente.',
            'url': self.success_url
        }
        return JsonResponse(data)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Crear Registro'
        context['entity'] = 'Calidad del Aire'
        context['list_url'] = reverse_lazy('dashboard:calidad_aire_list')
        context['action'] = 'add'
        return context


class CalidadAireUpdateView(SuperAdminOrAdminMixin, LoginRequiredMixin,  UpdateView):
    model = CalidadAire
    form_class = CalidadAireForm
    template_name = 'back/calidad_aire/view_editor.html'
    success_url = reverse_lazy('dashboard:calidad_aire_list')

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Ha ocurrido un error al editar el registro.',
            'errors': form.errors
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Registro creado exitosamente.',
            'url': self.success_url
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Editar Calidad del Aire'
        context['entity'] = 'Calidad del Aire'
        context['list_url'] = reverse_lazy('dashboard:calidad_aire_list')
        destino_widget = context['form'].fields['destino'].widget
        destino_widget.attrs.update({'readonly': 'readonly'})
        fecha_widget = context['form'].fields['fecha'].widget
        fecha_widget.attrs.update({'readonly': 'readonly'})
        context['edit_msg'] = 'Los Campos que no se pueden editar están sombreados'
        return context


class CalidadAireDeleteView(SuperAdminOrAdminMixin, LoginRequiredMixin, DeleteView):
    model = CalidadAire
    # template_name = 'back/delete.html'
    success_url = reverse_lazy('dashboard:calidad_aire_list')
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object()

        try:
            self.object.delete()
            return JsonResponse({'message': 'Eliminación exitosa.'})
        except Exception as e:
            return JsonResponse({'error': 'Error al eliminar el registro.'}, status=500)


class CalidadAireCargaMasivaView(SuperAdminOrAdminMixin, LoginRequiredMixin, View):
    form_class = CargaMasivaForm
    template_name = 'back/calidad_aire/carga_masiva.html'
    success_url = reverse_lazy('dashboard:calidad_aire_list')

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form, 'title': 'Carga Masiva'})


    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        archivo = request.FILES.get('archivo', None)
        if archivo:
            extension = os.path.splitext(archivo.name)[1]
            if extension == '.xlsx':
                registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas = self.procesar_archivo_xlsx(archivo)
            elif extension == '.csv':
                registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas = self.procesar_archivo_csv(archivo)
            else:
                messages.error(request, 'El archivo debe ser un archivo .xlsx o .csv')
                registros_incorrectos.append("El archivo debe ser un archivo .xlsx o .csv")
        else:
            messages.error(request, 'Debe seleccionar un archivo')
            registros_incorrectos.append("Debe seleccionar un archivo")

        if len(registros_incorrectos) > 0 or len(registros_existentes) > 0:
            messages.error(request, 'Hay errores de registros')
            datos_json = json.dumps(registros_incorrectos)
            
            return render(request, self.template_name, {
                'form': form,
                'title': 'Carga Masiva',
                'registros_correctos': registros_correctos,
                'registros_incorrectos': registros_incorrectos,
                'registros_existentes': registros_existentes,
                'descargar_url': datos_json,
                'num_filas_procesadas': num_filas_procesadas,
            })
            
        else:
            return HttpResponseRedirect(reverse('dashboard:calidad_aire_list'))
        
        

    def procesar_archivo_xlsx(self, archivo):
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        try:
            workbook = load_workbook(filename=archivo, read_only=True)
            worksheet = workbook.active
            filas = list(worksheet.rows)
            for i, row in enumerate(filas):
                if i == 0:
                    continue # Ignorar la primera fila si es el encabezado

                if not row:
                    continue  # Salta filas vacías

                num_filas_procesadas += 1
                # Limpieza de datos
                destino = clean_str_col(row[1].value)

                # Homologación de datos
                destino = homologar_columna_destino(destino)

                # Validar si el destino si es válido
                if destino not in CatalagoDestino.objects.values_list('destino', flat=True):
                    print(f"El destino {destino} no está en la tabla CatalagoDestino")
                    registros_incorrectos.append({'fecha': fecha_obj, 'destino': destino, 'calidad_del_aire': calidad_del_aire})
                    continue
                fecha = row[0].value.date()
                calidad_del_aire = row[2].value

                try:
                    # Validar los datos
                    fecha_str = str(fecha)
                    fecha_obj = datetime.datetime.strptime(fecha_str, '%Y-%m-%d').date()

                    # Buscar si la fila ya existe en la base de datos
                    inventario_existente = CalidadAire.objects.filter(fecha=fecha_obj, destino=destino)
                    if inventario_existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {row} ya existe en la base de datos")
                        registros_existentes.append({'fila': i, 'fecha': fecha_obj, 'destino': destino, 'calidad_del_aire': calidad_del_aire})
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        inventario = CalidadAire(fecha=fecha_obj, destino=destino, calidad_del_aire=calidad_del_aire)
                        inventario.save()
                        registros_correctos.append({'fila': i,  'fecha': fecha_obj, 'destino': destino, 'calidad_del_aire': calidad_del_aire})
                except (ValueError, TypeError) as e:
                    # Si los datos no son válidos, se guarda el número de fila en la lista de registros incorrectos
                    registros_incorrectos.append({'fila': i,'fecha': fecha, 'destino': destino, 'calidad_del_aire': calidad_del_aire, 'error': str(e)})
                    
        except FileNotFoundError:
                print(f"El archivo {archivo} no se pudo abrir")
                
        return registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas
    
    def procesar_archivo_csv(self, archivo):
        archivo = self.request.FILES['archivo']
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        try:
            datos = csv.DictReader(archivo.read().decode('latin-1').splitlines())
            # print(datos)
            for row in datos:
                num_filas_procesadas += 1

                # Limpieza de datos
                destino = clean_str_col(row['destino'])

                # Homologación de datos
                destino = homologar_columna_destino(destino)

                # Validar si el destino si es válido
                if destino not in CatalagoDestino.objects.values_list('destino', flat=True):
                    print(f"El destino {destino} no está en la tabla CatalagoDestino")
                    registros_incorrectos.append(datos)
                    continue
                fecha = row['fecha']
                calidad_del_aire = row['calidad_del_aire']

                try:
                    # Validar los datos
                    fecha_str = str(fecha)
                    fecha_obj = datetime.datetime.strptime(fecha_str, '%d/%m/%Y').date()

                    # Buscar si la fila ya existe en la base de datos
                    inventario_existente = CalidadAire.objects.filter(fecha=fecha_obj, destino=destino)
                    if inventario_existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {row} ya existe en la base de datos")
                        registros_existentes.append(datos)
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        inventario = CalidadAire(fecha=fecha_obj, destino=destino, calidad_del_aire=calidad_del_aire)
                        inventario.save()
                        registros_correctos.append(datos)
                except (ValueError, TypeError) as e:
                    print(f"Error al procesar la fila {row}: {e}")
                    registros_incorrectos.append(datos)
        except FileNotFoundError:
            print(f"No se encontró el archivo {archivo}")
        except Exception as e:
            print(f"Error al procesar el archivo {archivo}: {e}")
        return registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas



class DescargarArchivoAireView(SuperAdminOrAdminMixin, LoginRequiredMixin, View):

    def crear_archivo_excel(self, registros_incorrectos):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Add headers to the worksheet
        worksheet['A1'] = 'Fecha'
        worksheet['B1'] = 'Destino'
        worksheet['C1'] = 'calidad_del_aire'

        # Add the incorrect rows to the worksheet
        for i, row in enumerate(registros_incorrectos):
            fila = i + 2
            # worksheet.cell(row=fila, column=1, value=row['fila'])
            worksheet.cell(row=fila, column=1, value=row['fecha'])
            worksheet.cell(row=fila, column=2, value=row['destino'])
            worksheet.cell(row=fila, column=3, value=row['calidad_del_aire'])
            # worksheet.cell(row=fila, column=7, value=row['error'])

        # Set the column widths to auto-fit
        for column in worksheet.columns:
            max_length = 0
            column_name = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_name].width = adjusted_width

        # Create the response with the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=registros_incorrectos.xls'

        

        # workbook.save(response)
        return workbook
    
    def post(self, request, *args, **kwargs):
        # Obtener los registros incorrectos del cuerpo de la petición
        registros_incorrectos = json.loads(request.body)

        # Crear y enviar el archivo de Excel con las filas incorrectas
        workbook = self.crear_archivo_excel(registros_incorrectos)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=registros_incorrectos.xlsx'
        workbook.save(response)
        return response       
    
    