from django.shortcuts import render, redirect
from django.views.generic import CreateView, ListView, DetailView, UpdateView, DeleteView
from django.urls import reverse_lazy
from back.models import  *
from back.forms import *
from django.http import JsonResponse, HttpResponseRedirect
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.views import View
from django.contrib import messages
from openpyxl import load_workbook
import csv
import os
import datetime
from django.urls import reverse
import openpyxl
from django.http import HttpResponse
from django.template.loader import render_to_string

import json
from config.diccionarios import clean_str_col, homologar_columna_categoria, homologar_columna_destino, clean_str_col_des, parse_fecha
from django.contrib.auth.decorators import login_required, permission_required
from django.utils.decorators import method_decorator
from django.http import Http404
from django.core.exceptions import PermissionDenied
from back.mixins import *
from django.contrib.auth.decorators import user_passes_test
import logging

logger = logging.getLogger(__name__)

def es_admin_o_superadmin(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)



# Create your views here.
def is_ajax(request):
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'



class InversionPublicaListView(SuperAdminOrAdminMixin, LoginRequiredMixin, ListView):
    model = InversionPublica
    template_name = 'back/inversion_publica/list.html'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs) :
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'search':
                data = []
                for i in InversionPublica.objects.all():
                    data.append(i.toJSON())
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)
    

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Listado de Inversión publica'
        context['create_url'] = reverse_lazy('dashboard:inversion_publica_create')
        context['carga_masiva_url'] = reverse_lazy('dashboard:inversion_publica_carga_masiva')
        context['entity'] = 'Inversión publica'
        context['is_fuente'] = True
        return context


class  InversionPublicaCreateView(SuperAdminOrAdminMixin, LoginRequiredMixin, CreateView):
    model = InversionPublica
    form_class = InversionPublicaForm
    template_name = 'back/inversion_publica/create.html'
    success_url = reverse_lazy('dashboard:inversion_publica_list')


    def get_object(self, **kwargs):
        queryset = self.get_queryset()
        try:
            return queryset.get(**kwargs)
        except queryset.model.DoesNotExist:
            return None

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        replace_data = request.POST.get('replace_data')
        if form.is_valid():
            # Check if there is already a record with the same fecha, destino and categoria

            fecha = form.cleaned_data['fecha']
            destino = form.cleaned_data['destino']

            nombre_de_la_obra = form.cleaned_data['nombre_de_la_obra']

            try:
                existing_object = self.get_object(fecha=fecha, destino=destino , nombre_de_la_obra=nombre_de_la_obra)

            except InventarioHoteleroEntNac.DoesNotExist:
                existing_object = None

            existing_catalogo = CatalagoDestino.objects.filter(destino=destino).exists()
    
            # If there is no existing data, save the new data
            if not existing_catalogo :
                    data = {
                        'success': False,
                        'missingData': True,
                        'destino': destino,
                        'message': 'No existe la entidad en el catálogo de destinos',
                    }
                    return JsonResponse(data)

            # If there is existing data and replace_data is True, delete the existing data
            if existing_object and request.POST.get('replace_data') == 'on':
                # existing_object.delete()
                # Save the new data
                # self.object = form.save()

                for field in form.cleaned_data:
                    if field == 'replace_data':
                        continue
                    setattr(existing_object, field, form.cleaned_data[field])

                existing_object.save()

                data = {
                    'success': True,
                    'message': 'Data created successfully.',
                    'url': self.success_url,

                }
                return JsonResponse(data)

            # If there is existing data and replace_data is False, return an error

            if existing_object:
                data =  InversionPublica.objects.filter(fecha=fecha, destino=destino , nombre_de_la_obra=nombre_de_la_obra)

                data_list = list(data.values('fecha', 'destino', 'monto_total','nombre_de_la_obra', 'monto_de_inversion_municipal', 'monto_de_inversion_municipal', 'monto_de_inversion_estatal', 'monto_de_inversion_federal'))

                data_list2 = list(form.cleaned_data.values())

                table_html = render_to_string('back/inversion_publica/table.html',{'data_list': data_list, 'actual': True, 'data_list2': data_list2})

                datajsn = {
                    'success': False,
                    'message': 'Hubo un error al crear registro.',
                    'errors': 'Ya existe un registro con la misma fecha, destino y categoria',
                    'existing_object': table_html
                }

                return JsonResponse(datajsn)
            else:
                self.object = form.save()
                data = {
                    'success': True,
                    'message': 'Data created successfully.',
                    'url': self.success_url
                }
                return JsonResponse(data)
        else:
            data = {
                'success': False,
                'message': 'Error creating data.',
                'errors': form.errors,
                'format_errors': True
            }
            return JsonResponse(data)

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Hubo un error al crear registro.',
            'errors': form.errors
        }
        return JsonResponse(data)

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Registro creado exitosamente.',
            'url': self.success_url
        }
        return JsonResponse(data)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Crear Inventario Hotelero'
        context['entity'] = 'Inventario Hotelero'
        context['list_url'] = reverse_lazy('dashboard:inversion_publica_list')
        context['action'] = 'add'
        return context
    




class InversionPublicaUpdateView(SuperAdminOrAdminMixin, LoginRequiredMixin,  UpdateView):
    model = InversionPublica
    form_class = InversionPublicaForm
    template_name = 'back/inversion_publica/create.html'
    success_url = reverse_lazy('dashboard:inversion_publica_list')

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Ha ocurrido un error al editar el registro.',
            'errors': form.errors
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Registro creado exitosamente.',
            'url': self.success_url
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Editar Inversión publica'
        context['entity'] = 'Inversión publica'
        context['list_url'] = reverse_lazy('dashboard:inversion_publica_list')
        context['form'] = self.form_class(instance=self.object)       
        context['action'] = 'edit'
        return context


class InversionPublicaDeleteView(SuperAdminOrAdminMixin, LoginRequiredMixin, DeleteView):
    model = InversionPublica
    # template_name = 'back/delete.html'
    success_url = reverse_lazy('dashboard:inversion_publica_list')
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object()

        try:
            self.object.delete()
            return JsonResponse({'message': 'Eliminación exitosa.'})
        except Exception as e:
            return JsonResponse({'error': 'Error al eliminar el registro.'}, status=500)


class InversionPublicaCargaMasivaView(SuperAdminOrAdminMixin, LoginRequiredMixin, View):
    form_class = CargaMasivaForm
    template_name = 'back/inversion_publica/carga_masiva.html'
    success_url = reverse_lazy('dashboard:inversion_publica_list')

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form, 'title': 'Carga Masiva'})


    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        archivo = request.FILES.get('archivo', None)
        if archivo:
            extension = os.path.splitext(archivo.name)[1]
            if extension == '.xlsx':
                registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas = self.procesar_archivo_xlsx(archivo)
            elif extension == '.csv':
                registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas = self.procesar_archivo_csv(archivo)
            else:
                messages.error(request, 'El archivo debe ser un archivo .xlsx o .csv')
                registros_incorrectos.append("El archivo debe ser un archivo .xlsx o .csv")
        else:
            messages.error(request, 'Debe seleccionar un archivo')
            registros_incorrectos.append("Debe seleccionar un archivo")

        if len(registros_incorrectos) > 0 or len(registros_existentes) > 0:
            messages.error(request, 'Hay errores de registros')
            datos_json = json.dumps(registros_incorrectos)
            
            return render(request, self.template_name, {
                'form': form,
                'title': 'Carga Masiva',
                'registros_correctos': registros_correctos,
                'registros_incorrectos': registros_incorrectos,
                'registros_existentes': registros_existentes,
                'descargar_url': datos_json,
                'num_filas_procesadas': num_filas_procesadas,
            })
            
        else:
            return HttpResponseRedirect(reverse('dashboard:inversion_publica_list'))
        
        

    def procesar_archivo_xlsx(self, archivo):
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        try:
            workbook = load_workbook(filename=archivo, read_only=True)
            worksheet = workbook.active
            filas = list(worksheet.rows)
            for i, row in enumerate(filas):
                if i == 0:
                    continue # Ignorar la primera fila si es el encabezado

                if not row or all(cell.value is None for cell in row):
                    continue  # Salta filas vacías

                num_filas_procesadas += 1
                # Limpieza de datos
                destino = clean_str_col_des(row[0].value)

                # Homologación de datos
                # destino = homologar_columna_destino(destino)

                
                fecha_str = str(row[1].value).strip() if row[1].value else ''   
                nombre_de_la_obra = row[2].value
                monto_de_inversion_municipal = row[3].value
                monto_de_inversion_estatal = row[4].value
                monto_de_inversion_federal = row[5].value
                monto_total = str(row[6].value).strip() if row[6].value else ''
                

                datos = {
                    'fecha': fecha_str,
                    'destino': destino,
                    'nombre_de_la_obra': nombre_de_la_obra,
                    'monto_de_inversion_municipal': monto_de_inversion_municipal,
                    'monto_de_inversion_estatal': monto_de_inversion_estatal,
                    'monto_de_inversion_federal': monto_de_inversion_federal,
                    'monto_total': monto_total,
                }

                # Validar y convertir la fecha
                fecha = parse_fecha(fecha_str)
                if fecha is None:
                    error_msg = f"Formato de fecha inválido en la fila {i+1}: {fecha_str}. Debe estar en un formato válido."
                    logging.error(error_msg)
                    datos['errores'] = error_msg
                    registros_incorrectos.append(datos)
                    continue

                try:

                    # Validar si el destino es válido
                    if destino not in CatalagoDestino.objects.values_list('destino', flat=True):
                        error_msg = f"El destino {destino} no está en la tabla CatalagoDestino"
                        print(error_msg)
                        datos['errores'] = error_msg
                        registros_incorrectos.append(datos)
                        continue
                    # Validar si monto_total_str no es numérico
                    if not monto_total.isdigit():
                        error_msg = f"El monto_total '{monto_total}' no es un número válido."
                        print(error_msg)
                        datos['errores'] = error_msg
                        registros_incorrectos.append(datos)
                        continue

                    
                    monto_de_inversion_municipal_float = float(monto_de_inversion_municipal)
                    monto_de_inversion_estatal_float = float(monto_de_inversion_estatal)
                    monto_de_inversion_federal_float = float(monto_de_inversion_federal)
                    monto_total_float = float(monto_total)

                    # Buscar si la fila ya existe en la base de datos
                    inventario_existente = InversionPublica.objects.filter(
                        fecha=fecha, 
                        destino=destino, 
                        nombre_de_la_obra=nombre_de_la_obra
                    )
                    
                    if inventario_existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {row} ya existe en la base de datos")
                        registros_existentes.append({ 'fecha': fecha, 'destino': destino, 'nombre_de_la_obra': nombre_de_la_obra, 'monto_de_inversion_municipal': monto_de_inversion_municipal_float, 'monto_de_inversion_estatal': monto_de_inversion_estatal_float, 'monto_de_inversion_federal': monto_de_inversion_federal_float, 'monto_total': monto_total_float})
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        inventario = InversionPublica(fecha=fecha, destino=destino, nombre_de_la_obra=nombre_de_la_obra, monto_de_inversion_municipal=monto_de_inversion_municipal_float, monto_de_inversion_estatal=monto_de_inversion_estatal_float, monto_de_inversion_federal=monto_de_inversion_federal_float, monto_total=monto_total_float)
                        inventario.save()
                        registros_correctos.append({  'fecha': fecha, 'destino': destino, 'nombre_de_la_obra': nombre_de_la_obra, 'monto_de_inversion_municipal': monto_de_inversion_municipal_float, 'monto_de_inversion_estatal': monto_de_inversion_estatal_float, 'monto_de_inversion_federal': monto_de_inversion_federal_float, 'monto_total': monto_total_float})
                except (ValueError, TypeError) as e:
                    # Si los datos no son válidos, se guarda el número de fila en la lista de registros incorrectos
                    error_msg = f"Error al procesar la fila {datos}: {e}"
                    print(error_msg)
                    datos['errores'] = error_msg
                    registros_incorrectos.append(datos)
                    
        except FileNotFoundError:
                print(f"El archivo {archivo} no se pudo abrir")
                
        return registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas
    
    def procesar_archivo_csv(self, archivo):
        archivo = self.request.FILES['archivo']
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        try:
            datos = csv.DictReader(archivo.read().decode('latin-1').splitlines())
            # print(datos)
            for row in datos:
                num_filas_procesadas += 1

                # Limpieza de datos
                destino = clean_str_col(row['destino'])

                # Homologación de datos
                destino = homologar_columna_destino(destino)

                # Validar si el destino es válido
                if destino not in CatalagoDestino.objects.values_list('destino', flat=True):
                    print(f"El destino {destino} no está en la tabla CatalagoDestino")
                    registros_incorrectos.append(row)
                    continue

                fecha = row['fecha']
                nombre_de_la_obra = row['nombre_de_la_obra']
                monto_de_inversion_municipal = row['monto_de_inversion_municipal']
                monto_de_inversion_estatal = row['monto_de_inversion_estatal']
                monto_de_inversion_federal = row['monto_de_inversion_federal']
                monto_total = row['monto_total']

                try:
                    # Validar los datos
                    fecha_str = str(fecha)
                    fecha_obj = datetime.datetime.strptime(fecha_str, '%d/%m/%Y').date()
                    monto_de_inversion_municipal_float = float(monto_de_inversion_municipal)
                    monto_de_inversion_estatal_float = float(monto_de_inversion_estatal)
                    monto_de_inversion_federal_float = float(monto_de_inversion_federal)
                    monto_total_float = float(monto_total)

                    # Buscar si la fila ya existe en la base de datos
                    inventario_existente = InversionPublica.objects.filter(fecha=fecha_obj, destino=destino, nombre_de_la_obra=nombre_de_la_obra)
                    if inventario_existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {row} ya existe en la base de datos")
                        registros_existentes.append(row)
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        inventario = InversionPublica(fecha=fecha_obj, destino=destino, nombre_de_la_obra=nombre_de_la_obra, monto_de_inversion_municipal=monto_de_inversion_municipal_float, monto_de_inversion_estatal=monto_de_inversion_estatal_float, monto_de_inversion_federal=monto_de_inversion_federal_float, monto_total=monto_total_float)
                        inventario.save()
                        registros_correctos.append(row)
                except (ValueError, TypeError) as e:
                    print(f"Error al procesar la fila {row}: {e}")
                    registros_incorrectos.append(row)
        except FileNotFoundError:
            print(f"No se encontró el archivo {archivo}")
        except Exception as e:
            print(f"Error al procesar el archivo {archivo}: {e}")
        return registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas



class DescargarArchivoInversionPublicaView(SuperAdminOrAdminMixin, LoginRequiredMixin, View):

    def crear_archivo_excel(self, registros_incorrectos):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Add headers to the worksheet
        worksheet['A1'] = 'Fecha'
        worksheet['B1'] = 'Destino'
        worksheet['C1'] = 'nombre_de_la_obra'
        worksheet['D1'] = 'monto_de_inversion_municipal'
        worksheet['E1'] = 'monto_de_inversion_estatal'
        worksheet['F1'] = 'monto_de_inversion_federal'
        worksheet['G1'] = 'monto_total'
        worksheet['H1'] = 'errores'

        # Add the incorrect rows to the worksheet
        for i, row in enumerate(registros_incorrectos):
            fila = i + 2
            # worksheet.cell(row=fila, column=1, value=row['fila'])
            worksheet.cell(row=fila, column=1, value=row['fecha'])
            worksheet.cell(row=fila, column=2, value=row['destino'])
            worksheet.cell(row=fila, column=3, value=row['nombre_de_la_obra'])
            worksheet.cell(row=fila, column=4, value=row['monto_de_inversion_municipal'])
            worksheet.cell(row=fila, column=5, value=row['monto_de_inversion_estatal'])
            worksheet.cell(row=fila, column=6, value=row['monto_de_inversion_federal'])
            worksheet.cell(row=fila, column=7, value=row['monto_total'])
            worksheet.cell(row=fila, column=8, value=row['errores'])

        # Set the column widths to auto-fit
        for column in worksheet.columns:
            max_length = 0
            column_name = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_name].width = adjusted_width

        # Create the response with the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=registros_incorrectos.xls'

        

        # workbook.save(response)
        return workbook
    
    def post(self, request, *args, **kwargs):
        # Obtener los registros incorrectos del cuerpo de la petición
        registros_incorrectos = json.loads(request.body)

        # Crear y enviar el archivo de Excel con las filas incorrectas
        workbook = self.crear_archivo_excel(registros_incorrectos)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=registros_incorrectos.xlsx'
        workbook.save(response)
        return response    
    
           
    
    