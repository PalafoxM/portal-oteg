from django.shortcuts import render, redirect
from django.views.generic import CreateView, ListView, DetailView, UpdateView, DeleteView
from django.urls import reverse_lazy
from back.models import  *
from back.forms import *
from django.http import JsonResponse, HttpResponseRedirect

from django.views import View
from django.contrib import messages
from openpyxl import load_workbook
import csv
import os
import datetime
from django.urls import reverse
import openpyxl
from django.http import HttpResponse
from django.template.loader import render_to_string




# Create your views here.
def is_ajax(request):
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'
class InventarioHoteleroEntNacListView(ListView):
    model = InventarioHoteleroEntNac
    template_name = 'back/inventario_hotelero_ent_nac/list.html'

    

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Listado de inventario hotelero'
        context['create_url'] = reverse_lazy('dashboard:inventario_hotelero_ent_nac_create')
        context['carga_masiva_url'] = reverse_lazy('dashboard:inventario_hotelero_ent_nac_carga_masiva')
        context['entity'] = 'Inventario Hotelero'
        return context

class  InventarioHoteleroEntNacCreateView(CreateView):
    model = InventarioHoteleroEntNac
    form_class = InventarioHoteleroEntNacForm
    template_name = 'back/components/create_update.html'
    success_url = reverse_lazy('dashboard:inventario_hotelero_ent_nac_list')

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        if form.is_valid():
            self.object = form.save()
            data = {
                'success': True,
                'message': 'Registro creado exitosamente.',
                'url': self.success_url
            }
            return JsonResponse(data)
        else:
            data = {
                'success': False,
                'message': 'Ha ocurrido un error al crear un registro.',
                'errors': form.errors
            }
            return JsonResponse(data)
        

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Ha ocurrido un error al crear un registro.',
            'errors': form.errors
        }
        return JsonResponse(data)

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Registro creado exitosamente.',
            'url': self.success_url
        }
        return JsonResponse(data)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Crear Registro'
        context['entity'] = 'Inventario Hotelero'
        context['list_url'] = reverse_lazy('dashboard:inventario_hotelero_ent_nac_list')
        context['action'] = 'add'
        return context

class InventarioHoteleroEntNacUpdateView( UpdateView):
    model = InventarioHoteleroEntNac
    form_class = InventarioHoteleroEntNacForm
    template_name = 'back/components/create_update.html'
    success_url = reverse_lazy('dashboard:inventario_hotelero_ent_nac_list')

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Ha ocurrido un error al editar el registro.',
            'errors': form.errors
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Registro creado exitosamente.',
            'url': self.success_url
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Editar Inventario Hotelero'
        context['entity'] = 'Inventario Hotelero'
        context['list_url'] = reverse_lazy('dashboard:inventario_hotelero_ent_nac_list')
        context['form'] = self.form_class(instance=self.object)
        return context

class InventarioHoteleroEntNacDeleteView(DeleteView):
    model = InventarioHoteleroEntNac
    # template_name = 'back/delete.html'
    success_url = reverse_lazy('dashboard:inventario_hotelero_ent_nac_list')
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        success_url = self.get_success_url()
        self.object.delete()
        return HttpResponseRedirect(success_url)

class InventarioHoteleroEntNacCargaMasivaView(View):
    form_class = CargaMasivaForm
    template_name = 'back/inventario_hotelero_ent_nac/carga_masiva.html'
    success_url = reverse_lazy('dashboard:inventario_hotelero_ent_nac_list')

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form})


    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        archivo = request.FILES.get('archivo', None)
        if archivo:
            extension = os.path.splitext(archivo.name)[1]
            if extension == '.xlsx':
                registros_correctos, registros_incorrectos, registros_existentes = self.procesar_archivo_xlsx(archivo)
            elif extension == '.csv':
                registros_correctos, registros_incorrectos, registros_existentes = self.procesar_archivo_csv(archivo)
            else:
                messages.error(request, 'El archivo debe ser un archivo .xlsx o .csv')
                registros_incorrectos.append("El archivo debe ser un archivo .xlsx o .csv")
        else:
            messages.error(request, 'Debe seleccionar un archivo')
            registros_incorrectos.append("Debe seleccionar un archivo")

        if len(registros_incorrectos) > 0 or len(registros_existentes) > 0:
            messages.error(request, 'Hay errores de registros')
            # Cree y envíe el archivo de Excel con las filas incorrectas
            workbook = self.crear_archivo_excel(registros_incorrectos)
            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=registros_incorrectos.xlsx'
            workbook.save(response)
            
            return render(request, self.template_name, {
                'form': form,
                'registros_correctos': registros_correctos,
                'registros_incorrectos': registros_incorrectos,
                'registros_existentes': registros_existentes,
                'descargar_archivo': True
            })

            
        else:
            return HttpResponseRedirect(reverse('dashboard:inventario_hotelero_ent_nac_list'))
        
        return render(request, self.template_name, {
            'form': form,
            'registros_correctos': registros_correctos,
            'registros_incorrectos': registros_incorrectos,
            'registros_existentes': registros_existentes,
        })

    def procesar_archivo_xlsx(self, archivo):
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        try:
            workbook = load_workbook(filename=archivo, read_only=True)
            worksheet = workbook.active
            filas = list(worksheet.rows)
            for i, row in enumerate(filas):
                if i == 0:
                    continue # Ignorar la primera fila si es el encabezado
                destino = row[0].value
                fecha = row[1].value.date()
                categoria = row[2].value
                habitaciones = row[3].value
                establecimientos = row[4].value

                try:
                    # Validar los datos
                    fecha_str = str(fecha)
                    fecha_obj = datetime.datetime.strptime(fecha_str, '%Y-%m-%d').date()
                    # fecha_obj = datetime.datetime.strptime(fecha, '%Y-%m-%d').date()
                    habitaciones_int = int(habitaciones)
                    establecimientos_int = int(establecimientos)

                    # Buscar si la fila ya existe en la base de datos
                    inventario_existente = InventarioHoteleroEntNac.objects.filter(destino=destino, fecha=fecha_obj, categoria=categoria)
                    if inventario_existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {row} ya existe en la base de datos")
                        registros_existentes.append({'fila': i, 'destino': destino, 'fecha': fecha_obj, 'categoria': categoria, 'habitaciones': habitaciones_int, 'establecimientos': establecimientos_int})
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        inventario = InventarioHoteleroEntNac(destino=destino, fecha=fecha_obj, categoria=categoria, habitaciones=habitaciones_int, establecimientos=establecimientos_int)
                        inventario.save()
                        registros_correctos.append({'fila': i, 'destino': destino, 'fecha': fecha_obj, 'categoria': categoria, 'habitaciones': habitaciones_int, 'establecimientos': establecimientos_int})
                except (ValueError, TypeError) as e:
                    # Si los datos no son válidos, se guarda el número de fila en la lista de registros incorrectos
                    registros_incorrectos.append({'fila': i, 'destino': destino, 'fecha': fecha, 'categoria': categoria, 'habitaciones': habitaciones, 'establecimientos': establecimientos, 'error': str(e)})
                    
        except FileNotFoundError:
                print(f"El archivo {archivo} no se pudo abrir")
                
        return registros_correctos, registros_incorrectos, registros_existentes
    
    def procesar_archivo_csv(self, archivo):
        archivo = self.request.FILES['archivo']
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        try:
            datos = csv.DictReader(archivo.read().decode('latin-1').splitlines())
            # print(datos)
            for row in datos:
                destino = row['destino']
                fecha = row['fecha']
                categoria = row['categoria']
                habitaciones = row['habitaciones']
                establecimientos = row['establecimientos']

                try:
                    # Validar los datos
                    fecha_str = str(fecha)
                    fecha_obj = datetime.datetime.strptime(fecha_str, '%d/%m/%Y').date()
                    habitaciones_int = int(habitaciones)
                    establecimientos_int = int(establecimientos)

                    # Buscar si la fila ya existe en la base de datos
                    inventario_existente = InventarioHoteleroEntNac.objects.filter(destino=destino, fecha=fecha_obj, categoria=categoria)
                    if inventario_existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {row} ya existe en la base de datos")
                        registros_existentes.append(row)
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        inventario = InventarioHoteleroEntNac(destino=destino, fecha=fecha_obj, categoria=categoria, habitaciones=habitaciones_int, establecimientos=establecimientos_int)
                        inventario.save()
                        registros_correctos.append(row)
                except (ValueError, TypeError) as e:
                    print(f"Error al procesar la fila {row}: {e}")
                    registros_incorrectos.append(row)
        except FileNotFoundError:
            print(f"No se encontró el archivo {archivo}")
        except Exception as e:
            print(f"Error al procesar el archivo {archivo}: {e}")
        return registros_correctos, registros_incorrectos, registros_existentes
    
    def crear_archivo_excel(self, registros_incorrectos):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Add headers to the worksheet
        worksheet['A1'] = 'Fila'
        worksheet['B1'] = 'Destino'
        worksheet['C1'] = 'Fecha'
        worksheet['D1'] = 'Categoría'
        worksheet['E1'] = 'Habitaciones'
        worksheet['F1'] = 'Establecimientos'
        worksheet['G1'] = 'Error'

        # Add the incorrect rows to the worksheet
        for i, row in enumerate(registros_incorrectos):
            fila = i + 2
            # worksheet.cell(row=fila, column=1, value=row['fila'])
            worksheet.cell(row=fila, column=2, value=row['destino'])
            worksheet.cell(row=fila, column=3, value=row['fecha'])
            worksheet.cell(row=fila, column=4, value=row['categoria'])
            worksheet.cell(row=fila, column=5, value=row['habitaciones'])
            worksheet.cell(row=fila, column=6, value=row['establecimientos'])
            # worksheet.cell(row=fila, column=7, value=row['error'])

        # Set the column widths to auto-fit
        for column in worksheet.columns:
            max_length = 0
            column_name = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_name].width = adjusted_width

        # Create the response with the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=registros_incorrectos.xls'

        

        # workbook.save(response)
        return workbook

    
    
    
    