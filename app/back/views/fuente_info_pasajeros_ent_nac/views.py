from typing import Any, Dict
from django.shortcuts import render
from django.http import HttpRequest, JsonResponse, HttpResponseRedirect
from django.urls import reverse_lazy, reverse
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from back.models import *
from back.forms import *
from web.models import *
from django.http import HttpResponse
from django.contrib.auth.mixins import LoginRequiredMixin
from django.template.loader import render_to_string
# para archivo excel
from django.views import View
from django.contrib import messages
from openpyxl import load_workbook
import csv
import os
from datetime import datetime
from django.urls import reverse
import openpyxl
from django.http import HttpResponse
import json
from config.diccionarios import clean_str_col, homologar_columna_categoria, homologar_columna_destino
from django.contrib.auth.decorators import login_required, permission_required
from django.utils.decorators import method_decorator
from django.http import Http404
from django.core.exceptions import PermissionDenied
from back.mixins import *
from django.contrib.auth.decorators import user_passes_test

def es_admin_o_superadmin(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)


def is_ajax(request):
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'



class FuenteInfoPasajerosEntNacView(SuperAdminOrAdminMixin, LoginRequiredMixin, ListView):
    model = Pasajeros_Ent_Nac
    template_name = 'back/pasajeros_ent_nac/viewer.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Listado de Fuentes de Informacion de Pasajeros Entrada Nacional'
        context['create_url'] = reverse_lazy(
            'dashboard:fuente_info_pasajeros_ent_nac_create')
        context['entity'] = 'Fuentes de Informacion de Pasajeros Entrada Nacional'
        context['is_fuente'] = True
        context['carga_masiva_url'] = reverse_lazy('dashboard:fuente_info_pasajeros_ent_nac_carga_masiva')
        return context


class FuenteInfoPasajerosEntNacCreate(LoginRequiredMixin, CreateView):
    model = Pasajeros_Ent_Nac
    form_class = PasajerosEntNacForm
    template_name = 'back/pasajeros_ent_nac/create.html'
    success_url = reverse_lazy('dashboard:fuente_info_pasajeros_ent_nac')

    def get_object(self, **kwargs):
        queryset = self.get_queryset()
        try:
            return queryset.get(**kwargs)
        except queryset.model.DoesNotExist:
            return None

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        replace_data = request.POST.get('replace_data')
        if form.is_valid():
            # Check if there is already a record with the same fecha, destino and categoria

            aereopuerto = form.cleaned_data['aereopuerto']
            entidad = form.cleaned_data['entidad']
            ano = form.cleaned_data['ano']

            try:
                existing_object = self.get_object(
                    aereopuerto=aereopuerto, entidad=entidad, ano=ano)

            except Certificacion.DoesNotExist:
                existing_object = None

            existing_catalogo = CatalogoAeropuertos.objects.filter(aereopuerto__iexact=aereopuerto).exists()
            existing_catalogo2 = CatalogoAeropuertos.objects.filter(
                entidad__iexact=entidad).exists()
            # ALTER TABLE mytable MODIFY mycolumn VARCHAR(255) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci;

            # If there is no existing data, save the new data
            if not existing_catalogo or not existing_catalogo2:

                if not existing_catalogo and not existing_catalogo2:
                    data = {
                        'success': False,
                        'missingData': True,
                        'destino': aereopuerto,
                        'categoria': entidad,
                        'message': 'No existe el aeropuerto ni la entidad en el catalogo',
                    }
                    return JsonResponse(data)

                if not existing_catalogo:
                    data = {
                        'success': False,
                        'missingData': True,
                        'destino': aereopuerto,
                        'message': 'No existe el aeropuerto en el catalogo',
                    }
                    return JsonResponse(data)

                if not existing_catalogo2:
                    data = {
                        'success': False,
                        'missingData': True,
                        'categoria': entidad,
                        'message': 'No existe la entidad en el catalogo',
                    }
                    return JsonResponse(data)

            # If there is existing data and replace_data is True, delete the existing data
            if existing_object and request.POST.get('replace_data') == 'on':
                # existing_object.delete()
                # Save the new data
                # self.object = form.save()

                for field in form.cleaned_data:
                    if field == 'replace_data':
                        continue
                    setattr(existing_object, field, form.cleaned_data[field])

                existing_object.save()

                data = {
                    'success': True,
                    'message': 'Data created successfully.',
                    'url': self.success_url,

                }
                return JsonResponse(data)

            # If there is existing data and replace_data is False, return an error

            if existing_object:
                data = Pasajeros_Ent_Nac.objects.filter(
                    aereopuerto=aereopuerto, entidad=entidad, ano=ano)

                data_list = list(data.values('aereopuerto', 'entidad', 'ano', 'nacionales', 'regulares', 'nacionales_regulares',
                                             'internacionales_regulares', 'charters', 'charters_nacionales', 'charters_internacionales'))

                data_list2 = list(form.cleaned_data.values())
                table_html = render_to_string('back/pasajeros_ent_nac/table.html', {
                                              'data_list': data_list, 'actual': True, 'data_list2': data_list2})

                datajsn = {
                    'success': False,
                    'message': 'Hubo un error al crear registro.',
                    'errors': 'Ya existe un registro con la misma fecha, destino , PasajerosEntNac y museo o zona arqueologica',
                    'existing_object': table_html
                }

                return JsonResponse(datajsn)
            else:
                self.object = form.save()
                data = {
                    'success': True,
                    'message': 'Data created successfully.',
                    'url': self.success_url
                }
                return JsonResponse(data)
        else:
            data = {
                'success': False,
                'message': 'Error creating data.',
                'errors': form.errors,
                'format_errors': True
            }
            return JsonResponse(data)

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Hubo un error al crear registro.',
            'errors': form.errors
        }
        return JsonResponse(data)

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Registro creado exitosamente.',
            'url': self.success_url
        }
        return JsonResponse(data)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Crear una fuente'
        context['entity'] = 'Glosario'
        context['list_url'] = reverse_lazy('dashboard:fuente_info_pasajeros_ent_nac')
        context['action'] = 'add'
        return context


class FuenteInfoPasajerosEntNacUpdate (UpdateView):
    model = Pasajeros_Ent_Nac
    form_class =    PasajerosEntNacForm
    template_name = 'back/fuente_info_certificacion/view_editor.html'
    success_url = reverse_lazy('dashboard:fuente_info_certificacion')

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Hubo un error al subir los datos.',
            'errors': form.errors
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Fuente creada exitosamente.',
            'url': self.success_url
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)

        context['list_url'] = reverse_lazy('dashboard:fuente_info_pasajeros_ent_nac')
        # Set the widget for the 'destino' field to read-only text input

        context['form'].fields['ano'].widget.attrs['readonly'] = True
        context['form'].fields['aereopuerto'].widget.attrs['readonly'] = True
        context['form'].fields['entidad'].widget.attrs['readonly'] = True

        context['title'] = 'Editar fuente de informacion'

        context['edit_msg'] = 'Los Campos Año , Aeropuerto y Entidad no se pueden editar'
        return context


class FuenteInfoPasajerosEntNacDelete (DeleteView):
    model = Pasajeros_Ent_Nac
    success_url = reverse_lazy('dashboard:fuente_info_pasajeros_ent_nac')
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object()

        try:
            self.object.delete()
            return JsonResponse({'message': 'Eliminación exitosa.'})
        except Exception as e:
            return JsonResponse({'error': 'Error al eliminar el registro.'}, status=500)


class PasajerosEntNacCargaMasivaView(SuperAdminOrAdminMixin, LoginRequiredMixin, View):
    form_class = CargaMasivaForm
    template_name = 'back/pasajeros_ent_nac/carga_masiva.html'
    success_url = reverse_lazy('dashboard:fuente_info_pasajeros_ent_nac')

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form, 'title': 'Carga Masiva de Pasajeros Ent Nac'})

    def convert_to_serializable(self, obj):
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        raise TypeError(
            f'Object of type {obj.__class__.__name__} is not JSON serializable')

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        archivo = request.FILES.get('archivo', None)
        if archivo:
            extension = os.path.splitext(archivo.name)[1]
            if extension == '.xlsx':
                registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas = self.procesar_archivo_xlsx(
                    archivo)
            elif extension == '.csv':
                registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas = self.procesar_archivo_csv(
                    archivo)
            else:
                messages.error(
                    request, 'El archivo debe ser un archivo .xlsx o .csv')
                registros_incorrectos.append(
                    "El archivo debe ser un archivo .xlsx o .csv")
        else:
            messages.error(request, 'Debe seleccionar un archivo')
            registros_incorrectos.append("Debe seleccionar un archivo")

        if len(registros_incorrectos) > 0 or len(registros_existentes) > 0:
            messages.error(request, 'Hay errores de registros')
            datos_json = json.dumps(
                registros_incorrectos, default=self.convert_to_serializable)

            return render(request, self.template_name, {
                'form': form,
                'title': 'Carga Masiva de PasajerosEntNac',
                'registros_correctos': registros_correctos,
                'registros_incorrectos': registros_incorrectos,
                'registros_existentes': registros_existentes,
                'descargar_url': datos_json,
                'num_filas_procesadas': num_filas_procesadas,
            })

        else:
            return HttpResponseRedirect(reverse('dashboard:fuente_info_pasajeros_ent_nac'))

    def procesar_archivo_xlsx(self, archivo):
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        try:
            workbook = load_workbook(filename=archivo, read_only=True)
            worksheet = workbook.active
            filas = list(worksheet.rows)
            for i, row in enumerate(filas):
                if i == 0:
                    continue # Ignorar la primera fila si es el encabezado

                if not row or all(cell.value is None for cell in row):
                    continue  # Salta filas vacías

                num_filas_procesadas += 1

                # Limpieza de datos
                entidad = clean_str_col(row[1].value)

                aereopuerto = row[0].value
                 
                ano = row[2].value
                nacionales = row[3].value
                internacionales = row[4].value
                regulares = row[5].value
                nacionales_regulares = row[6].value
                internacionales_regulares = row[7].value
                charters = row[8].value
                charters_nacionales = row[9].value
                charters_internacionales = row[10].value


                datos = {
                    'aereopuerto':aereopuerto,
                    'entidad':entidad,
                    'ano':ano,
                    'nacionales':nacionales,
                    "internacionales":internacionales,
                    'regulares':regulares,
                    'nacionales_regulares':nacionales_regulares,
                    'internacionales_regulares':internacionales_regulares,
                    'charters':charters,
                    'charters_nacionales':charters_nacionales,
                    'charters_internacionales':charters_internacionales,
                }

                try:
                    if entidad not in CatalogoEntidad.objects.values_list('entidad', flat=True):
                        print(f"La entidad {entidad} no está en la tabla CatalagoDestinoAeropuerto")
                        registros_incorrectos.append(datos)
                        continue

                    # Buscar si la fila ya existe en la base de datos
                    existente = Pasajeros_Ent_Nac.objects.filter(
                        aereopuerto = aereopuerto,
                        entidad = entidad,
                        ano = ano,
                    )
                    if existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {row} ya existe en la base de datos")
                        registros_existentes.append(datos)
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        db = Pasajeros_Ent_Nac(
                            aereopuerto = aereopuerto,
                            entidad = entidad,
                            ano = ano,
                            nacionales = nacionales,
                            internacionales = internacionales,
                            regulares = regulares,
                            nacionales_regulares = nacionales_regulares,
                            internacionales_regulares = internacionales_regulares,
                            charters = charters,
                            charters_nacionales = charters_nacionales,
                            charters_internacionales = charters_internacionales,
                        )
                        db.save()
                        registros_correctos.append(datos)
                except (ValueError, TypeError) as e:
                    print(f"Error al procesar la fila {datos}: {e}")
                    registros_incorrectos.append(datos)

        except FileNotFoundError:
            print(f"El archivo {archivo} no se pudo abrir")

        return registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas

    def procesar_archivo_csv(self, archivo):
        archivo = self.request.FILES['archivo']
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        try:
            datos = csv.DictReader(
                archivo.read().decode('latin-1').splitlines())
            # print(datos)
            for row in datos:
                num_filas_procesadas += 1

                # Limpieza de datos
                entidad = clean_str_col(row['entidad'])

                aereopuerto = row['aereopuerto'] 
                 
                ano = row['ano'] 
                nacionales = row['nacionales'] 
                regulares = row['regulares'] 
                nacionales_regulares = row['nacionales_regulares'] 
                internacionales_regulares = row['internacionales_regulares'] 
                charters = row['charters'] 
                charters_nacionales = row['charters_nacionales'] 
                charters_internacionales = row['charters_internacionales'] 


                datos = {
                    'aereopuerto':aereopuerto,
                    'entidad':entidad,
                    'ano':ano,
                    'nacionales':nacionales,
                    'regulares':regulares,
                    'nacionales_regulares':nacionales_regulares,
                    'internacionales_regulares':internacionales_regulares,
                    'charters':charters,
                    'charters_nacionales':charters_nacionales,
                    'charters_internacionales ':charters_internacionales,
                }

                try:
                    if entidad not in CatalagoDestino.objects.values_list('entidad', flat=True):
                        print(f"La entidad {entidad} no está en la tabla CatalagoDestinoAeropuerto")
                        registros_incorrectos.append(datos)
                        continue

                    # Buscar si la fila ya existe en la base de datos
                    existente = Pasajeros_Ent_Nac.objects.filter(
                        aereopuerto = aereopuerto,
                        entidad = entidad,
                        ano = ano,
                    )
                    if existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {row} ya existe en la base de datos")
                        registros_existentes.append(datos)
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        db = Pasajeros_Ent_Nac(
                            aereopuerto = aereopuerto,
                            entidad = entidad,
                            ano = ano,
                            nacionales = nacionales,
                            regulares = regulares,
                            nacionales_regulares = nacionales_regulares,
                            internacionales_regulares = internacionales_regulares,
                            charters = charters,
                            charters_nacionales = charters_nacionales,
                            charters_internacionales = charters_internacionales,
                        )
                        db.save()
                        registros_correctos.append(datos)
                except (ValueError, TypeError) as e:
                    print(f"Error al procesar la fila {datos}: {e}")
                    registros_incorrectos.append(datos)
        except FileNotFoundError:
            print(f"No se encontró el archivo {archivo}")
        except Exception as e:
            print(f"Error al procesar el archivo {archivo}: {e}")
        return registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas



class PasajerosEntNacDescargarArchivoView(SuperAdminOrAdminMixin, LoginRequiredMixin, View):

    def crear_archivo_excel(self, registros_incorrectos):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Obtener los nombres y verbose_name de los campos del modelo Pasajeros_Ent_Nac
        fields = Pasajeros_Ent_Nac._meta.get_fields()
        column_labels = [field.verbose_name for field in fields if field.name != 'id']
        column_names = [field.name for field in fields if field.name != 'id']

        # Escribir los encabezados de las columnas
        for i, campo in enumerate(column_labels):
            columna = i + 1
            worksheet.cell(row=1, column=columna, value=campo)

        # Obtener los datos del modelo Pasajeros_Ent_Nac
        datos = registros_incorrectos

        # Escribir los valores en las celdas correspondientes
        fila = 2
        for registro in datos:
            for i, campo in enumerate(column_names):
                if campo != 'id':  # Omitir la clave 'id'
                    columna = i + 1
                    valor = registro[campo]
                    worksheet.cell(row=fila, column=columna, value=valor)
            fila += 1

        # Set the column widths to auto-fit
        for column in worksheet.columns:
            max_length = 0
            column_name = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_name].width = adjusted_width

        # Create the response with the Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=gasto_derrama_registros_incorrectos.xls'

        # workbook.save(response)
        return workbook

    def post(self, request, *args, **kwargs):
        # Obtener los registros incorrectos del cuerpo de la petición
        registros_incorrectos = json.loads(request.body)

        # Crear y enviar el archivo de Excel con las filas incorrectas
        workbook = self.crear_archivo_excel(registros_incorrectos)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=gasto_derrama_registros_incorrectos.xlsx'
        workbook.save(response)
        return response