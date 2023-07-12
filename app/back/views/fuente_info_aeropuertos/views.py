from typing import Any, Dict
from django.shortcuts import render
from django.http import HttpRequest, JsonResponse, HttpResponseRedirect
from django.urls import reverse_lazy, reverse
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from back.models import *
from back.forms import *
from web.models import *
from django.http import HttpResponse
from django.contrib.auth.mixins import LoginRequiredMixin
from django.template.loader import render_to_string
# para archivo excel
from django.views import View
from django.contrib import messages
from openpyxl import load_workbook
import csv
import os
from datetime import datetime
from django.urls import reverse
import openpyxl
from django.http import HttpResponse
import json
from config.diccionarios import clean_str_col, homologar_columna_categoria, homologar_columna_destino
from django.contrib.auth.decorators import login_required, permission_required
from django.utils.decorators import method_decorator
from django.http import Http404
from django.core.exceptions import PermissionDenied
from back.mixins import *
from django.contrib.auth.decorators import user_passes_test

def es_admin_o_superadmin(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)


def is_ajax(request):
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'



class FuenteInfoAeropuerto(SuperAdminOrAdminMixin, LoginRequiredMixin, ListView):
    model = Aeropuerto
    template_name = 'back/fuente_info_aeropuertos/list.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Listado de Fuentes de Informacion de Aeropuerto'
        context['create_url'] = reverse_lazy('dashboard:fuente_info_aeropuertos_create')
        context['entity'] = 'Aeropuerto'
        context['is_fuente'] = True
        context['carga_masiva_url'] = reverse_lazy('dashboard:fuente_info_aeropuertos_carga_masiva')
        return context  


class FuenteInfoAeropuertoCreate(SuperAdminOrAdminMixin, LoginRequiredMixin, CreateView):
    model = Aeropuerto
    form_class = AeropuertoForm 
    template_name = 'back/fuente_info_aeropuertos/create.html'
    success_url = reverse_lazy('dashboard:fuente_info_aeropuertos')

    # def get_object(self, **kwargs):
    #     queryset = self.get_queryset()
    #     try:
    #         return queryset.get(**kwargs)
    #     except queryset.model.DoesNotExist:
    #         return None

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        replace_data = request.POST.get('replace_data')
        if form.is_valid():
            # Check if there is already a record with the same fecha, destino and categoria

            fecha = form.cleaned_data['fecha']

        
            try:
                existing_object = self.get_object(fecha=fecha)

            except Aeropuerto.DoesNotExist:
                existing_object = None

            # If there is existing data and replace_data is True, delete the existing data
            if existing_object and request.POST.get('replace_data') == 'on':
                # existing_object.delete()
                # Save the new data
                # self.object = form.save()

                for field in form.cleaned_data:
                    if field == 'replace_data':
                        continue
                    setattr(existing_object, field, form.cleaned_data[field])

                existing_object.save()

                data = {
                    'success': True,
                    'message': 'Data created successfully.',
                    'url': self.success_url,

                }
                return JsonResponse(data)

            # If there is existing data and replace_data is False, return an error

            if existing_object:
                data = Aeropuerto.objects.filter(fecha=fecha)
                data_list = list(data.values('fecha', 'pasajeros_aeropuerto_gto', 'pasajeros_nacionales', 'pasajeros_internacionales', 'vuelos'))
                data_list2 = list(form.cleaned_data.values())
                table_html = render_to_string('back/fuente_info_aeropuertos/table.html',{'data_list': data_list, 'actual': True, 'data_list2': data_list2})

                datajsn = {
                    'success': False,
                    'message': 'Hubo un error al crear registro.',
                    'errors': 'Ya existe un registro con la misma fecha, destino , origen y museo o zona arqueologica',
                    'existing_object': table_html
                }

                return JsonResponse(datajsn)
            else:
                self.object = form.save()
                data = {
                    'success': True,
                    'message': 'Data created successfully.',
                    'url': self.success_url
                }
                return JsonResponse(data)
        else:
            data = {
                'success': False,
                'message': 'Error creating data.',
                'errors': form.errors,
                'format_errors': True
            }
            return JsonResponse(data)

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Hubo un error al crear registro.',
            'errors': form.errors
        }
        return JsonResponse(data)

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Registro creado exitosamente.',
            'url': self.success_url
        }
        return JsonResponse(data)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Crear una fuente Aeropuerto'
        context['entity'] = 'Aeropuerto'
        context['list_url'] = reverse_lazy('dashboard:fuente_info_aeropuertos')
        context['action'] = 'add'
        return context



class FuenteInfoAeropuertoUpdate(SuperAdminOrAdminMixin, LoginRequiredMixin, UpdateView):
    model = Aeropuerto
    form_class = AeropuertoForm
    template_name = 'back/fuente_info_aeropuertos/view_editor.html'
    success_url = reverse_lazy('dashboard:fuente_info_aeropuertos')

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Hubo un error al subir los datos.',
            'errors': form.errors
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Fuente creada exitosamente.',
            'url': self.success_url
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['list_url'] = reverse_lazy('dashboard:fuente_info_aeropuertos')
            # Set the widget for the 'destino' field to read-only text input
        context['form'].fields['fecha'].widget.attrs['readonly'] = True

        context['title'] = 'Editar fuente'
        context['edit_msg'] = 'El campo Año no pude ser editado.'

        return context
    


class FuenteInfoAeropuertoDelete(SuperAdminOrAdminMixin, LoginRequiredMixin, DeleteView):
    model = Aeropuerto
    success_url = reverse_lazy('dashboard:fuente_info_aeropuertos')
    
    def post(self, request: HttpRequest, *args: str, **kwargs: Any) -> HttpResponse:
        return super().post(request, *args, **kwargs)


class AeropuertoCargaMasivaView(SuperAdminOrAdminMixin, LoginRequiredMixin, View):
    form_class = CargaMasivaForm
    template_name = 'back/fuente_info_aeropuertos/carga_masiva.html'
    success_url = reverse_lazy('dashboard:fuente_info_aeropuertos')

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form, 'title': 'Carga Masiva de Aeropuerto'})


    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        archivo = request.FILES.get('archivo', None)
        if archivo:
            extension = os.path.splitext(archivo.name)[1]
            if extension == '.xlsx':
                registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas = self.procesar_archivo_xlsx(archivo)
            elif extension == '.csv':
                registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas = self.procesar_archivo_csv(archivo)
            else:
                messages.error(request, 'El archivo debe ser un archivo .xlsx o .csv')
                registros_incorrectos.append("El archivo debe ser un archivo .xlsx o .csv")
        else:
            messages.error(request, 'Debe seleccionar un archivo')
            registros_incorrectos.append("Debe seleccionar un archivo")

        if len(registros_incorrectos) > 0 or len(registros_existentes) > 0:
            messages.error(request, 'Hay errores de registros')
            datos_json = json.dumps(registros_incorrectos)
            
            return render(request, self.template_name, {
                'form': form,
                'title': 'Carga Masiva de Aeropuerto',
                'registros_correctos': registros_correctos,
                'registros_incorrectos': registros_incorrectos,
                'registros_existentes': registros_existentes,
                'descargar_url': datos_json,
                'num_filas_procesadas': num_filas_procesadas,
            })
            
        else:
            return HttpResponseRedirect(reverse('dashboard:fuente_info_aeropuertos'))
        
    def procesar_archivo_xlsx(self, archivo):
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        try:
            workbook = load_workbook(filename=archivo, read_only=True)
            worksheet = workbook.active
            filas = list(worksheet.rows)
            for i, row in enumerate(filas):
                if i == 0:
                    continue # Ignorar la primera fila si es el encabezado
                num_filas_procesadas += 1

                fecha_str = row[3].value.date().strftime('%Y-%m-%d') if len(row) > 0 and row[0].value else ''
                fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date() if fecha_str else ''
                
                vuelos = row[4].value if len(row) > 5 else 0
                pasajeros_aeropuerto_gto = row[0].value if len(row) > 0 else 0
                pasajeros_nacionales =row[1].value if len(row) > 1 else 0
                pasajeros_internacionales = row[2].value if len(row) > 2 else 0

                datos = {
                    "fecha": fecha_str,
                    "vuelos": vuelos,
                    "pasajeros_aeropuerto_gto": pasajeros_aeropuerto_gto,
                    "pasajeros_nacionales": pasajeros_nacionales,
                    "pasajeros_internacionales": pasajeros_internacionales,
                }

                try:

                    # Buscar si la fila ya existe en la base de datos
                    existente = Aeropuerto.objects.filter(
                        fecha = fecha_str, 
                        vuelos = vuelos, 
                    )
                    if existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {row} ya existe en la base de datos")
                        registros_existentes.append(datos)
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        db = Aeropuerto(
                            fecha = fecha_obj,
                            vuelos = vuelos,
                            pasajeros_aeropuerto_gto = pasajeros_aeropuerto_gto,
                            pasajeros_nacionales = pasajeros_nacionales,
                            pasajeros_internacionales = pasajeros_internacionales,
                        )
                        db.save()
                        registros_correctos.append(datos)
                except (ValueError, TypeError) as e:
                    print(f"Error al procesar la fila {datos}: {e}")
                    registros_incorrectos.append(datos)
                    
        except FileNotFoundError:
                print(f"El archivo {archivo} no se pudo abrir")
                
        return registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas
    
    def procesar_archivo_csv(self, archivo):
        archivo = self.request.FILES['archivo']
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        try:
            datos = csv.DictReader(archivo.read().decode('latin-1').splitlines())
            # print(datos)
            for row in datos:
                num_filas_procesadas += 1

                fecha = row['fecha']
                fecha_str = str(fecha)
                fecha_str = fecha_str.split()[0] if fecha_str else ''  # Eliminar la parte de la hora si existe la fecha
                fecha_obj = datetime.datetime.strptime(fecha_str, '%Y-%m-%d').date()

                vuelos = row['vuelos']
                pasajeros_aeropuerto_gto = row['pasajeros_aeropuerto_gto']
                pasajeros_nacionales = row['pasajeros_nacionales']
                pasajeros_internacionales = row['pasajeros_internacionales']

                datos = {
                    "fecha": fecha_str,
                    "vuelos": vuelos,
                    "pasajeros_aeropuerto_gto": pasajeros_aeropuerto_gto,
                    "pasajeros_nacionales": pasajeros_nacionales,
                    "pasajeros_internacionales": pasajeros_internacionales,
                }

                try:

                    # Buscar si la fila ya existe en la base de datos
                    existente = Aeropuerto.objects.filter(
                        fecha = fecha, 
                        vuelos = vuelos, 
                    )
                    if existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {row} ya existe en la base de datos")
                        registros_existentes.append(datos)
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        db = Aeropuerto(
                            fecha = fecha,
                            vuelos = vuelos,
                            pasajeros_aeropuerto_gto = pasajeros_aeropuerto_gto,
                            pasajeros_nacionales = pasajeros_nacionales,
                            pasajeros_internacionales = pasajeros_internacionales,
                        )
                        db.save()
                        registros_correctos.append(datos)
                except (ValueError, TypeError) as e:
                    print(f"Error al procesar la fila {datos}: {e}")
                    registros_incorrectos.append(datos)
        except FileNotFoundError:
            print(f"No se encontró el archivo {archivo}")
        except Exception as e:
            print(f"Error al procesar el archivo {archivo}: {e}")
        return registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas



class AeropuertoDescargarArchivoView(SuperAdminOrAdminMixin, LoginRequiredMixin, View):

    def crear_archivo_excel(self, registros_incorrectos):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Escribir encabezados de columna
        worksheet['A1'] = 'pasajeros_aeropuerto_gto'
        worksheet['B1'] = 'pasajeros_nacionales'
        worksheet['C1'] = 'pasajeros_internacionales'
        worksheet['D1'] = 'fecha'
        worksheet['E1'] = 'vuelos'

        # Add the incorrect rows to the worksheet
        for i, row in enumerate(registros_incorrectos):
            fila = i + 2
            worksheet.cell(row=fila, column=1, value=row['pasajeros_aeropuerto_gto'])
            worksheet.cell(row=fila, column=2, value=row['pasajeros_nacionales'])
            worksheet.cell(row=fila, column=3, value=row['pasajeros_internacionales'])
            worksheet.cell(row=fila, column=4, value=row['fecha'])
            worksheet.cell(row=fila, column=5, value=row['vuelos'])



        # Set the column widths to auto-fit
        for column in worksheet.columns:
            max_length = 0
            column_name = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_name].width = adjusted_width

        # Create the response with the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=gasto_derrama_registros_incorrectos.xls'

        

        # workbook.save(response)
        return workbook
    
    def post(self, request, *args, **kwargs):
        # Obtener los registros incorrectos del cuerpo de la petición
        registros_incorrectos = json.loads(request.body)

        # Crear y enviar el archivo de Excel con las filas incorrectas
        workbook = self.crear_archivo_excel(registros_incorrectos)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=gasto_derrama_registros_incorrectos.xlsx'
        workbook.save(response)
        return response       