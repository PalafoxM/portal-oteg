from typing import Any
from django import http
from django.shortcuts import render
from django.http import HttpRequest, JsonResponse, HttpResponseRedirect
from django.urls import reverse_lazy, reverse
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from back.models import *
from back.forms import *
from web.models import *
from django.shortcuts import render
from django.http import HttpResponse
import csv
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_GET
from django.views.decorators.http import require_POST
from django.shortcuts import get_object_or_404
#serializers
# render to string
from django.template.loader import render_to_string
from django.utils.decorators import method_decorator
# para archivo excel
from django.views import View
from django.contrib import messages
from openpyxl import load_workbook
import csv
import os
from datetime import datetime
from django.urls import reverse
import openpyxl
from django.http import HttpResponse
import json
from config.diccionarios import clean_str_col, homologar_columna_categoria, homologar_columna_destino
from django.contrib.auth.decorators import login_required, permission_required
from django.utils.decorators import method_decorator
from django.http import Http404
from django.core.exceptions import PermissionDenied
from back.mixins import *
from django.contrib.auth.decorators import user_passes_test

def es_admin_o_superadmin(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)


def is_ajax(request):
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'



class FuenteInfoPerfilVisitanteDestinos (SuperAdminOrAdminMixin, LoginRequiredMixin, ListView):
    model = FuenteInfoPerfilVisitanteDestino
    template_name = 'back/fuente_info_perfil_visitante_destinos/viewer.html'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs) :
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'search':
                data = []
                for i in FuenteInfoPerfilVisitanteDestino.objects.all():
                    data.append(i.toJSON())
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)
    

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Listado de Fuentes de Información de PV Destinos'
        context['create_url'] = reverse_lazy('dashboard:fuente_info_perfil_visitante_destinos_create')
        context['entity'] = 'Fuentes de Información de PV Eventos'
        context['is_fuente'] = True
        context['carga_masiva_url'] = reverse_lazy('dashboard:fuente_info_perfil_visitante_destinos_carga_masiva')
        return context  

   
class FuenteInfoPerfilVisitanteDestinosCreate(SuperAdminOrAdminMixin, LoginRequiredMixin, CreateView):

    model = FuenteInfoPerfilVisitanteDestino    

    form_class = FuenteInfoPerfilVisitanteDestinoForm
    template_name = 'back/fuente_info_perfil_visitante_destinos/create.html'
    success_url = reverse_lazy('dashboard:fuente_info_perfil_visitante_destinos')

    def get_object(self, **kwargs):
        queryset = self.get_queryset()
        try:
            return queryset.get(**kwargs)
        except queryset.model.DoesNotExist:
            return None

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        replace_data = request.POST.get('replace_data')
        if form.is_valid():
            # Check if there is already a record with the same fecha, destino and categoria

            fecha = form.cleaned_data['fecha']
            ano = form.cleaned_data['ano']
            destino = form.cleaned_data['destino']


            try:
                existing_object = self.get_object(fecha=fecha, destino=destino, ano=ano)

            except FuenteInfoPerfilVisitanteEvento.DoesNotExist:
                existing_object = None

            existing_catalogo = CatalagoDestino.objects.filter(destino=destino).exists()
            # ALTER TABLE mytable MODIFY mycolumn VARCHAR(255) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci;

            # If there is no existing data, save the new data
            if not existing_catalogo:

                if not existing_catalogo:
                    data = {
                        'success': False,
                        'missingData': True,
                        'destino': destino,
                        'message': 'No existe el la entidad en el catalogo',
                    }
                    return JsonResponse(data)

            # If there is existing data and replace_data is True, delete the existing data
            if existing_object and request.POST.get('replace_data') == 'on':
                # existing_object.delete()
                # Save the new data
                # self.object = form.save()

                for field in form.cleaned_data:
                    if field == 'replace_data':
                        continue
                    setattr(existing_object, field, form.cleaned_data[field])

                existing_object.save()

                data = {
                    'success': True,
                    'message': 'Data created successfully.',
                    'url': self.success_url,

                }
                return JsonResponse(data)

            # If there is existing data and replace_data is False, return an error

            if existing_object:
                data =  FuenteInfoPerfilVisitanteDestino.objects.filter(fecha=fecha, destino=destino, ano=ano)
                data_list = list(data.values( 'ano',
    'folio',
    'herramienta',
    'fecha',
    'temporada',
    'destino',
    'residencia',
    'tipo_asistente',
    'municipio',
    'estado',
    'pais',
    'origen',
    'motivo_visita',
    'motivo_visita_otro',
    'segmento',
    'tipo_hospedaje',
    'tipo_visitante',
    'estadia_dias',
    'estadia_hrs',
    'acompanantes',
    'acompanantes_maxmin',
    'medio_transporte_edo',
    'tiene_fam',
    'visita_fam',
    'sat_hospedaje',
    'sat_ayb',
    'sat_atractivos',
    'sat_tours',
    'sat_central',
    'sat_aeropuerto',
    'sat_carretera',
    'sat_infotur',
    'sat_estacionamiento',
    'sat_hospitalidad',
    'sat_seguridad',
    'sat_experiencia',
    'sat_accesibilidad',
    'sat_senaletica',
    'sat_transporte',
    'sat_limpieza',
    'sat_eventos',
    'sat_protocolos',
    'sat_precios',
    'recomendacion_destino',
    'retorno_destino',
    'nps_destino',
    'nps_destino_categoria',
    'nps_hotel',
    'nps_ayb',
    'nps_atractivos',
    'nps_tours',
    'vio_escucho_noticias',
    'impacto_noticias',
    'identifico_practicas_sust',
    'edad',
    'nse',
    'sexo',
    'proposito_visita_destino_estado',
    'codigo_encuesta_ano',
    ))
                data_list2 = list(form.cleaned_data.values())
                table_html = render_to_string('back/fuente_info_perfil_visitante_destinos/table.html',{'data_list': data_list, 'actual': True, 'data_list2': data_list2})

                datajsn = {
                    'success': False,
                    'message': 'Hubo un error al crear registro.',
                    'errors': 'Ya existe un registro con la misma fecha, destino , origen y museo o zona arqueologica',
                    'existing_object': table_html
                }

                return JsonResponse(datajsn)
            else:
                self.object = form.save()
                data = {
                    'success': True,
                    'message': 'Data created successfully.',
                    'url': self.success_url
                }
                return JsonResponse(data)
        else:
            data = {
                'success': False,
                'message': 'Error creating data.',
                'errors': form.errors,
                'format_errors': True
            }
            return JsonResponse(data)

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Hubo un error al crear registro.',
            'errors': form.errors
        }
        return JsonResponse(data)

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Registro creado exitosamente.',
            'url': self.success_url
        }
        return JsonResponse(data)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Crear Fuente de Información de PV Destinos'
        context['entity'] = 'Fuentes de Información de PV Destinos'
        context['list_url'] = reverse_lazy('dashboard:fuente_info_perfil_visitante_destinos')
        context['action'] = 'add'
        return context
    


class FuenteInfoPerfilVisitanteDestinosUpdate(SuperAdminOrAdminMixin, LoginRequiredMixin, UpdateView):
    model = FuenteInfoPerfilVisitanteDestino
    form_class = FuenteInfoPerfilVisitanteDestinoForm
    template_name = 'back/fuente_info_perfil_visitante_destinos/view_editor.html'
    success_url = reverse_lazy('dashboard:fuente_info_perfil_visitante_destinos')

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Hubo un error al subir los datos.',
            'errors': form.errors
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Fuente creada exitosamente.',
            'url': self.success_url
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['list_url'] = reverse_lazy('dashboard:fuente_info_perfil_visitante_destinos')
            # Set the widget for the 'destino' field to read-only text input
        context['form'].fields['ano'].widget.attrs['readonly'] = True
        context['form'].fields['folio'].widget.attrs['readonly'] = True
        context['form'].fields['fecha'].widget.attrs['readonly'] = True
        context['form'].fields['destino'].widget.attrs['readonly'] = True
  
        context['title'] = 'Editar fuente de información de PV Destinos'

        context['edit_msg'] = 'Los Campos año , folio, fecha, destino y nombre_evento no se pueden editar'

        return context



class FuenteInfoPerfilVisitanteDestinosDelete(SuperAdminOrAdminMixin, LoginRequiredMixin, DeleteView):
    model = FuenteInfoPerfilVisitanteDestino
    success_url = reverse_lazy('dashboard:fuente_info_perfil_visitante_destinos')
    
    def post(self, request: HttpRequest, *args: str, **kwargs: Any) -> HttpResponse:
        return super().post(request, *args, **kwargs)


class PerfilVisitanteDestinosCargaMasivaView(SuperAdminOrAdminMixin, LoginRequiredMixin, View):
    form_class = CargaMasivaForm
    template_name = 'back/fuente_info_perfil_visitante_destinos/carga_masiva.html'
    success_url = reverse_lazy('dashboard:fuente_info_perfil_visitante_destinos')

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form, 'title': 'Carga Masiva de Perfil Visitante Destinos'})


    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        archivo = request.FILES.get('archivo', None)
        if archivo:
            extension = os.path.splitext(archivo.name)[1]
            if extension == '.xlsx':
                registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas = self.procesar_archivo_xlsx(archivo)
            elif extension == '.csv':
                registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas = self.procesar_archivo_csv(archivo)
            else:
                messages.error(request, 'El archivo debe ser un archivo .xlsx o .csv')
                registros_incorrectos.append("El archivo debe ser un archivo .xlsx o .csv")
        else:
            messages.error(request, 'Debe seleccionar un archivo')
            registros_incorrectos.append("Debe seleccionar un archivo")

        if len(registros_incorrectos) > 0 or len(registros_existentes) > 0:
            messages.error(request, 'Hay errores de registros')
            datos_json = json.dumps(registros_incorrectos)
            
            return render(request, self.template_name, {
                'form': form,
                'title': 'Carga Masiva de Perfil Visitante Destinos',
                'registros_correctos': registros_correctos,
                'registros_incorrectos': registros_incorrectos,
                'registros_existentes': registros_existentes,
                'descargar_url': datos_json,
                'num_filas_procesadas': num_filas_procesadas,
            })
            
        else:
            return HttpResponseRedirect(reverse('dashboard:fuente_info_aerolinea'))
        
    def procesar_archivo_xlsx(self, archivo):
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        try:
            workbook = load_workbook(filename=archivo, read_only=True)
            worksheet = workbook.active
            filas = list(worksheet.rows)
            for i, row in enumerate(filas):
                if i == 0:
                    continue # Ignorar la primera fila si es el encabezado
                num_filas_procesadas += 1
                
                fecha_str = row[3].value.date().strftime('%Y-%m-%d') if len(row) > 2 and row[2].value else ''
                fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date() if fecha_str else ''

                # Limpieza de datos
                destino = clean_str_col(row[5].value)

                # Homologación de datos
                destino = homologar_columna_destino(destino)
                
                ano=row[0].value
                folio=row[1].value
                herramienta=row[2].value
                temporada=row[4].value
                residencia=row[6].value
                tipo_asistente=row[7].value
                municipio=row[8].value
                estado=row[9].value
                pais=row[10].value
                origen=row[11].value
                motivo_visita=row[12].value
                motivo_visita_otro=row[13].value
                segmento=row[14].value
                tipo_hospedaje=row[15].value
                tipo_visitante=row[16].value
                estadia_dias=row[17].value
                estadia_hrs=row[18].value
                acompanantes=row[19].value
                acompanantes_maxmin=row[20].value
                medio_transporte_edo=row[21].value
                tiene_fam=row[22].value
                visita_fam=row[23].value
                sat_hospedaje=row[24].value
                sat_ayb=row[25].value
                sat_atractivos=row[26].value
                sat_tours=row[27].value
                sat_central=row[28].value
                sat_aeropuerto=row[29].value
                sat_carretera=row[30].value
                sat_infotur=row[31].value
                sat_estacionamiento=row[32].value
                sat_hospitalidad=row[33].value
                sat_seguridad=row[34].value
                sat_experiencia=row[35].value
                sat_accesibilidad=row[36].value
                sat_senaletica=row[37].value
                sat_transporte=row[38].value
                sat_limpieza=row[39].value
                sat_eventos=row[40].value
                sat_protocolos=row[41].value
                sat_precios=row[42].value
                recomendacion_destino=row[43].value
                retorno_destino=row[44].value
                nps_destino=row[45].value
                nps_destino_categoria=row[46].value
                nps_hotel=row[47].value
                nps_ayb=row[48].value
                nps_atractivos=row[49].value
                nps_tours=row[50].value
                vio_escucho_noticias=row[51].value
                impacto_noticias=row[52].value
                identifico_practicas_sust=row[53].value
                edad=row[54].value
                nse=row[55].value
                sexo=row[56].value
                proposito_visita_destino_estado=row[57].value
                codigo_encuesta_ano=row[58].value



                datos = {
                    'ano': ano,
                    'folio': folio,
                    'herramienta': herramienta,
                    'fecha': fecha_str,
                    'temporada': temporada,
                    'destino': destino,
                    'residencia': residencia,
                    'tipo_asistente': tipo_asistente,
                    'municipio': municipio,
                    'estado': estado,
                    'pais': pais,
                    'origen': origen,
                    'motivo_visita': motivo_visita,
                    'motivo_visita_otro': motivo_visita_otro,
                    'segmento': segmento,
                    'tipo_hospedaje': tipo_hospedaje,
                    'tipo_visitante': tipo_visitante,
                    'estadia_dias': estadia_dias,
                    'estadia_hrs': estadia_hrs,
                    'acompanantes': acompanantes,
                    'acompanantes_maxmin': acompanantes_maxmin,
                    'medio_transporte_edo': medio_transporte_edo,
                    'tiene_fam': tiene_fam,
                    'visita_fam': visita_fam,
                    'sat_hospedaje': sat_hospedaje,
                    'sat_ayb': sat_ayb,
                    'sat_atractivos': sat_atractivos,
                    'sat_tours': sat_tours,
                    'sat_central': sat_central,
                    'sat_aeropuerto': sat_aeropuerto,
                    'sat_carretera': sat_carretera,
                    'sat_infotur': sat_infotur,
                    'sat_estacionamiento': sat_estacionamiento,
                    'sat_hospitalidad': sat_hospitalidad,
                    'sat_seguridad': sat_seguridad,
                    'sat_experiencia': sat_experiencia,
                    'sat_accesibilidad': sat_accesibilidad,
                    'sat_senaletica': sat_senaletica,
                    'sat_transporte': sat_transporte,
                    'sat_limpieza': sat_limpieza,
                    'sat_eventos': sat_eventos,
                    'sat_protocolos': sat_protocolos,
                    'sat_precios': sat_precios,
                    'recomendacion_destino': recomendacion_destino,
                    'retorno_destino': retorno_destino,
                    'nps_destino': nps_destino,
                    'nps_destino_categoria': nps_destino_categoria,
                    'nps_hotel': nps_hotel,
                    'nps_ayb': nps_ayb,
                    'nps_atractivos': nps_atractivos,
                    'nps_tours': nps_tours,
                    'vio_escucho_noticias': vio_escucho_noticias,
                    'impacto_noticias': impacto_noticias,
                    'identifico_practicas_sust': identifico_practicas_sust,
                    'edad': edad,
                    'nse': nse,
                    'sexo': sexo,
                    'proposito_visita_destino_estado': proposito_visita_destino_estado,
                    'codigo_encuesta_ano': codigo_encuesta_ano
                }


                try:
                    if destino not in CatalagoDestino.objects.values_list('destino', flat=True):
                        print(f"El destino {destino} no está en la tabla CatalagoDestino")
                        registros_incorrectos.append(datos)
                        continue

                    # Buscar si la fila ya existe en la base de datos
                    existente = FuenteInfoPerfilVisitanteDestino.objects.filter(
                        fecha=fecha_obj, 
                        destino=destino, 
                        ano=ano
                    )
                    if existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {row} ya existe en la base de datos")
                        registros_existentes.append(datos)
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        db = FuenteInfoPerfilVisitanteDestino(
                            ano=ano,
                            folio=folio,
                            herramienta=herramienta,
                            fecha=fecha,
                            temporada=temporada,
                            destino=destino,
                            residencia=residencia,
                            tipo_asistente=tipo_asistente,
                            municipio=municipio,
                            estado=estado,
                            pais=pais,
                            origen=origen,
                            motivo_visita=motivo_visita,
                            motivo_visita_otro=motivo_visita_otro,
                            segmento=segmento,
                            tipo_hospedaje=tipo_hospedaje,
                            tipo_visitante=tipo_visitante,
                            estadia_dias=estadia_dias,
                            estadia_hrs=estadia_hrs,
                            acompanantes=acompanantes,
                            acompanantes_maxmin=acompanantes_maxmin,
                            medio_transporte_edo=medio_transporte_edo,
                            tiene_fam=tiene_fam,
                            visita_fam=visita_fam,
                            sat_hospedaje=sat_hospedaje,
                            sat_ayb=sat_ayb,
                            sat_atractivos=sat_atractivos,
                            sat_tours=sat_tours,
                            sat_central=sat_central,
                            sat_aeropuerto=sat_aeropuerto,
                            sat_carretera=sat_carretera,
                            sat_infotur=sat_infotur,
                            sat_estacionamiento=sat_estacionamiento,
                            sat_hospitalidad=sat_hospitalidad,
                            sat_seguridad=sat_seguridad,
                            sat_experiencia=sat_experiencia,
                            sat_accesibilidad=sat_accesibilidad,
                            sat_senaletica=sat_senaletica,
                            sat_transporte=sat_transporte,
                            sat_limpieza=sat_limpieza,
                            sat_eventos=sat_eventos,
                            sat_protocolos=sat_protocolos,
                            sat_precios=sat_precios,
                            recomendacion_destino=recomendacion_destino,
                            retorno_destino=retorno_destino,
                            nps_destino=nps_destino,
                            nps_destino_categoria=nps_destino_categoria,
                            nps_hotel=nps_hotel,
                            nps_ayb=nps_ayb,
                            nps_atractivos=nps_atractivos,
                            nps_tours=nps_tours,
                            vio_escucho_noticias=vio_escucho_noticias,
                            impacto_noticias=impacto_noticias,
                            identifico_practicas_sust=identifico_practicas_sust,
                            edad=edad,
                            nse=nse,
                            sexo=sexo,
                            proposito_visita_destino_estado=proposito_visita_destino_estado,
                            codigo_encuesta_ano=codigo_encuesta_ano,
                        )
                        db.save()
                        registros_correctos.append(datos)
                except (ValueError, TypeError) as e:
                    print(f"Error al procesar la fila {datos}: {e}")
                    registros_incorrectos.append(datos)
                    
        except FileNotFoundError:
                print(f"El archivo {archivo} no se pudo abrir")
                
        return registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas
    
    def procesar_archivo_csv(self, archivo):
        archivo = self.request.FILES['archivo']
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        try:
            datos = csv.DictReader(archivo.read().decode('latin-1').splitlines())
            # print(datos)
            for row in datos:
                num_filas_procesadas += 1

                fecha = row['fecha']
                fecha_str = str(fecha)
                fecha_str = fecha_str.split()[0] if fecha_str else ''  # Eliminar la parte de la hora si existe la fecha
                fecha_obj = datetime.datetime.strptime(fecha_str, '%Y-%m-%d').date()

                # Limpieza de datos
                destino = clean_str_col(row['destino'])

                # Homologación de datos
                destino = homologar_columna_destino(destino)

                ano=row['ano'],
                folio=row['folio'],
                herramienta=row['herramienta'],
                temporada=row['temporada'],
                residencia=row['residencia'],
                tipo_asistente=row['tipo_asistente'],
                municipio=row['municipio'],
                estado=row['estado'],
                pais=row['pais'],
                origen=row['origen'],
                motivo_visita=row['motivo_visita'],
                motivo_visita_otro=row['motivo_visita_otro'],
                segmento=row['segmento'],
                tipo_hospedaje=row['tipo_hospedaje'],
                tipo_visitante=row['tipo_visitante'],
                estadia_dias=row['estadia_dias'],
                estadia_hrs=row['estadia_hrs'],
                acompanantes=row['acompanantes'],
                acompanantes_maxmin=row['acompanantes_maxmin'],
                medio_transporte_edo=row['medio_transporte_edo'],
                tiene_fam=row['tiene_fam'],
                visita_fam=row['visita_fam'],
                sat_hospedaje=row['sat_hospedaje'],
                sat_ayb=row['sat_ayb'],
                sat_atractivos=row['sat_atractivos'],
                sat_tours=row['sat_tours'],
                sat_central=row['sat_central'],
                sat_aeropuerto=row['sat_aeropuerto'],
                sat_carretera=row['sat_carretera'],
                sat_infotur=row['sat_infotur'],
                sat_estacionamiento=row['sat_estacionamiento'],
                sat_hospitalidad=row['sat_hospitalidad'],
                sat_seguridad=row['sat_seguridad'],
                sat_experiencia=row['sat_experiencia'],
                sat_accesibilidad=row['sat_accesibilidad'],
                sat_senaletica=row['sat_senaletica'],
                sat_transporte=row['sat_transporte'],
                sat_limpieza=row['sat_limpieza'],
                sat_eventos=row['sat_eventos'],
                sat_protocolos=row['sat_protocolos'],
                sat_precios=row['sat_precios'],
                recomendacion_destino=row['recomendacion_destino'],
                retorno_destino=row['retorno_destino'],
                nps_destino=row['nps_destino'],
                nps_destino_categoria=row['nps_destino_categoria'],
                nps_hotel=row['nps_hotel'],
                nps_ayb=row['nps_ayb'],
                nps_atractivos=row['nps_atractivos'],
                nps_tours=row['nps_tours'],
                vio_escucho_noticias=row['vio_escucho_noticias'],
                impacto_noticias=row['impacto_noticias'],
                identifico_practicas_sust=row['identifico_practicas_sust'],
                edad=row['edad'],
                nse=row['nse'],
                sexo=row['sexo'],
                proposito_visita_destino_estado=row['proposito_visita_destino_estado'],
                codigo_encuesta_ano=row['codigo_encuesta_ano'],



                datos = {
                    'ano': ano,
                    'folio': folio,
                    'herramienta': herramienta,
                    'fecha': fecha,
                    'temporada': temporada,
                    'destino': destino,
                    'residencia': residencia,
                    'tipo_asistente': tipo_asistente,
                    'municipio': municipio,
                    'estado': estado,
                    'pais': pais,
                    'origen': origen,
                    'motivo_visita': motivo_visita,
                    'motivo_visita_otro': motivo_visita_otro,
                    'segmento': segmento,
                    'tipo_hospedaje': tipo_hospedaje,
                    'tipo_visitante': tipo_visitante,
                    'estadia_dias': estadia_dias,
                    'estadia_hrs': estadia_hrs,
                    'acompanantes': acompanantes,
                    'acompanantes_maxmin': acompanantes_maxmin,
                    'medio_transporte_edo': medio_transporte_edo,
                    'tiene_fam': tiene_fam,
                    'visita_fam': visita_fam,
                    'sat_hospedaje': sat_hospedaje,
                    'sat_ayb': sat_ayb,
                    'sat_atractivos': sat_atractivos,
                    'sat_tours': sat_tours,
                    'sat_central': sat_central,
                    'sat_aeropuerto': sat_aeropuerto,
                    'sat_carretera': sat_carretera,
                    'sat_infotur': sat_infotur,
                    'sat_estacionamiento': sat_estacionamiento,
                    'sat_hospitalidad': sat_hospitalidad,
                    'sat_seguridad': sat_seguridad,
                    'sat_experiencia': sat_experiencia,
                    'sat_accesibilidad': sat_accesibilidad,
                    'sat_senaletica': sat_senaletica,
                    'sat_transporte': sat_transporte,
                    'sat_limpieza': sat_limpieza,
                    'sat_eventos': sat_eventos,
                    'sat_protocolos': sat_protocolos,
                    'sat_precios': sat_precios,
                    'recomendacion_destino': recomendacion_destino,
                    'retorno_destino': retorno_destino,
                    'nps_destino': nps_destino,
                    'nps_destino_categoria': nps_destino_categoria,
                    'nps_hotel': nps_hotel,
                    'nps_ayb': nps_ayb,
                    'nps_atractivos': nps_atractivos,
                    'nps_tours': nps_tours,
                    'vio_escucho_noticias': vio_escucho_noticias,
                    'impacto_noticias': impacto_noticias,
                    'identifico_practicas_sust': identifico_practicas_sust,
                    'edad': edad,
                    'nse': nse,
                    'sexo': sexo,
                    'proposito_visita_destino_estado': proposito_visita_destino_estado,
                    'codigo_encuesta_ano': codigo_encuesta_ano
                }


                try:
                    if destino not in CatalagoDestino.objects.values_list('destino', flat=True):
                        print(f"El destino {destino} no está en la tabla CatalagoDestino")
                        registros_incorrectos.append(datos)
                        continue

                    # Buscar si la fila ya existe en la base de datos
                    existente = FuenteInfoPerfilVisitanteDestino.objects.filter(
                        fecha=fecha_obj, 
                        destino=destino, 
                        ano=ano
                    )
                    if existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {row} ya existe en la base de datos")
                        registros_existentes.append(datos)
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        db = FuenteInfoPerfilVisitanteDestino(
                            ano=ano,
                            folio=folio,
                            herramienta=herramienta,
                            fecha=fecha,
                            temporada=temporada,
                            destino=destino,
                            residencia=residencia,
                            tipo_asistente=tipo_asistente,
                            municipio=municipio,
                            estado=estado,
                            pais=pais,
                            origen=origen,
                            motivo_visita=motivo_visita,
                            motivo_visita_otro=motivo_visita_otro,
                            segmento=segmento,
                            tipo_hospedaje=tipo_hospedaje,
                            tipo_visitante=tipo_visitante,
                            estadia_dias=estadia_dias,
                            estadia_hrs=estadia_hrs,
                            acompanantes=acompanantes,
                            acompanantes_maxmin=acompanantes_maxmin,
                            medio_transporte_edo=medio_transporte_edo,
                            tiene_fam=tiene_fam,
                            visita_fam=visita_fam,
                            sat_hospedaje=sat_hospedaje,
                            sat_ayb=sat_ayb,
                            sat_atractivos=sat_atractivos,
                            sat_tours=sat_tours,
                            sat_central=sat_central,
                            sat_aeropuerto=sat_aeropuerto,
                            sat_carretera=sat_carretera,
                            sat_infotur=sat_infotur,
                            sat_estacionamiento=sat_estacionamiento,
                            sat_hospitalidad=sat_hospitalidad,
                            sat_seguridad=sat_seguridad,
                            sat_experiencia=sat_experiencia,
                            sat_accesibilidad=sat_accesibilidad,
                            sat_senaletica=sat_senaletica,
                            sat_transporte=sat_transporte,
                            sat_limpieza=sat_limpieza,
                            sat_eventos=sat_eventos,
                            sat_protocolos=sat_protocolos,
                            sat_precios=sat_precios,
                            recomendacion_destino=recomendacion_destino,
                            retorno_destino=retorno_destino,
                            nps_destino=nps_destino,
                            nps_destino_categoria=nps_destino_categoria,
                            nps_hotel=nps_hotel,
                            nps_ayb=nps_ayb,
                            nps_atractivos=nps_atractivos,
                            nps_tours=nps_tours,
                            vio_escucho_noticias=vio_escucho_noticias,
                            impacto_noticias=impacto_noticias,
                            identifico_practicas_sust=identifico_practicas_sust,
                            edad=edad,
                            nse=nse,
                            sexo=sexo,
                            proposito_visita_destino_estado=proposito_visita_destino_estado,
                            codigo_encuesta_ano=codigo_encuesta_ano,
                        )
                        db.save()
                        registros_correctos.append(datos)
                except (ValueError, TypeError) as e:
                    print(f"Error al procesar la fila {datos}: {e}")
                    registros_incorrectos.append(datos)
        except FileNotFoundError:
            print(f"No se encontró el archivo {archivo}")
        except Exception as e:
            print(f"Error al procesar el archivo {archivo}: {e}")
        return registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas



class PerfilVisitanteDestinosDescargarArchivoView(SuperAdminOrAdminMixin, LoginRequiredMixin, View):

    def crear_archivo_excel(self, registros_incorrectos):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Obtener los nombres y verbose_name de los campos del modelo FuenteInfoPerfilVisitanteDestino
        fields = FuenteInfoPerfilVisitanteDestino._meta.get_fields()
        column_labels = [field.verbose_name for field in fields if field.name != 'id']
        column_names = [field.name for field in fields if field.name != 'id']

        # Escribir los encabezados de las columnas
        for i, campo in enumerate(column_labels):
            columna = i + 1
            worksheet.cell(row=1, column=columna, value=campo)

        # Obtener los datos del modelo FuenteInfoPerfilVisitanteDestino
        datos = registros_incorrectos

        # Escribir los valores en las celdas correspondientes
        fila = 2
        for registro in datos:
            for i, campo in enumerate(column_names):
                if campo != 'id':  # Omitir la clave 'id'
                    columna = i + 1
                    valor = registro[campo]
                    worksheet.cell(row=fila, column=columna, value=valor)
            fila += 1



        # Set the column widths to auto-fit
        for column in worksheet.columns:
            max_length = 0
            column_name = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_name].width = adjusted_width

        # Create the response with the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=gasto_derrama_registros_incorrectos.xls'

        

        # workbook.save(response)
        return workbook
    
    def post(self, request, *args, **kwargs):
        # Obtener los registros incorrectos del cuerpo de la petición
        registros_incorrectos = json.loads(request.body)

        # Crear y enviar el archivo de Excel con las filas incorrectas
        workbook = self.crear_archivo_excel(registros_incorrectos)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=gasto_derrama_registros_incorrectos.xlsx'
        workbook.save(response)
        return response      