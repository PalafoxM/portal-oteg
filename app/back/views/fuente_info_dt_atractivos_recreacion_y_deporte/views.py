from typing import Any, Dict
from django.shortcuts import render
from django.http import HttpRequest, JsonResponse, HttpResponseRedirect
from django.urls import reverse_lazy, reverse
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from back.models import *
from back.forms import *
from web.models import *
from django.http import HttpResponse
from django.contrib.auth.mixins import LoginRequiredMixin
from django.template.loader import render_to_string
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime, date

# para archivo excel
from django.views import View
from django.contrib import messages
from openpyxl import load_workbook
import csv
import os
from django.urls import reverse
import openpyxl
from django.http import HttpResponse
import json
from config.diccionarios import clean_str_col, homologar_columna_categoria, homologar_columna_destino


def is_ajax(request):
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'


class FuenteInfoDirectorioActivosRecreacionYDeporte(ListView):
    model = DirectorioActivosRecreacionYDeporte
    template_name = 'back/fuente_info_dt_atractivos_recreacion_y_deporte/list.html'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'search':
                data = []
                for i in DirectorioActivosRecreacionYDeporte.objects.all():
                    data.append(i.toJSON())
            else:
                data.append({'error': 'Ha ocurrido un error'})
        except Exception as e:
            data.append({'error': str(e)})
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Listado  de Directorio Activos Recreacion Y Deporte'
        context['create_url'] = reverse_lazy(
            'dashboard:fuente_info_dt_atractivos_recreacion_y_deporte_create')
        context['entity'] = 'Fuentes de Informacion de DirectorioActivosRecreacionYDeporte'
        context['is_fuente'] = True
        context['carga_masiva_url'] = reverse_lazy(
            'dashboard:fuente_info_dt_atractivos_recreacion_y_deporte_carga_masiva')

        return context


class FuenteInfoDirectorioActivosRecreacionYDeporteCreate (CreateView):
    model = DirectorioActivosRecreacionYDeporte
    form_class = DirectorioActivosRecreacionYDeporteForm
    template_name = 'back/fuente_info_dt_atractivos_recreacion_y_deporte/create.html'
    success_url = reverse_lazy('dashboard:fuente_info_dt_atractivos_recreacion_y_deporte')

    def get_object(self, **kwargs):
        queryset = self.get_queryset()
        try:
            return queryset.get(**kwargs)
        except queryset.model.DoesNotExist:
            return None
        
    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        replace_data = request.POST.get('replace_data')
        if form.is_valid():
            # Check if there is already a record with the same fecha, destino and categoria

            destino = form.cleaned_data['destino']
            entidad = form.cleaned_data['entidad']
            giro = form.cleaned_data['giro']
            clave_del_giro = form.cleaned_data['clave_del_giro']

            try:
                existing_object = self.get_object(destino=destino, entidad=entidad, giro=giro, clave_del_giro=clave_del_giro)

            except DirectorioActivosRecreacionYDeporte.DoesNotExist:
                existing_object = None

            existing_catalogo = CatalagoDestino.objects.filter(
                destino__iexact=destino).exists()
            existing_catalogo_entidad = CatalogoEntidad.objects.filter(
                entidad__iexact=entidad).exists()
            # ALTER TABLE mytable MODIFY mycolumn VARCHAR(255) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci;

            # If there is no existing data, save the new data
            if not existing_catalogo or not existing_catalogo_entidad:

                if not existing_catalogo_entidad and not existing_catalogo:
                    data = {
                        'success': False,
                        'missingData': True,
                        'entidad': entidad,
                        'destino': destino,
                        'message': 'No existe la entidad y el destino en el catalogo',
                    }
                    return JsonResponse(data)

                if not existing_catalogo_entidad:
                    data = {
                        'success': False,
                        'missingData': True,
                        'entidad': entidad,
                        'message': 'No existe la entidad en el catalogo',
                    }
                    return JsonResponse(data)

                if not existing_catalogo:
                    data = {
                        'success': False,
                        'missingData': True,
                        'destino': destino,
                        'message': 'No existe el destino en el catalogo',
                    }
                    return JsonResponse(data)

            # If there is existing data and replace_data is True, delete the existing data
            if existing_object and request.POST.get('replace_data') == 'on':
                # existing_object.delete()
                # Save the new data
                # self.object = form.save()

                for field in form.cleaned_data:
                    if field == 'replace_data':
                        continue
                    setattr(existing_object, field, form.cleaned_data[field])

                existing_object.save()

                data = {
                    'success': True,
                    'message': 'Data created successfully.',
                    'url': self.success_url,

                }
                return JsonResponse(data)

            # If there is existing data and replace_data is False, return an error

            if existing_object:
                data = DirectorioActivosRecreacionYDeporte.objects.filter( destino=destino, entidad=entidad, giro=giro, clave_del_giro=clave_del_giro)   
                data_list = list(data.values('giro',
                                             'clave_del_giro',
                                             'entidad',
                                             'clave_entidad',
                                             'destino',
                                            'clave_municipio',
                                             'nombre_comercial',
                                             'razon_social',
                                             'rfc',
                                             'calle',
                                             'numero',
                                             'colonia',
                                             'codigo_postal',
                                             'lada',
                                             'telefono_1',
                                             'telefono_2',
                                             'celular',
                                             'correo_electronico',
                                             'sitio_web',
                                             'ret',
                                             'rnt',))
                

                data_list2 = list(form.cleaned_data.values())
                table_html = render_to_string('back/fuente_info_dt_atractivos_recreacion_y_deporte/table.html', {
                                              'data_list': data_list, 'actual': True, 'data_list2': data_list2})

                datajsn = {
                    'success': False,
                    'message': 'Hubo un error al crear registro.',
                    'errors': 'Ya existe un registro con la misma fecha, destino , origen y museo o zona arqueologica',
                    'existing_object': table_html
                }

                return JsonResponse(datajsn)
            else:
                self.object = form.save()
                data = {
                    'success': True,
                    'message': 'Data created successfully.',
                    'url': self.success_url
                }
                return JsonResponse(data)
        else:
            data = {
                'success': False,
                'message': 'Error creating data.',
                'errors': form.errors,
                'format_errors': True
            }
            return JsonResponse(data)

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Hubo un error al crear registro.',
            'errors': form.errors
        }
        return JsonResponse(data)

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Registro creado exitosamente.',
            'url': self.success_url
        }
        return JsonResponse(data)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Crear una fuente'
        context['entity'] = 'Glosario'
        context['list_url'] = reverse_lazy('dashboard:fuente_info_dt_atractivos_recreacion_y_deporte')
        context['action'] = 'add'
        return context


class FuenteInfoDirectorioActivosRecreacionYDeporteUpdate (UpdateView):
    model = DirectorioActivosRecreacionYDeporte
    form_class = DirectorioActivosRecreacionYDeporteForm
    template_name = 'back/fuente_info_dt_atractivos_recreacion_y_deporte/view_editor.html'
    success_url = reverse_lazy('dashboard:fuente_info_dt_atractivos_recreacion_y_deporte')

    def form_invalid(self, form):
        response = super().form_invalid(form)
        data = {
            'success': False,
            'message': 'Hubo un error al subir los datos.',
            'errors': form.errors
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def form_valid(self, form):
        response = super().form_valid(form)
        data = {
            'success': True,
            'message': 'Fuente creada exitosamente.',
            'url': self.success_url
        }
        if is_ajax(self.request):
            return JsonResponse(data)
        else:
            return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['list_url'] = reverse_lazy(
            'dashboard:fuente_info_dt_atractivos_recreacion_y_deporte')
        # Set the widget for the 'destino' field to read-only text input
        # context['form'].fields['ano'].widget = forms.TextInput(attrs={'class': 'form-control', 'readonly': 'readonly'})
        # context['form'].fields['giro'].widget = forms.TextInput(attrs={'class': 'form-control', 'readonly': 'readonly'})
        # context['form'].fields['destino'].widget = forms.TextInput(attrs={'class': 'form-control', 'readonly': 'readonly'})

        context['form'].fields['entidad'].widget = forms.TextInput(attrs={'class': 'form-control', 'readonly': 'readonly'})
        context['form'].fields['giro'].widget = forms.TextInput(attrs={'class': 'form-control', 'readonly': 'readonly'})
        context['form'].fields['destino'].widget = forms.TextInput(attrs={'class': 'form-control', 'readonly': 'readonly'})
        context['form'].fields['clave_del_giro'].widget = forms.TextInput(attrs={'class': 'form-control', 'readonly': 'readonly'})
        
        context['title'] = 'Editar fuente'
        context['edit_msg'] = 'Los Campos Destino y Año no pueden ser editados'

        return context


class FuenteInfoDirectorioActivosRecreacionYDeporteDelete (DeleteView):
    model = DirectorioActivosRecreacionYDeporte
    success_url = reverse_lazy('dashboard:fuente_info_dt_atractivos_recreacion_y_deporte')

    def post(self, request: HttpRequest, *args: str, **kwargs: Any) -> HttpResponse:
        return super().post(request, *args, **kwargs)


class DirectorioActivosRecreacionYDeporteCargaMasivaView(View):
    form_class = CargaMasivaForm
    template_name = 'back/fuente_info_dt_atractivos_recreacion_y_deporte/carga_masiva.html'
    success_url = reverse_lazy('dashboard:fuente_info_dt_atractivos_recreacion_y_deporte')

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form, 'title': 'Carga Masiva de Directorio Activos Recreacion Y Deporte'})

    def convert_to_serializable(self, obj):
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        raise TypeError(
            f'Object of type {obj.__class__.__name__} is not JSON serializable')

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        archivo = request.FILES.get('archivo', None)
        if archivo:
            extension = os.path.splitext(archivo.name)[1]
            if extension == '.xlsx':
                registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas = self.procesar_archivo_xlsx(
                    archivo)
            elif extension == '.csv':
                registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas = self.procesar_archivo_csv(
                    archivo)
            else:
                messages.error(
                    request, 'El archivo debe ser un archivo .xlsx o .csv')
                registros_incorrectos.append(
                    "El archivo debe ser un archivo .xlsx o .csv")
        else:
            messages.error(request, 'Debe seleccionar un archivo')
            registros_incorrectos.append("Debe seleccionar un archivo")

        if len(registros_incorrectos) > 0 or len(registros_existentes) > 0:
            messages.error(request, 'Hay errores de registros')
            datos_json = json.dumps(
                registros_incorrectos, default=self.convert_to_serializable)

            return render(request, self.template_name, {
                'form': form,
                'title': 'Carga Masiva de Directorio Activos Recreacion Y Deporte',
                'registros_correctos': registros_correctos,
                'registros_incorrectos': registros_incorrectos,
                'registros_existentes': registros_existentes,
                'descargar_url': datos_json,
                'num_filas_procesadas': num_filas_procesadas,
            })

        else:
            return HttpResponseRedirect(reverse('dashboard:fuente_info_dt_atractivos_recreacion_y_deporte'))

    def procesar_archivo_xlsx(self, archivo):
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        try:
            workbook = load_workbook(filename=archivo, read_only=True)
            worksheet = workbook.active
            filas = list(worksheet.rows)
            for i, row in enumerate(filas):
                if i == 0:
                    continue  # Ignorar la primera fila si es el encabezado
                num_filas_procesadas += 1

                # Limpieza de datos
                destino = clean_str_col(row[4].value)
                entidad = clean_str_col(row[2].value)

                # Homologación de datos
                destino = homologar_columna_destino(destino)

                giro = row[0].value
                clave_del_giro = row[1].value
                # entidad 
                clave_entidad = row[3].value
                # destino 
                clave_municipio = row[5].value
                nombre_comercial = row[6].value
                razon_social = row[7].value
                rfc = row[8].value
                calle = row[9].value
                numero = row[10].value
                colonia = row[11].value
                codigo_postal = row[12].value
                lada = row[13].value
                telefono_1 = row[14].value
                telefono_2 = row[15].value
                celular = row[16].value
                correo_electronico = row[17].value
                sitio_web = row[18].value
                ret = row[19].value
                rnt = row[20].value

                datos = {
                    'giro': giro,
                    'clave_del_giro': clave_del_giro,
                    'entidad': entidad,
                    'clave_entidad': clave_entidad,
                    'destino': destino,
                    'clave_municipio': clave_municipio,
                    'nombre_comercial': nombre_comercial,
                    'razon_social': razon_social,
                    'rfc': rfc,
                    'calle': calle,
                    'numero': numero,
                    'colonia': colonia,
                    'codigo_postal': codigo_postal,
                    'lada': lada,
                    'telefono_1': telefono_1,
                    'telefono_2': telefono_2,
                    'celular': celular,
                    'correo_electronico': correo_electronico,
                    'sitio_web': sitio_web,
                    'ret': ret,
                    'rnt': rnt
                }

                try:
                    if destino not in CatalagoDestino.objects.values_list('destino', flat=True):
                        print(f"El destino {destino} no está en la tabla CatalagoDestinoAeropuerto")
                        registros_incorrectos.append(datos)
                        continue
                    if entidad not in CatalogoEntidad.objects.values_list('entidad', flat=True):
                        print(f"La entidad {entidad} no está en la tabla CatalagoDestinoAeropuerto")
                        registros_incorrectos.append(datos)
                        continue

                    # Buscar si la fila ya existe en la base de datos
                    existente = DirectorioActivosRecreacionYDeporte.objects.filter(
                        giro=giro,
                        clave_del_giro=clave_del_giro,
                        entidad=entidad,
                        destino=destino,
                    )
                    if existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {row} ya existe en la base de datos")
                        registros_existentes.append(datos)
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        db = DirectorioActivosRecreacionYDeporte(
                            giro=giro,
                            clave_del_giro=clave_del_giro,
                            entidad=entidad,
                            clave_entidad=clave_entidad,
                            destino=destino,
                            clave_municipio=clave_municipio,
                            nombre_comercial=nombre_comercial,
                            razon_social=razon_social,
                            rfc=rfc,
                            calle=calle,
                            numero=numero,
                            colonia=colonia,
                            codigo_postal=codigo_postal,
                            lada=lada,
                            telefono_1=telefono_1,
                            telefono_2=telefono_2,
                            celular=celular,
                            correo_electronico=correo_electronico,
                            sitio_web=sitio_web,
                            ret=ret,
                            rnt=rnt
                        )
                        db.save()
                        registros_correctos.append(datos)
                except (ValueError, TypeError) as e:
                    print(f"Error al procesar la fila {datos}: {e}")
                    registros_incorrectos.append(datos)

        except FileNotFoundError:
            print(f"El archivo {archivo} no se pudo abrir")

        return registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas

    def procesar_archivo_csv(self, archivo):
        archivo = self.request.FILES['archivo']
        registros_correctos, registros_incorrectos, registros_existentes = [], [], []
        num_filas_procesadas = 0
        try:
            datos = csv.DictReader(
                archivo.read().decode('latin-1').splitlines())
            # print(datos)
            for row in datos:
                num_filas_procesadas += 1

                # Limpieza de datos
                destino = clean_str_col(row['destino'])
                entidad = clean_str_col(row['tipo_visitante'])

                # Homologación de datos
                destino = homologar_columna_destino(destino)

                giro = row['giro']
                clave_del_giro = row['clave_del_giro']
                # entidad = row['entidad']
                clave_entidad = row['clave_entidad']
                # destino = row['destino']
                clave_municipio = row['clave_municipio']
                nombre_comercial = row['nombre_comercial']
                razon_social = row['razon_social']
                rfc = row['rfc']
                calle = row['calle']
                numero = row['numero']
                colonia = row['colonia']
                codigo_postal = row['codigo_postal']
                lada = row['lada']
                telefono_1 = row['telefono_1']
                telefono_2 = row['telefono_2']
                celular = row['celular']
                correo_electronico = row['correo_electronico']
                sitio_web = row['sitio_web']
                ret = row['ret']
                rnt = row['rnt']

                datos = {
                    'giro': giro,
                    'clave_del_giro': clave_del_giro,
                    'entidad': entidad,
                    'clave_entidad': clave_entidad,
                    'destino': destino,
                    'clave_municipio': clave_municipio,
                    'nombre_comercial': nombre_comercial,
                    'razon_social': razon_social,
                    'rfc': rfc,
                    'calle': calle,
                    'numero': numero,
                    'colonia': colonia,
                    'codigo_postal': codigo_postal,
                    'lada': lada,
                    'telefono_1': telefono_1,
                    'telefono_2': telefono_2,
                    'celular': celular,
                    'correo_electronico': correo_electronico,
                    'sitio_web': sitio_web,
                    'ret': ret,
                    'rnt': rnt
                }

                try:
                    if destino not in CatalagoDestino.objects.values_list('destino', flat=True):
                        print(f"El destino {destino} no está en la tabla CatalagoDestinoAeropuerto")
                        registros_incorrectos.append(datos)
                        continue
                    if entidad not in CatalogoEntidad.objects.values_list('entidad', flat=True):
                        print(f"La entidad {entidad} no está en la tabla CatalagoDestinoAeropuerto")
                        registros_incorrectos.append(datos)
                        continue

                    # Buscar si la fila ya existe en la base de datos
                    existente = DirectorioActivosRecreacionYDeporte.objects.filter(
                        giro=giro,
                        clave_del_giro=clave_del_giro,
                        entidad=entidad,
                        destino=destino,
                    )
                    if existente.exists():
                        # Si ya existe, se omite la fila y se guarda en la lista de registros incorrectos
                        print(f"La fila {row} ya existe en la base de datos")
                        registros_existentes.append(datos)
                    else:
                        # Si no existe, se guarda la nueva instancia del modelo en la base de datos y se guarda en la lista de registros correctos
                        db = DirectorioActivosRecreacionYDeporte(
                            giro=giro,
                            clave_del_giro=clave_del_giro,
                            entidad=entidad,
                            clave_entidad=clave_entidad,
                            destino=destino,
                            clave_municipio=clave_municipio,
                            nombre_comercial=nombre_comercial,
                            razon_social=razon_social,
                            rfc=rfc,
                            calle=calle,
                            numero=numero,
                            colonia=colonia,
                            codigo_postal=codigo_postal,
                            lada=lada,
                            telefono_1=telefono_1,
                            telefono_2=telefono_2,
                            celular=celular,
                            correo_electronico=correo_electronico,
                            sitio_web=sitio_web,
                            ret=ret,
                            rnt=rnt
                        )
                        db.save()
                        registros_correctos.append(datos)
                except (ValueError, TypeError) as e:
                    print(f"Error al procesar la fila {datos}: {e}")
                    registros_incorrectos.append(datos)
        except FileNotFoundError:
            print(f"No se encontró el archivo {archivo}")
        except Exception as e:
            print(f"Error al procesar el archivo {archivo}: {e}")
        return registros_correctos, registros_incorrectos, registros_existentes, num_filas_procesadas


class DirectorioActivosRecreacionYDeporteDescargarArchivoView(View):

    def crear_archivo_excel(self, registros_incorrectos):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Obtener los nombres y verbose_name de los campos del modelo DirectorioActivosRecreacionYDeporte
        fields = DirectorioActivosRecreacionYDeporte._meta.get_fields()
        column_labels = [field.verbose_name for field in fields if field.name != 'id']
        column_names = [field.name for field in fields if field.name != 'id']

        # Escribir los encabezados de las columnas
        for i, campo in enumerate(column_labels):
            columna = i + 1
            worksheet.cell(row=1, column=columna, value=campo)

        # Obtener los datos del modelo DirectorioActivosRecreacionYDeporte
        datos = registros_incorrectos

        # Escribir los valores en las celdas correspondientes
        fila = 2
        for registro in datos:
            for i, campo in enumerate(column_names):
                if campo != 'id':  # Omitir la clave 'id'
                    columna = i + 1
                    valor = registro[campo]
                    worksheet.cell(row=fila, column=columna, value=valor)
            fila += 1

        # Set the column widths to auto-fit
        for column in worksheet.columns:
            max_length = 0
            column_name = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_name].width = adjusted_width

        # Create the response with the Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=gasto_derrama_registros_incorrectos.xls'

        # workbook.save(response)
        return workbook

    def post(self, request, *args, **kwargs):
        # Obtener los registros incorrectos del cuerpo de la petición
        registros_incorrectos = json.loads(request.body)

        # Crear y enviar el archivo de Excel con las filas incorrectas
        workbook = self.crear_archivo_excel(registros_incorrectos)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=gasto_derrama_registros_incorrectos.xlsx'
        workbook.save(response)
        return response
